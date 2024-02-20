from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from itertools import combinations
from datetime import datetime
import os
import logging
import sqlite3
from tkinter import messagebox as mb
import pandas as pd
from tabulate import tabulate
from collections import namedtuple, defaultdict
import win32com.client

import collection_func as cf
from rastr_model import RastrModel
from automation import Automation
from common import Common
from print_xl import PrintXL
# import report
log_g_s = logging.getLogger(f'__main__.{__name__}')


class CalcModel(Common):
    """
    Расчет нормативных возмущений.
    """

    def __init__(self, config: dict):
        """
        :param config: Задание и настройки программы
        """
        super(CalcModel, self).__init__()
        self.config = self.config | config
        RastrModel.config = config['Settings']

        RastrModel.overwrite_new_file = 'question'
        self.info_srs = None  # СРС
        self.comb_id = 1
        self.all_folder = False  # Не перебирать вложенные папки
        self.set_comb = None  # {количество отключений: контроль ДТН, 1:"ДДТН",2:"АДТН"}
        # self.auto_shunt = {}

        self.control_I = None
        self.control_U = None
        self.restore_only_state = True

        # DF для хранения токовых перегрузок и недопустимого снижения U
        self.srs_xl = pd.DataFrame()  # Перечень отключений из excel

        self.book_path: str = ''  # Путь к файлу excel.
        self.book_db: str = ''  # Путь к файлу db.
        self.pa = None  # Объект Automation
        self.task_full_name = ''  # Путь к файлу задания rg2.

        self.disable_df_vetv = pd.DataFrame()

        # {((ip, iq, np), disable_scheme, comb.repair_scheme):list((ip, iq, np), ...)}
        # отключаемая ветвь: (ветви из списка отключаемых ветвей, загрузка которых изменяется)
        self.disable_effect = defaultdict(list)

        self.num_pic, self.name_pic = list(Common.read_title(self.config['name_pic']))

        self.info_action = None
        RastrModel.all_rm = pd.DataFrame()
        self.all_comb = pd.DataFrame()
        self.all_actions = pd.DataFrame()  # действие оперативного персонала или ПА

        # Для хранения имен файлов с рисунками и имен рисунков
        self.all_pic = pd.DataFrame(dtype='str', columns=['pic_id',
                                                          'comb_id',
                                                          'active_id',
                                                          'Наименование файла',
                                                          'Наименование рисунка'])
        self.breach = {'i': pd.DataFrame(), 'high_u': pd.DataFrame(), 'low_u': pd.DataFrame()}
        # Для хранения токовой загрузки контролируемых элементов в пределах одной РМ a формате df и добавления в db.
        self.save_i_rm = None

    def run(self):
        """
        Запуск расчета нормативных возмущений (НВ) в РМ.
        """
        # test_run('calc')
        log_g_s.info('Запуск расчета нормативных возмущений (НВ) в расчетной модели (РМ).')
        if "*" in self.config["calc_folder"]:
            self.config["calc_folder"] = self.config["calc_folder"].replace('*', '')
            self.all_folder = True

        if not os.path.exists(self.config["calc_folder"]):
            raise ValueError(f'Не найден путь: {self.config["calc_folder"]}.')

        # папка для сохранения результатов
        self.config['folder_result_calc'] = self.config["calc_folder"] + r"\result"
        if os.path.isfile(self.config["calc_folder"]):
            self.config['folder_result_calc'] = os.path.dirname(self.config["calc_folder"]) + r"\result"
        if not os.path.exists(self.config['folder_result_calc']):
            os.mkdir(self.config['folder_result_calc'])  # создать папку result

        self.config['name_time'] = f"{self.config['folder_result_calc']}" \
                                   f"\\{datetime.now().strftime('%d-%m-%Y %H-%M-%S')}"

        if self.config['cb_disable_excel']:
            self.srs_xl = pd.read_excel(self.config['srs_XL_path'],
                                        sheet_name=self.config['srs_XL_sheets'])

            self.srs_xl = self.srs_xl[self.srs_xl['Статус'] != '-']
            self.srs_xl.drop(columns=['Примечание', 'Статус'], inplace=True)
            self.srs_xl.dropna(how='all', axis=0, inplace=True)
            # self.srs_xl.dropna(how='all', axis=1, inplace=True)
            # for col in self.srs_xl.columns:
            #     self.srs_xl[col] = self.srs_xl[col].astype(str).str.split('#').str[0]
            self.srs_xl.fillna(0, inplace=True)

        # Цикл, если несколько файлов задания.
        if self.config['CB_Import_Rg2'] and os.path.isdir(self.config["Import_file"]):
            task_files = os.listdir(self.config["Import_file"])
            task_files = list(filter(lambda x: x.endswith('.rg2'), task_files))
            for task_file in task_files:  # цикл по файлам '.rg2' в папке
                self.task_full_name = os.path.join(self.config["Import_file"], task_file)
                log_g_s.info(f'Текущий файл задания: {self.task_full_name}')
                self.run_calc_task()
        else:
            if self.config['CB_Import_Rg2']:
                self.task_full_name = self.config['Import_file']
                self.run_calc_task()
            else:
                self.run_calc_task()

        self.the_end()
        self.save_config(self.config, 'calc')
        mb.showinfo("Инфо", self.config['end_info'])

    @staticmethod
    def gen_comb_xl(rm, df: pd.DataFrame) -> pd.DataFrame:
        """
        Генератор комбинаций из XL
        :param rm:
        :param df:
        :return:  комбинацию comb_xl
        """
        for _, row in df.iterrows():
            comb_xl = pd.DataFrame(columns=['table',
                                            'index',
                                            'status_repair',
                                            'key',
                                            's_key',
                                            'repair_scheme',
                                            'disable_scheme',
                                            'double_repair_scheme',
                                            'double_repair_scheme_copy'])
            double_repair = True if row['Ключ рем.1'] and row['Ключ рем.2'] else False
            for key_type, scheme_xl_name in (('Ключ откл.', 'Схема при отключении'),
                                             ('Ключ рем.1', 'Ремонтная схема1'),
                                             ('Ключ рем.2', 'Ремонтная схема2')):
                key = row[key_type]
                if key:
                    key = str(key)
                    status_repair = False if key_type == 'Ключ откл.' else True
                    table, s_key = rm.recognize_key(key=key, back='tab s_key')
                    index = rm.index(table_name=table, key_int=s_key)
                    if table and index >= 0:
                        repair_scheme = False
                        disable_scheme = False
                        double_repair_scheme = False
                        double_repair_scheme_copy = False
                        # Если в колонке «Схема при отключении» или «Ремонтная схема» содержится «*», то значение поля
                        # дополняется из соответствующих полей disable_scheme, repair_scheme, double_repair_scheme РМ.
                        scheme_xl = row[scheme_xl_name]
                        add_scheme = []
                        if scheme_xl:
                            scheme_xl = scheme_xl.split('#')[0].replace(' ', '')
                            if '*' in scheme_xl:
                                scheme_xl = scheme_xl.replace('*', '')

                                if status_repair:
                                    add_scheme = rm.t_scheme[table]['repair_scheme'].get(s_key, False)
                                    if double_repair:
                                        double_repair_scheme_copy = \
                                            rm.t_scheme[table]['double_repair_scheme'].get(s_key, False)
                                else:
                                    add_scheme = rm.t_scheme[table]['disable_scheme'].get(s_key, False)
                            scheme_xl = cf.split_task_action(scheme_xl)
                            if add_scheme:
                                if scheme_xl:
                                    scheme_xl.append(add_scheme)
                                else:
                                    scheme_xl = add_scheme
                        if scheme_xl:
                            if status_repair:
                                repair_scheme = scheme_xl
                            else:
                                disable_scheme = scheme_xl

                        comb_xl.loc[len(comb_xl.index)] = [table,
                                                           index,
                                                           status_repair,
                                                           key,
                                                           s_key,
                                                           repair_scheme,
                                                           disable_scheme,
                                                           double_repair_scheme,
                                                           double_repair_scheme_copy]
                    else:
                        log_g_s.info(f'Задание комбинаций их XL: в РМ не найден ключ {key!r}')
                        log_g_s.info(tabulate(row, headers='keys', tablefmt='psql'))
                        continue
            if not len(comb_xl):
                continue
            if comb_xl['double_repair_scheme'].any():
                CalcModel.find_double_repair_scheme(comb_xl)
            yield comb_xl

    def run_calc_task(self):
        """
        Запуск расчета с текущим файлом импорта задания или без него.
        """
        xlApp = None

        if os.path.isdir(self.config["calc_folder"]):
            # папка с вложенными папками
            if self.all_folder:
                for address, dir_, file_ in os.walk(self.config["calc_folder"]):
                    self.cycle_rm(folder_calc=address)
            # папка без вложенных папок
            else:
                self.cycle_rm(folder_calc=self.config["calc_folder"])
        # один файл
        elif os.path.isfile(self.config["calc_folder"]):
            rm = RastrModel(self.config["calc_folder"])
            if not rm.code_name_rg2:
                raise ValueError(f'Имя файла {self.config["calc_folder"]!r} не подходит.')
            self.calc_file(rm=rm)

        # Сохранить таблицы в SQL.
        con = sqlite3.connect(self.book_db)
        for key in self.breach:
            self.breach[key].to_sql(key, con, if_exists="replace")
        name_df = {'all_rm': RastrModel.all_rm,
                   'all_comb': self.all_comb,
                   'all_actions': self.all_actions,
                   'all_pic': self.all_pic}
        for key in name_df:
            name_df[key].to_sql(key, con, if_exists="replace")
        save_i_for_xl = None
        if self.config['cb_save_i']:
            save_i_for_xl = pd.read_sql_query("""
            SELECT s_key, "Контролируемые элементы", "Год", "Сезон макс/мин", 
            "Темп.(°C)", "Кол. откл. эл.", 
            count(*) AS "Кол.СРС", 
            "Наименование СРС", 
            max(i_max) AS "Iрасч.,A", 
            i_dop_r AS "Iддтн,А", 
            i_zag AS "Iзагр. ддтн,%", 
            i_dop_r_av AS "Iадтн,А", 
            i_zag_av AS "Iзагр. адтн,%"
            FROM (
            SELECT *
            FROM save_i AS si
               INNER JOIN all_actions AS aa
                  ON si.comb_id = aa.comb_id AND si.active_id = aa.active_id
               INNER JOIN all_comb AS ac
                  ON ac.comb_id = aa.comb_id
               INNER JOIN all_rm AS ar
                  ON ar.rm_id = ac.rm_id
            )
            GROUP BY s_key, "Год", "Сезон макс/мин", "Темп.(°C)", "Кол. откл. эл.";
            """, con)

        con.commit()
        con.close()

        log_g_s.debug(f'Запись параметров режима в excel.')
        full_breach = {}
        for key in self.breach:
            if len(self.breach[key]):
                full_breach[key] = (RastrModel.all_rm.merge(self.all_comb)
                                    .merge(self.all_actions)
                                    .merge(self.breach[key]))

                for col in ["Отключение", "Ремонт 1", "Ремонт 2", "Доп. имя"]:
                    for col_df in full_breach[key].columns:
                        if col in col_df:
                            full_breach[key].fillna(value={col_df: 0}, inplace=True)
                            full_breach[key].loc[full_breach[key][col_df] == 0, col_df] = '-'
                # https://www.geeksforgeeks.org/how-to-write-pandas-dataframes-to-multiple-excel-sheets/
                mode = 'a' if os.path.exists(self.book_path) else 'w'
                with pd.ExcelWriter(path=self.book_path, mode=mode) as writer:
                    full_breach[key].to_excel(excel_writer=writer,
                                              float_format="%.2f",
                                              index=False,
                                              freeze_panes=(1, 1),
                                              sheet_name=key)
        crash = self.all_actions[self.all_actions.alive == 0]
        if len(crash):
            full_breach['crash'] = (RastrModel.all_rm.merge(self.all_comb)
                                    .merge(self.all_actions[self.all_actions.alive == 0]))
            mode = 'a' if os.path.exists(self.book_path) else 'w'
            with (pd.ExcelWriter(path=self.book_path, mode=mode) as writer):
                full_breach['crash'].to_excel(excel_writer=writer,
                                              float_format="%.2f",
                                              index=False,
                                              freeze_panes=(1, 1),
                                              sheet_name='crash')
            (RastrModel.all_rm.merge(self.all_comb).merge(self.all_actions[self.all_actions.alive == 0]))

        if self.config['cb_save_i']:
            mode = 'a' if os.path.exists(self.book_path) else 'w'
            with (pd.ExcelWriter(path=self.book_path, mode=mode) as writer):
                save_i_for_xl.to_excel(excel_writer=writer,
                                       float_format="%.2f",
                                       index=False,
                                       freeze_panes=(1, 1),
                                       sheet_name='Макс.ток')
        # todo Сохранить в Excel таблицу перегрузки.
        sheet_name_pic = 'Рисунки'
        if len(self.all_pic):
            with pd.ExcelWriter(path=self.book_path,
                                mode='a' if os.path.exists(self.book_path) else 'w') as writer:
                self.all_pic[['Наименование файла', 'Наименование рисунка']].to_excel(excel_writer=writer,
                                                                                      startrow=1,
                                                                                      index=False,
                                                                                      freeze_panes=(5, 1),
                                                                                      sheet_name=sheet_name_pic)
            book = load_workbook(self.book_path)
            sheet_pic = book[sheet_name_pic]

            sheet_pic.insert_rows(1, amount=3)

            sheet_pic['A1'] = 'Формат листа (3 - А3, 4 - А4):'
            sheet_pic['A2'] = 'Ориентация(1 - книжная, 0 - альбомная):'
            sheet_pic['A3'] = 'Имя папки с файлами rg2:'
            sheet_pic['B1'] = 3
            sheet_pic['B2'] = 0
            sheet_pic['B3'] = self.config['folder_result_calc']
            thins = Side(border_style="thin", color="000000")
            for col in ['A', 'B']:
                sheet_pic.column_dimensions[col].width = 100
                for r in ['1', '2', '3']:
                    sheet_pic[col + r].alignment = Alignment(horizontal='left')
                    sheet_pic[col + r].border = Border(thins, thins, thins, thins)
                    sheet_pic[col + r].fill = PatternFill(fill_type='solid', fgColor="00B1E76E")
            PrintXL.create_table(sheet=sheet_pic,
                                 sheet_name=sheet_name_pic,
                                 point_start='A5')
            book.save(self.book_path)

            # Сохранить макрос rbs.
            rbs = ''
            with open('help\Сделать рисунки в word.rbs') as f:
                for readline in f.readlines():
                    rbs += readline
            rbs = rbs.replace('папка с файлами', self.book_path)
            f2 = open(self.book_path.rsplit('.', 1)[0] + '.rbs', 'w')
            f2.write(rbs)
            f2.close()

        # Сводная.
        if len(full_breach):
            log_g_s.info(f'Формирование сводных таблиц ({self.book_path}).')
            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.ScreenUpdating = False  # Обновление экрана
            try:
                book = xlApp.Workbooks.Open(self.book_path)
            except Exception:
                raise Exception(f'Ошибка при открытии файла {self.book_path=}')

            for key in full_breach:
                try:
                    sheet = book.sheets[key]
                except Exception:
                    raise Exception(f'Не найден лист: {key}')
                # Создать объект таблица из всего диапазона листа.

                tabl_overload = sheet.ListObjects.Add(SourceType=1, Source=sheet.Range(sheet.UsedRange.address))
                tabl_overload.Name = f"Таблица_{key}"
                pt_cache = book.PivotCaches().add(1, tabl_overload)  # Создать КЭШ xlDatabase, ListObjects

                task_pivot = namedtuple('task_pivot',
                                        ['sheet_name', 'pivot_table_name', 'data_field'])
                task_pivot_cur = None
                if key == "i":
                    task_pivot_cur = task_pivot('Сводная_I', "Свод_I",
                                                dict(i_max="Iрасч.,A",
                                                     i_dop_r="Iддтн,A",
                                                     i_zag="Iзагр. ддтн,%",
                                                     i_dop_r_av="Iадтн,A",
                                                     i_zag_av="Iзагр. адтн,%"))
                elif key == "low_u":
                    task_pivot_cur = task_pivot('Сводная_Umin', "Свод_Umin",
                                                dict(vras="Uр, кВ",
                                                     umin="МДН, кВ",
                                                     otv_min="Uр, % от МДН",
                                                     umin_av="АДН,кВ",
                                                     otv_min_av="Uр, % от АДН"))
                elif key == "high_u":
                    task_pivot_cur = task_pivot('Сводная_Umax', "Свод_Umax",
                                                dict(vras="Uр, кВ",
                                                     umax="Uнр, кВ",
                                                     otv_max="Uр, % от Uнр"))
                elif key == "crash":
                    task_pivot_cur = task_pivot('Сводная_не_сходятся', "Свод_crash",
                                                dict(alive='Режим не сошелся'))

                RowFields = [col for col in ["Контролируемые элементы", "Отключение", "Ремонт 1", "Ремонт 2"]
                             if col in full_breach[key].columns]

                ColumnFields = (["Год", "Сезон макс/мин"] +
                                [col for col in full_breach[key].columns if "Доп. имя" in col])

                sheet_pivot = book.Sheets.Add()
                sheet_pivot.Name = task_pivot_cur.sheet_name

                pt = pt_cache.CreatePivotTable(TableDestination=task_pivot_cur.sheet_name + "!R1C1",
                                               TableName=task_pivot_cur.pivot_table_name)
                pt.ManualUpdate = True  # True не обновить сводную
                pt.AddFields(RowFields=RowFields,
                             ColumnFields=ColumnFields,
                             PageFields=["Имя файла", 'Кол. откл. эл.', 'End', 'Наименование СРС', 'Контроль ДТН',
                                         'Темп.(°C)'],
                             AddToTable=False)
                for field_df, field_pt in task_pivot_cur.data_field.items():
                    pt.AddDataField(Field=pt.PivotFields(field_df),
                                    Caption=field_pt,
                                    Function=-4136)  # xlMax -4136 xlSum -4157
                    pt.PivotFields(field_pt).NumberFormat = "0"

                if len(task_pivot_cur.data_field) > 1:
                    pt.PivotFields("Контролируемые элементы").ShowDetail = True  # группировка
                pt.RowAxisLayout(1)  # 1 xlTabularRow показывать в табличной форме!!!!
                if len(task_pivot_cur.data_field) > 1:
                    pt.DataPivotField.Orientation = 1  # xlRowField = 1 "Значения" в столбцах или строках xlColumnField
                pt.RowGrand = False  # Удалить строку общих итогов
                pt.ColumnGrand = False  # Удалить столбец общих итогов
                pt.MergeLabels = True  # Объединять одинаковые ячейки
                pt.HasAutoFormat = False  # Не обновлять ширину при обновлении
                pt.NullString = "--"  # Заменять пустые ячейки
                pt.PreserveFormatting = False  # Сохранять формат ячеек при обновлении
                pt.ShowDrillIndicators = False  # Показывать кнопки свертывания
                for row in RowFields + ColumnFields:
                    pt.PivotFields(row).Subtotals = [False, False, False, False, False, False, False, False, False,
                                                     False, False, False]  # промежуточные итоги и фильтры
                if len(task_pivot_cur.data_field) > 1:
                    field = list(task_pivot_cur.data_field)[2]
                    pt.PivotFields(field).Orientation = 3  # xlPageField = 3
                    pt.PivotFields(field).CurrentPage = "(All)"  #
                pt.TableStyle2 = ""  # стиль
                pt.ColumnRange.ColumnWidth = 10  # ширина строк
                pt.RowRange.ColumnWidth = 20
                pt.DataBodyRange.HorizontalAlignment = -4108  # xlCenter = -4108
                pt.TableRange1.WrapText = True  # перенос текста в ячейке
                for i in range(7, 13):
                    pt.TableRange1.Borders(i).LineStyle = 1  # лево
                # Условное форматирование
                if key != 'crash':
                    for i in range(3, len(task_pivot_cur.data_field) + 1, 2):
                        dpz = pt.DataBodyRange.Rows(i).Cells(1)
                        dpz.FormatConditions.AddColorScale(2)  # ColorScaleType:=2
                        dpz.FormatConditions(dpz.FormatConditions.count).SetFirstPriority()
                        dpz.FormatConditions(1).ColorScaleCriteria(1).Type = 0  # xlConditionValueNumber = 0
                        if list(task_pivot_cur.data_field)[2] == 'i_zag':
                            dpz.FormatConditions(1).ColorScaleCriteria(1).Value = 100
                        else:
                            dpz.FormatConditions(1).ColorScaleCriteria(1).Value = 0
                        dpz.FormatConditions(1).ColorScaleCriteria(1).FormatColor.ThemeColor = 1  # xlThemeColorDark1=1
                        dpz.FormatConditions(1).ColorScaleCriteria(1).FormatColor.TintAndShade = 0
                        dpz.FormatConditions(1).ColorScaleCriteria(2).Type = 2  # xlConditionValueHighestValue = 2
                        dpz.FormatConditions(1).ColorScaleCriteria(2).FormatColor.ThemeColor = 3 + i  # номер темы
                        dpz.FormatConditions(1).ColorScaleCriteria(2).FormatColor.TintAndShade = -0.249977111117893
                        dpz.FormatConditions(1).ScopeType = 2  # xlDataFieldScope = 2 применить ко всем значениям поля
                pt.ManualUpdate = False  # обновить сводную
            book.Save()
            book.Close()
        else:
            log_g_s.info('Отклонений параметров режима от допустимых значений не выявлено.')

        # Вставить таблицы К-О в word.
        if self.config['cb_tab_KO']:
            log_g_s.info('Вставить таблицы К-О в word.')

            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.Visible = False
            book = xlApp.Workbooks.Open(self.book_path)

            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            word.ScreenUpdating = False
            doc = word.Documents.Add()  # doc = word.Documents.Open(r"I:\file.docx")

            doc.PageSetup.PageWidth = 29.7 * 28.35  # CentimetersToPoints( format_list_i (2) ) 1 см = 28,35
            doc.PageSetup.PageHeight = 42.0 * 28.35  # CentimetersToPoints( format_list_i (1) )
            doc.PageSetup.Orientation = 1  # 1 книжная или 0 альбомная

            cursor = word.Selection
            cursor.Font.Size = 12
            cursor.Font.Name = "Times New Roman"
            cursor.EndKey(Unit=6)  # перейти в конец текста

            for i in range(1, book.Worksheets.Count + 1):
                if book.Worksheets(i).name[:1].isnumeric():
                    sheet = book.Worksheets(i)
                    cursor.TypeText(Text=sheet.Cells(1, 1).value)
                    cursor.TypeParagraph()

                    sheet.Range(sheet.UsedRange.address.replace('$', '').replace('A1', 'A2')).Copy()
                    # cursor.PasteExcelTable(LinkedToExcel=False, WordFormatting=False, RTF=False)
                    cursor.PasteAndFormat(Type=13)  # 13 Вставить в виде рисунка.
                    cursor.InsertBreak(Type=0)
                    # разрыв:7 страницы с новой строки, 0-в той же строке,1 и 8 колонки,
                    # 2-5 раздела со след стр,6 и 9-11 перенос на новую стр

            word.ScreenUpdating = True
            doc.SaveAs2(FileName=self.config['name_time'] + ' таблицы К-О.docx')  # FileFormat=16 .docx
            doc.Close()

        if xlApp:
            xlApp.Visible = True
            xlApp.ScreenUpdating = True  # обновление экрана

    def cycle_rm(self, folder_calc: str):
        """
        Цикл по файлам.
        :param folder_calc:
        """

        files_calc = os.listdir(folder_calc)  # список всех файлов в папке
        rm_files = list(filter(lambda x: x.endswith('.rg2'), files_calc))

        for rastr_file in rm_files:  # цикл по файлам '.rg2' в папке
            log_g_s.info("\n\n")
            if self.config["Filter_file"] and self.file_count == self.config["file_count_max"]:
                break  # Если включен фильтр файлов проверяем количество расчетных файлов.
            full_name = os.path.join(folder_calc, rastr_file)
            rm = RastrModel(full_name)
            # если включен фильтр файлов и имя стандартизовано
            if not rm.code_name_rg2:
                log_g_s.info(f'Имя файл {full_name} не распознано.')
                continue
            if self.config["Filter_file"]:
                if not rm.test_name(condition=self.config["calc_criterion"],
                                    info=f'Имя файла {full_name} не подходит.'):
                    continue  # пропускаем, если не соответствует фильтру
            self.calc_file(rm)

    def calc_file(self, rm):
        """
        Рассчитать РМ.
        """

        if self.config['cb_save_i']:
            self.save_i_rm = pd.DataFrame()

        self.set_comb = {}  # {количество отключений: контроль ДТН, 1:"ДДТН",2:"АДТН"}
        self.file_count += 1
        self.book_path = self.config['name_time'] + ' результаты расчетов.xlsx'
        self.book_db = self.config['name_time'] + ' данные.db'
        rm.load()
        rm.rastr.CalcIdop(rm.info_file["Темп.(°C)"], 0.0, "")
        log_g_s.info(f'Выполнен расчет ДТН для температуры: {rm.info_file["Темп.(°C)"]} °C.')

        if self.config['cor_rm']['add']:
            rm.cor_rm_from_txt(self.config['cor_rm']['txt'])

        # Импорт из РМ c ИД.
        if self.task_full_name:  # :task_full_name: полный путь к текущему файлу задания
            # "таблица: node, vetv; тип: 2; поле: disable_scheme, automation; выборка: sel"
            for row in self.config['txt_Import_Rg2'].split('\n'):
                row = row.replace(' ', '').split('#')[0]  # удалить текст после '#'
                if row:
                    rm.txt_import_rm(type_import='файл', description=f'({self.task_full_name});{row}')

        # Подготовка.
        rm.voltage_fix_frame()
        # if self.config['CalcSetWindow']['skrm']:
        #     self.auto_shunt = rm.auto_shunt_rec(selection='')

        # Добавить поле index в таблицы.
        rm.fill_field_index('vetv,node,Generator')

        # Поля для сортировки ветвей и др.
        rm.add_fields_in_table(name_tables='vetv', fields='temp,temp1', type_fields=1)

        # Поля для контроля напряжений
        rm.add_fields_in_table(name_tables='node', fields='umin_av', type_fields=1)
        rm.add_fields_in_table(name_tables='node', fields='otv_min', type_fields=1,
                               prop=((5, 'if(sta=0) (-vras+umin)/umin*100:0'),),
                               replace=True)
        rm.add_fields_in_table(name_tables='node', fields='otv_min_av', type_fields=1,
                               prop=((5, 'if(sta=0) (-vras+umin_av)/umin_av*100:0'),),
                               replace=True)
        rm.add_fields_in_table(name_tables='node', fields='otv_max', type_fields=1,
                               prop=((5, 'if(sta=0) (vras-umax)/umax*100:0'),))
        # Поля для загрузки ветвей
        rm.add_fields_in_table(name_tables='vetv', fields='i_zag_av', type_fields=1,
                               prop=((5, 'if(ktr!=0) zag_it_av:zag_i_av'), (12, 1000),))

        # Поля для автоматики, что бы не было ошибок
        rm.add_fields_in_table(name_tables='vetv,node,Generator',
                               fields='repair_scheme,double_repair_scheme,disable_scheme,automation,dname',
                               type_fields=2)
        # Поля с ключами таблиц
        rm.add_fields_in_table(name_tables='vetv', fields='key', type_fields=2,
                               prop=((5, '"ip="+str(ip)+"&iq="+str(iq)+"&np="+str(np)'),))
        rm.add_fields_in_table(name_tables='node', fields='key', type_fields=2,
                               prop=((5, '"ny="+str(ny)'),))
        rm.add_fields_in_table(name_tables='Generator', fields='key', type_fields=2,
                               prop=((5, '"Num="+str(Num)'),))

        rm.add_fields_in_table(name_tables='vetv,node', fields='all_control', type_fields=3)
        # В поле all_disable складываем элементы авто отмеченные и отмеченные в поле comb_field
        rm.add_fields_in_table(name_tables='vetv,node,Generator', fields='all_disable', type_fields=3)

        if self.config['CalcSetWindow']['pa']:
            self.pa = Automation(rm)
            if not self.pa.exist:
                self.config['CalcSetWindow']['pa'] = False

        # Сохранить текущее состояние РМ
        rm.save_value_fields()

        # Контролируемые элементы сети.
        if self.config['cb_control']:
            log_g_s.debug('Определение контролируемых элементов сети.')
            # all_control для отметки всех контролируемых узлов и ветвей (авто и field)

            if self.config['cb_control_field']:
                control_field = self.config['control_field'].replace(' ', '')
                log_g_s.debug(f'Отмеченный в поле [{control_field}] элементы добавлены в контролируемые.')
                if control_field:
                    for table_name in ['vetv', 'node']:
                        rm.group_cor(tabl=table_name,
                                     param="all_control",
                                     selection=control_field,
                                     formula='1')

            if self.config['cb_control_sel']:
                control_sel = self.config['control_sel'].replace(' ', '')
                log_g_s.debug(f'Элементы по выборке {control_sel}  добавлены в контролируемые.')
                if control_sel:
                    table = rm.rastr.tables('node')
                    table.setsel(control_sel)
                    if table.count:
                        ny_sel = [x[0] for x in table.writesafearray("ny", "000")]
                        sel_v = set()
                        for ny in ny_sel:
                            for ip, iq, np_ in rm.ny_join_vetv[ny]:
                                sel_v.add((ip, iq, np_, 1))
                        rm.rastr.tables('vetv').ReadSafeArray(2, 'ip,iq,np,all_control', list(sel_v))
                        rm.rastr.tables('node').ReadSafeArray(2, 'ny,all_control', [(ny, 1) for ny in ny_sel])
                    else:
                        raise ValueError(f'По выборке {control_sel} не найдены узлы в РМ.')
                else:  # Контролировать все узлы и ветви.
                    rm.rastr.Tables("node").cols.item("all_control").Calc("1")
                    rm.rastr.Tables("vetv").cols.item("all_control").Calc("1")

            # all_control_groupid для отметки всех контролируемых ветвей и ветвей с теми же groupid
            if not self.config['cb_tab_KO']:
                log_g_s.info('Добавление в контролируемые элементы ветвей по groupid.')
                table = rm.rastr.tables('vetv')
                table.setsel('all_control & groupid>0')
                for gr in set(table.writesafearray('groupid', "000")):
                    rm.group_cor(tabl='vetv',
                                 param="all_control",
                                 selection=f"groupid={gr[0]}",
                                 formula=1)

            if self.config['cb_tab_KO']:
                log_g_s.debug('Инициализация таблицы "контролируемые - отключаемые" элементы.')
                rm.rastr.tables('vetv').cols.item("temp").calc('ip.uhom')
                rm.rastr.tables('vetv').cols.item("temp1").calc('iq.uhom')
                self.control_I = rm.df_from_table(table_name='vetv',
                                                  fields='index,dname,name,temp,temp1,i_dop_r,i_dop_r_av,groupid'
                                                         ',key,tip',  # ip, iq, np
                                                  setsel="all_control")
                dname_list = []
                for dname, name in zip(list(self.control_I.dname), list(self.control_I.name)):
                    if dname.strip():
                        dname_list.append(dname)
                    else:
                        dname_list.append(name)
                self.control_I.dname = dname_list
                self.control_I.drop(['name'], axis=1, inplace=True)

                if len(self.control_I):
                    # Сортировка
                    self.control_I['uhom'] = (self.control_I[['temp', 'temp1']].max(axis=1) * 10000 +
                                              self.control_I[['temp', 'temp1']].min(axis=1))
                    self.control_I.sort_values(by=['tip', 'uhom', 'dname'],  # столбцы сортировки
                                               ascending=(False, False, True),  # обратный порядок
                                               inplace=True)  # изменить df
                    self.control_I.drop(['temp', 'temp1', 'uhom', 'tip'], axis=1, inplace=True)
                    self.control_I['i_dop_r'] = self.control_I['i_dop_r'].round(0).astype(int)
                    self.control_I['i_dop_r_av'] = self.control_I['i_dop_r_av'].round(0).astype(int)
                    self.control_I.rename(columns={'i_dop_r': 'ДДТН, А',
                                                   'i_dop_r_av': 'АДТН, А',
                                                   'dname': 'Контролируемый элемент'}, inplace=True)
                    self.control_I.set_index('index', inplace=True)
                    self.control_I = self.control_I.T
                    self.control_I.index = pd.MultiIndex.from_product([['-'], ['-'], self.control_I.index])

                self.control_U = rm.df_from_table(table_name='node',
                                                  fields='index,dname,name,umin,umin_av,uhom',  # ny,umax
                                                  setsel="all_control")

                dname_list = []
                for dname, name in zip(list(self.control_U.dname), list(self.control_U.name)):
                    if dname.strip():
                        dname_list.append(dname)
                    else:
                        dname_list.append(name)
                self.control_U.dname = dname_list
                self.control_U.drop(['name'], axis=1, inplace=True)

                if len(self.control_U):
                    self.control_U.sort_values(by=['uhom', 'dname'],  # столбцы сортировки
                                               ascending=(False, True),  # обратный порядок
                                               inplace=True)  # изменить df
                    self.control_U.drop(['uhom'], axis=1, inplace=True)
                    self.control_U['umin'] = self.control_U['umin'].round(1)
                    self.control_U['umin_av'] = self.control_U['umin_av'].round(1)
                    self.control_U.rename(columns={'umin': 'МДН, кВ',
                                                   'umin_av': 'АДН, кВ',
                                                   'dname': 'Контролируемый элемент'}, inplace=True)
                    self.control_U.set_index('index', inplace=True)
                    self.control_U = self.control_U.T
                    self.control_U.index = pd.MultiIndex.from_product([['-'], ['-'], self.control_U.index])

        # Нормальная схема сети
        self.info_srs = dict()  # СРС
        self.info_srs['Наименование СРС'] = 'Нормальная схема сети.'
        self.info_srs["Наименование СРС без()"] = 'Нормальная схема сети'
        self.info_srs['comb_id'] = self.comb_id
        self.info_srs['Кол. откл. эл.'] = 0
        self.info_srs['Контроль ДТН'] = 'ДДТН'
        self.info_srs['rm_id'] = RastrModel.rm_id
        log_g_s.info(f"Сочетание {self.comb_id}: {self.info_srs['Наименование СРС']}")
        self.do_action(rm)

        # Отключаемые элементы сети.
        if self.config['cb_disable_comb']:
            # self.set_comb[0] = 'ДДТН'
            # Выбор количества одновременно отключаемых элементов
            # н-1
            if self.config['SRS']['n-1']:
                self.set_comb[1] = 'ДДТН'
            # н-2
            if self.config['CalcSetWindow']['gost']:
                if self.config['SRS']['n-2_abv'] and rm.gost_abv:
                    self.set_comb[2] = 'AДТН'
                if self.config['SRS']['n-2_gd'] and rm.gost_gd:
                    self.set_comb[2] = 'ДДТН'
            else:
                if self.config['SRS']['n-2_abv'] or self.config['SRS']['n-2_gd']:
                    self.set_comb[2] = 'ДДТН'
            # н-3
            if self.config['SRS']['n-3']:
                if self.config['CalcSetWindow']['gost']:
                    if rm.gost_gd:
                        self.set_comb[3] = 'АДТН'
                else:
                    self.set_comb[3] = 'ДДТН'
            log_g_s.info(f'Расчетные СРС: {self.set_comb}.')

            if self.config['cb_auto_disable']:
                # Выбор отключаемых элементов автоматически из выборки в таблице узлы
                # Отметка в таблицах ветви и узлы нужного поля
                rm.network_analysis(field='all_disable',
                                    selection_node_for_disable=self.config['auto_disable_choice'])
            else:
                rm.network_analysis(disable_on=False)

            # Выбор отключаемых элементов из отмеченных в поле comb_field
            if self.config['cb_comb_field']:
                # Добавит поле отметки отключений, если их нет в какой-то таблице
                rm.add_fields_in_table(name_tables='vetv,node,Generator', fields=self.config['comb_field'],
                                       type_fields=3)
                for table_name in ['vetv', 'node', 'Generator']:
                    rm.group_cor(tabl=table_name,
                                 param="all_disable",
                                 selection=self.config['comb_field'],
                                 formula='1')

            # Создать df отключаемых узлов и ветвей и генераторов. Сортировка.
            # Генераторы
            disable_df_gen = rm.df_from_table(table_name='Generator',
                                              fields='index,key,Num',  # ,Num,NodeState,Node
                                              setsel="all_disable")
            disable_df_gen['table'] = 'Generator'
            disable_df_gen.rename(columns={'Num': 's_key'}, inplace=True)
            # Узлы
            disable_df_node = rm.df_from_table(table_name='node',
                                               fields='index,name,uhom,key,ny',
                                               setsel="all_disable")
            # disable_df_node.index = self.disable_df_node['index']

            disable_df_node['table'] = 'node'
            disable_df_node.sort_values(by=['uhom', 'name'],  # столбцы сортировки
                                        ascending=(False, True),  # обратный порядок
                                        inplace=True)  # изменить df
            disable_df_node.drop(['name'], axis=1, inplace=True)
            disable_df_node.rename(columns={'ny': 's_key'}, inplace=True)
            # Ветви
            self.disable_df_vetv = rm.df_from_table(table_name='vetv',
                                                    fields='index,name,key,temp,temp1,tip,ip,iq,np,i_zag',
                                                    setsel="all_disable")
            self.disable_df_vetv['table'] = 'vetv'
            self.disable_df_vetv['uhom'] = (self.disable_df_vetv[['temp', 'temp1']].max(axis=1) * 10000 +
                                            self.disable_df_vetv[['temp', 'temp1']].min(axis=1))
            self.disable_df_vetv.sort_values(by=['tip', 'uhom', 'name'],  # столбцы сортировки
                                             ascending=(False, False, True),  # обратный порядок
                                             inplace=True)  # изменить df
            self.disable_df_vetv['s_key'] = None
            for i in self.disable_df_vetv.index:
                self.disable_df_vetv.at[i, 's_key'] = (self.disable_df_vetv.at[i, 'ip'],
                                                       self.disable_df_vetv.at[i, 'iq'],
                                                       self.disable_df_vetv.at[i, 'np'],)

            self.disable_df_vetv.drop(['temp', 'temp1', 'tip', 'name', 'ip', 'iq', 'np'], axis=1, inplace=True)

            log_g_s.info(f'Количество отключаемых элементов сети:'
                         f' ветвей - {len(self.disable_df_vetv.axes[0])},'
                         f' узлов - {len(disable_df_node.axes[0])},'
                         f' генераторов - {len(disable_df_gen.axes[0])}.')

            disable_df_all = pd.concat([self.disable_df_vetv.drop(['i_zag'], axis=1),
                                        disable_df_node,
                                        disable_df_gen])
            # Фильтр комбинаций
            if not self.config['SRS']['n-1'] or not (self.config['SRS']['n-2_abv']
                                                     or self.config['SRS']['n-2_gd']
                                                     or self.config['SRS']['n-3']):
                self.config['filter_comb'] = False

            if self.config['filter_comb']:
                self.disable_df_vetv.set_index('index', inplace=True)
                self.disable_df_vetv.drop(['table', 'uhom'], axis=1, inplace=True)

            # Цикл по всем возможным сочетаниям отключений
            for n_, self.info_srs['Контроль ДТН'] in self.set_comb.items():  # Цикл н-1 н-2 н-3.
                if n_ > len(disable_df_all):
                    break
                log_g_s.info(f"Количество отключаемых элементов в комбинации: {n_} ({self.info_srs['Контроль ДТН']}).")
                if n_ == 1:
                    disable_all = disable_df_all.copy()
                else:
                    disable_all = \
                        disable_df_all[(disable_df_all['uhom'] > 300) | (disable_df_all['table'] != 'node')]
                disable_all.drop(['uhom'], axis=1, inplace=True)
                name_columns = list(disable_all.columns)
                disable_all = tuple(disable_all.itertuples(index=False, name=None))

                for comb in combinations(disable_all, r=n_):  # Цикл по комбинациям.
                    log_g_s.debug(f'Комбинация элементов {comb}')
                    comb_df = pd.DataFrame(data=comb, columns=name_columns)
                    unique_set_actions = []

                    comb_df['repair_scheme'] = False
                    comb_df['disable_scheme'] = False
                    comb_df['double_repair_scheme'] = False
                    for index in comb_df.index:
                        t = comb_df.loc[index, 'table']
                        k = comb_df.loc[index, 's_key']
                        for nm in ('repair_scheme', 'disable_scheme', 'double_repair_scheme'):
                            value = rm.t_scheme[t][nm].get(k, False)
                            if value:
                                comb_df.at[index, nm] = 1  # долбаный глюк at
                                comb_df.at[index, nm] = value

                    # Если нет дополнительных изменений сети, то всего 1 сочетание.
                    if not comb_df[['disable_scheme', 'repair_scheme', 'double_repair_scheme']].any().any():
                        comb_df['status_repair'] = True
                        comb_df.loc[0, 'status_repair'] = False
                        self.calc_comb(rm, comb_df)
                        continue
                    comb_df['double_repair_scheme_copy'] = comb_df['double_repair_scheme']
                    # Цикл по всем возможным комбинациям внутри сочетания, вызванные ремонтами и отключениями.
                    # Под i понимаем номер отключаемого элемента, остальные в ремонте.
                    # Если -1, то ремонт всех элементов.

                    i_min = 0 if len(comb_df) == 3 else -1
                    for i in range(n_ - 1, i_min - 1, -1):  # От последнего до первого элемента или -1.

                        # Если в ремонте 2 элемента.
                        double_repair = True if (n_ == 2 and i == -1) or (n_ == 3) else False
                        if self.info_srs['Контроль ДТН'] == "AДТН" and double_repair and n_ == 2:
                            continue  # Не расчетный случай по ГОСТ.

                        comb_df['status_repair'] = True  # Истина, если элемент в ремонте. Ложь отключен.
                        if i != -1:
                            comb_df.loc[i, 'status_repair'] = False

                        comb_df['double_repair_scheme'] = False
                        double_repair_scheme = []
                        if double_repair:
                            double_repair_scheme = self.find_double_repair_scheme(comb_df)

                        # Суммировать текущий набор изменений сети в set и проверить на уникальность.
                        set_actions = set()
                        for _, row in comb_df.iterrows():
                            if row['status_repair']:
                                if double_repair_scheme:
                                    set_actions.add(tuple(double_repair_scheme))
                                else:
                                    if row['repair_scheme']:
                                        set_actions.add(tuple(row['repair_scheme']))
                            else:
                                if row['disable_scheme']:
                                    set_actions.add(tuple(row['disable_scheme']))

                        if set_actions not in unique_set_actions:
                            unique_set_actions.append(set_actions)
                            self.calc_comb(rm, comb_df)

        # Отключаемые элементы сети по excel.
        if self.config['cb_disable_excel']:
            if self.srs_xl.empty:
                raise ValueError(f'Таблица отключений из xl отсутствует.')
            # self.srs_xl.fillna(0, inplace=True)
            comb_xl = self.gen_comb_xl(rm, self.srs_xl)
            for comb in comb_xl:
                self.info_srs['Контроль ДТН'] = 'ДДТН'
                if self.config['CalcSetWindow']['gost']:
                    if comb.shape[0] == 3 or (comb.shape[0] == 2 and rm.gost_abv):
                        self.info_srs['Контроль ДТН'] = 'АДТН'
                    if rm.gost_abv and (comb.shape[0] == 3 or (comb.shape[0] == 2 and all(comb['status_repair']))):
                        log_g_s.info(f'Сочетание отклонено по ГОСТ: ')
                        log_g_s.info(tabulate(comb, headers='keys', tablefmt='psql'))
                        continue
                self.calc_comb(rm, comb, source='xl')

        # Доработка перечня перегрузок РМ
        comb_min = min(self.all_comb[self.all_comb.rm_id == RastrModel.rm_id].comb_id.to_list())

        for key in self.breach:
            if len(self.breach[key]):
                sel = self.breach[key].comb_id >= comb_min
                if len(self.breach[key][sel]):
                    tabl = rm.vetv_name if key == 'i' else rm.node_name
                    self.breach[key].loc[sel, 'Контролируемые элементы'] = \
                        self.breach[key].loc[sel, ['s_key']] \
                            .merge(tabl, how='left') \
                            .set_index(self.breach[key].loc[sel].index)['Контролируемые элементы']
        if self.config['cb_save_i']:
            self.save_i_rm = self.save_i_rm.merge(rm.vetv_name, how='left')
            con = sqlite3.connect(self.book_db)
            self.save_i_rm.to_sql('save_i', con, if_exists="append")
            con.commit()
            con.close()
        # Вывод таблиц К-О в excel
        if self.config['cb_tab_KO'] and (len(self.control_I) or len(self.control_U)):
            name_sheet = f'{self.file_count}_{rm.info_file["Имя файла"]}'.replace('[', '').replace(']', '')[:28]
            control_df_dict = {}
            if len(self.control_I):
                control_df_dict[name_sheet + '{I}'] = self.control_I
                self.control_I = None
            if len(self.control_U):
                control_df_dict[name_sheet + '{U}'] = self.control_U
                self.control_U = None
            # https://www.geeksforgeeks.org/how-to-write-pandas-dataframes-to-multiple-excel-sheets/

            # mode = 'a' if os.path.exists(self.book_path) else 'w'
            if not os.path.exists(self.book_path):
                Workbook().save(self.book_path)

            with pd.ExcelWriter(path=self.book_path, mode='a', engine="openpyxl") as writer:
                for name_sheet, df_control in control_df_dict.items():
                    if '{I}' in name_sheet:
                        # Поиск столбцов с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                        # https/www.geeksforgeeks.org/how-to-find-drop-duplicate-columns-in-a-pandas-dataframe/
                        df_control_head = df_control.iloc[:4].T  # включая groupid
                        duplicated_true = df_control_head.duplicated(keep=False)
                        groupid_true = df_control.loc['-', '-', 'groupid'] > 0
                        selection_columns = duplicated_true & groupid_true  # выборка в столбцах df_control для проверки

                        dict_equals = defaultdict(list)  # {номер:[перечень индексов столбцов с одинаковыми колонками]}
                        if selection_columns.any():
                            df_control_head = df_control_head[selection_columns]
                            duplicated_unique = df_control_head.drop_duplicates()
                            for i in range(len(duplicated_unique)):
                                col_unique = duplicated_unique.iloc[i, :]
                                for ii in range(len(df_control_head)):
                                    control_col = df_control_head.iloc[ii, :]
                                    if col_unique.equals(control_col):
                                        dict_equals[str(i)].append(int(control_col.name))
                        # Объединить столбцы с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                        if dict_equals:
                            for cols in dict_equals.values():
                                df_control[cols[0]] = df_control[cols].max(axis=1)
                                df_control.drop(columns=cols[1:], inplace=True)

                    df_control.to_excel(excel_writer=writer,
                                        sheet_name=name_sheet,
                                        header=False,
                                        startrow=1,
                                        freeze_panes=(2, 3),
                                        index=True)

            # Форматирование таблиц Отключение - Контроль
            wb = load_workbook(self.book_path)
            for name_sheet in control_df_dict:
                ws = wb[name_sheet]
                num_tab, name_tab = Common.read_title(self.config['te_tab_KO_info'])
                ws['A1'] = f'{name_tab[0]}{num_tab + self.file_count - 1}{name_tab[1]} {rm.info_file["Имя режима"]}'
                ws['A2'] = 'Наименование режима'
                ws['B2'] = 'Номер режима'
                ws['C2'] = 'Наименование параметра'
                # ws.merge_cells('A2:B4')
                thins = Side(border_style="thin", color="000000")
                max_column_lit = get_column_letter(ws.max_column)
                ws.merge_cells(f'A1:{max_column_lit}1')

                # Данные
                for row in range(3, ws.max_row + 1):
                    for col in range(4, ws.max_column + 1):
                        ws.cell(row, col).border = Border(thins, thins, thins, thins)
                        if ws.cell(row, 3).value in ['I, А', 'U, кВ']:
                            ws.cell(row, col).font = Font(bold=True)
                        if 'I, %' in ws.cell(row, 3).value:
                            if ws.cell(row, col).value >= 100:
                                ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor="00FF9900")
                        if 'U, %' in ws.cell(row, 3).value:
                            if ws.cell(row, col).value > 0:
                                ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor="00FF9900")
                # Колонки
                for litter, L in {'A': 35, 'B': 6, 'C': 17}.items():
                    ws.column_dimensions[litter].width = L
                for n in range(4, ws.max_column + 1):
                    ws[f'{get_column_letter(n)}2'].alignment = Alignment(textRotation=90, wrap_text=True,
                                                                         horizontal="center", vertical="center")
                    ws.column_dimensions[get_column_letter(n)].width = 9
                    ws[f'{get_column_letter(n)}2'].font = Font(bold=True)
                    ws[f'{get_column_letter(n)}2'].border = Border(thins, thins, thins, thins)
                # Строки
                if '{I}' in name_sheet:
                    ws.row_dimensions[5].hidden = True  # Скрыть
                    ws.row_dimensions[6].hidden = True
                for n in range(1, ws.max_row + 1):
                    ws[f'A{n}'].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws[f'B{n}'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    ws[f'C{n}'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 145
            wb.save(self.book_path)

    def calc_comb(self, rm, comb: pd.DataFrame, source: str = 'combinatorics'):
        """
        Смоделировать отключение элементов в комбинации.
        :param rm:
        :param comb:
        :param source: Варианты: 'combinatorics' или 'xl'
        :return:
        """
        # Фильтр н-2-3
        if self.config['filter_comb'] and len(comb) > 1 and source == 'combinatorics':
            if not len(comb.loc[(comb.table == 'node') | comb.double_repair_scheme]):
                count_effect = 0  # если 1 элемент из пары влияет на другой, то прибавляем 1
                s_key0 = comb.s_key[0]
                s_key1 = comb.s_key[1]
                el0 = (s_key0, comb.disable_scheme[0], comb.repair_scheme[0],)
                el1 = (s_key1, comb.disable_scheme[1], comb.repair_scheme[1],)
                # Если хотя бы 1 оказывает влияние на загрузку второго
                if ((s_key1 in self.disable_effect.get(el0, [])) or
                        (s_key0 in self.disable_effect.get(el1, []))):
                    count_effect += 1
                    # Если ветви состоят в одном транзите
                    if rm.v__num_transit.get(s_key0, 0) == rm.v__num_transit.get(s_key1, 1):
                        log_g_s.debug(f'В одном транзите: {tabulate(comb, headers="keys", tablefmt="psql")}')
                        return False

                elif len(comb) == 3:
                    s_key2 = comb.s_key[2]
                    el2 = (s_key2, comb.disable_scheme[2], comb.repair_scheme[2])

                    if ((s_key2 in self.disable_effect.get(el0, [])) or
                            (s_key0 in self.disable_effect.get(el2, []))):
                        count_effect += 1
                        if rm.v__num_transit.get(s_key0, 0) == rm.v__num_transit.get(s_key2, 2):
                            log_g_s.debug(f'В одном транзите: {tabulate(comb, headers="keys", tablefmt="psql")}')
                            return False
                    if ((s_key2 in self.disable_effect.get(el1, [])) or
                            (s_key1 in self.disable_effect.get(el2, []))):
                        count_effect += 1
                        if rm.v__num_transit.get(s_key1, 1) == rm.v__num_transit.get(s_key2, 2):
                            log_g_s.debug(f'В одном транзите: {tabulate(comb, headers="keys", tablefmt="psql")}')
                            return False

                if count_effect < len(comb) - 1:
                    log_g_s.debug(f'Комбинация отклонена фильтром: {tabulate(comb, headers="keys", tablefmt="psql")}')
                    return False

        # Восстановление схемы
        if self.restore_only_state:
            for name_table in rm.data_save_sta:
                rm.rastr.tables(name_table).ReadSafeArray(2,
                                                          rm.data_columns_sta[name_table],
                                                          rm.data_save_sta[name_table])
            log_g_s.debug('Состояние элементов сети восстановлено.')
        else:
            for name_table in rm.data_save:
                rm.rastr.tables(name_table).ReadSafeArray(2,
                                                          rm.data_columns[name_table],
                                                          rm.data_save[name_table])
            self.restore_only_state = True
            log_g_s.debug('Состояние элементов сети и параметров восстановлено.')
        # log_g_s.debug(tabulate(comb, headers='keys', tablefmt='psql'))
        comb.sort_values(by='status_repair', inplace=True)

        # Для добавления в 'Наименование СРС' данных о disable_scheme, double_repair_scheme и repair_scheme
        comb['scheme_info'] = ''
        log_g_s.debug(tabulate(comb, headers='keys', tablefmt='psql'))

        # Отключение элементов
        repair2_one = True  # Для выполнения действия с двойным отключением на 2-м элементе.
        for i in comb.index:
            if not rm.sta(table_name=comb.loc[i, 'table'],
                          ndx=comb.loc[i, 'index']):  # отключить элемент
                log_g_s.info(f'Комбинация отклонена: элемент {rm.t_name[comb.loc[i, "table"]][comb.loc[i, "s_key"]]!r}'
                             f' отключен в исходной РМ.')
                return False
            scheme_info = ''

            # Ремонтная схема
            if comb.loc[i, 'status_repair']:
                if comb.loc[i, 'double_repair_scheme']:
                    if repair2_one:
                        repair2_one = False
                    else:
                        scheme_info = self.perform_action(rm, comb.loc[i, 'double_repair_scheme'])
                else:
                    if comb.loc[i, 'repair_scheme']:
                        scheme_info = self.perform_action(rm, comb.loc[i, 'repair_scheme'])

            # Схема при отключении
            if (not comb.loc[i, 'status_repair']) and comb.loc[i, 'disable_scheme']:
                scheme_info = self.perform_action(rm, comb.loc[i, 'disable_scheme'])

            if scheme_info:
                comb.loc[i, 'scheme_info'] = f' ({scheme_info})'
        log_g_s.debug('Элементы сети из сочетания отключены.')

        # Имя сочетания
        for k in ['Отключение', 'Ключ откл.', 'Ремонт 1', 'Ключ рем.1', 'Ремонт 2', 'Ключ рем.2']:
            self.info_srs.pop(k, None)

        dname = rm.t_name[comb["table"].iloc[0]][comb["s_key"].iloc[0]]
        if comb.iloc[0]["status_repair"]:
            all_name_srs = 'Ремонт '
            self.info_srs['Ремонт 1'] = dname + comb['scheme_info'].iloc[0]
            self.info_srs['Ключ рем.1'] = RastrModel.key_to_str(comb["s_key"].iloc[0])
        else:
            all_name_srs = 'Отключение '
            self.info_srs['Отключение'] = dname + comb['scheme_info'].iloc[0]
            self.info_srs['Ключ откл.'] = RastrModel.key_to_str(comb["s_key"].iloc[0])

        name_srs_base = all_name_srs + dname
        all_name_srs += dname + comb['scheme_info'].iloc[0]
        if len(comb) > 1:
            dname = rm.t_name[comb["table"].iloc[1]][comb["s_key"].iloc[1]]
            all_name_srs += ' при ремонте' if 'Откл' in all_name_srs else ' и'
            all_name_srs += f' {dname}{comb["scheme_info"].iloc[1]}'
            name_srs_base += ' при ремонте' if 'Откл' in all_name_srs else ' и'
            name_srs_base += f' {dname}'
            if comb.iloc[0]["status_repair"]:
                self.info_srs['Ремонт 2'] = dname + comb["scheme_info"].iloc[1]
                self.info_srs['Ключ рем.2'] = comb["s_key"].iloc[1]
            else:
                self.info_srs['Ремонт 1'] = dname + comb["scheme_info"].iloc[1]
                self.info_srs['Ключ рем.1'] = comb["s_key"].iloc[1]
        if len(comb) == 3:
            dname = rm.t_name[comb["table"].iloc[2]][comb["s_key"].iloc[2]]
            all_name_srs += f', {dname}{comb["scheme_info"].iloc[2]}'
            name_srs_base += f', {dname}'
            self.info_srs['Ремонт 2'] = dname + comb["scheme_info"].iloc[2]
            self.info_srs['Ключ рем.2'] = comb["s_key"].iloc[2]

        self.info_srs['Наименование СРС без()'] = name_srs_base  # re.sub(r'\(.+\)', '', all_name_srs).strip()
        all_name_srs += '.'

        self.info_srs['Наименование СРС'] = all_name_srs.strip()
        self.info_srs['comb_id'] = self.comb_id
        self.info_srs['Кол. откл. эл.'] = comb.shape[0]
        log_g_s.info(f"Сочетание {self.comb_id}: {all_name_srs}")

        self.do_action(rm, comb)

    def perform_action(self, rm, task_action: list) -> str:
        """
        Выполнить действия, записанные в поле repair_scheme, disable_scheme.
        :param task_action: list("10", "2")
        :param rm:
        :return: Наименование внесенных изменений в расчетное НВ.
        """
        info = []
        # if not type(task_action) == tuple:
        #     task_action = tuple(task_action)
        for task_action_i in task_action:
            names, actions = self.pa.scheme_description(number=task_action_i)
            for i, action in enumerate(actions):
                name = rm.cor_rm_from_txt(action)
                if name:
                    if self.restore_only_state:
                        self.test_not_only_sta(name)
                    if names[i]:
                        name = names[i]
                    info.append(name)

        all_info = ', '.join(info) if info else ''
        return all_info

    def test_not_only_sta(self, txt):
        """
        Проверка на наличие изменений в сети кроме состояния.
        :param txt: Строка сформированная group_cor
        """
        for i in ['нагрузки', 'генерации', 'ktr', 'pn', 'qn', 'pg', 'qg', 'vzd', 'bsh', 'P']:
            # список параметров сверять с функцией group_cor, data_columns
            if i in txt:
                self.restore_only_state = False
                break

    def do_action(self, rm, comb=pd.DataFrame()):
        """
        Цикл по действиям ПА для ввода режима в область допустимых значений.
        :param rm:
        :param comb:
        """
        self.info_srs['comb_id'] = self.comb_id
        self.all_comb = pd.concat([self.all_comb, pd.Series(self.info_srs).to_frame().T],
                                  axis=0, ignore_index=True)
        self.info_action = dict()
        self.info_action['comb_id'] = self.comb_id
        self.info_action['active_id'] = 1
        self.info_action['End'] = False
        self.info_action['alive'] = 1
        # Если False - значит есть ПА, True - конец расчета сочетания (перегрузку нечем ликвидировать или отсутствует).
        # Цикл по действиям (ПА или ОП)
        while True:
            overloads = self.do_control(rm, comb)
            if self.config['CalcSetWindow']['pa'] and self.pa.active(overloads):  # TODO overloads не задано
                self.info_action['Action'] += self.pa.execute_action_pa(rm)
            else:
                if self.config['CalcSetWindow']['pa']:
                    self.pa.reset()
                self.info_action['End'] = True

            self.all_actions = pd.concat([self.all_actions, pd.Series(self.info_action).to_frame().T],
                                         axis=0, ignore_index=True)
            if self.info_action['End']:
                self.comb_id += 1  # код комбинации
                return

            self.info_action['active_id'] += 1
            for k in ['АРВ', 'СКРМ', 'Action']:
                self.info_action.pop(k, None)

    def do_control(self, rm, comb=pd.DataFrame()):
        """
        Проверка параметров режима.
        Заполняет таблицу Контроль - Отключение.
        Наполняет self.breach['i', 'low_u', 'high_u'].
        return overloads
        """
        log_g_s.debug(f'Проверка параметров УР.')
        violation = False
        test_rgm = rm.rgm('do_control')
        if self.config['CalcSetWindow']['avr'] and len(comb):
            self.info_action['АРВ'] = rm.node_include()
            if 'Восстановлено' in self.info_action['АРВ']:
                test_rgm = rm.rgm('Перерасчет после действия АВР.')
        # if self.config['CalcSetWindow']['skrm']:
        #     self.info_action['СКРМ'] = rm.auto_shunt_cor(all_auto_shunt=self.auto_shunt)
        #     if self.info_action['СКРМ']:
        #         test_rgm = rm.rgm('do_control')

        if not test_rgm:
            self.info_action['alive'] = 0
            log_g_s.debug(f'Режим развалился.')
            return None
        else:
            # Сохранение загрузки отключаемых элементов в н-1 для фильтра
            if self.config['filter_comb'] and len(comb) == 1 and comb.table[0] == 'vetv':
                table = rm.rastr.tables('vetv')
                table.setsel("all_disable")

                for index, i_zag in table.writesafearray('index,i_zag', "000"):
                    difference_i_zag = abs(i_zag - self.disable_df_vetv.loc[index, 'i_zag'])
                    if difference_i_zag > self.config['filter_comb_val']:
                        log_g_s.info((comb.s_key[0], comb.disable_scheme[0], comb.repair_scheme[0]))
                        log_g_s.info(self.disable_df_vetv.loc[index, 's_key'])
                        self.disable_effect[(comb.s_key[0],
                                             comb.disable_scheme[0],
                                             comb.repair_scheme[0])].append(self.disable_df_vetv.loc[index, 's_key'])

                # col_i_zag = rm.df_from_table(table_name='vetv', fields='index,i_zag', setsel="all_disable")
                # col_i_zag.set_index('index', inplace=True)
                # col_name = comb['key'][0]
                # for col in ['repair_scheme', 'disable_scheme']:
                #     if comb[col][0]:
                #         col_name += comb[col][0]
                # col_i_zag.rename(columns={'i_zag': col_name}, inplace=True)
                # self.disable_df_vetv = pd.concat([self.disable_df_vetv, col_i_zag], axis=1)
                # self.disable_df_vetv[col_name] = self.disable_df_vetv['i_zag'] - self.disable_df_vetv[col_name]

            # проверка на наличие перегрузок ветвей (ЛЭП, трансформаторов, выключателей)
            if self.info_srs['Контроль ДТН'] == 'АДТН':
                selection_v = 'all_control & i_zag_av > 0.1004'
                selection_n = 'all_control & vras<umin_av & !sta'
            else:
                selection_v = 'all_control & i_zag > 0.1004'
                selection_n = 'all_control & vras<umin & !sta'

            tv = rm.rastr.tables('vetv')
            tv.SetSel(selection_v)
            # overloads_i = None
            # high_voltage = None
            # low_voltages = None
            if tv.count:
                overloads_i = rm.df_from_table(table_name='vetv',
                                               fields='s_key,'  # 'Ключ контроль,'
                                                      'txt_zag,'  # 'txt_zag,' 
                                                      'i_max,'  # 'Iрасч.(A),'
                                                      'i_dop_r,'  # 'Iддтн(A),'
                                                      'i_zag,'  # 'Iзагр.ддтн(%),'
                                                      'i_dop_r_av,'  # 'Iадтн(A),'
                                                      'i_zag_av',  # 'Iзагр.адтн(%),'
                                               setsel=selection_v)
                overloads_i['comb_id'] = self.comb_id
                overloads_i['active_id'] = self.info_action['active_id']
                self.breach['i'] = pd.concat([self.breach['i'], overloads_i], axis=0, ignore_index=True)
                violation = True
                log_g_s.info(f'Выявлено {len(overloads_i)} превышений ДТН.')
            # проверка на наличие недопустимого снижение напряжения
            tn = rm.rastr.tables('node')
            tn.SetSel(selection_n)
            if tn.count:
                low_voltages = rm.df_from_table(table_name='node',
                                                fields='ny,'  # 'Ключ контроль,'
                                                # 'txt_zag,'  # 'txt_zag,'
                                                # todo сделать что бы в txt_zag были значения узлов?
                                                       'vras,'  # 'Uрасч.(кВ),'
                                                       'umin,'  # 'Uмин.доп.(кВ),'
                                                       'umin_av,'  # 'U ав.доп.(кВ),'
                                                       'otv_min,'
                                                # отклонение vras от 'Uмин.доп.' (%)
                                                       'otv_min_av',
                                                # отклонение vras от 'U ав.доп.' (%)
                                                setsel=selection_n)
                low_voltages.rename(columns={'ny': 's_key'}, inplace=True)
                low_voltages['comb_id'] = self.comb_id
                low_voltages['active_id'] = self.info_action['active_id']
                self.breach['low_u'] = pd.concat([self.breach['low_u'], low_voltages], axis=0, ignore_index=True)
                violation = True
                log_g_s.info(f'Выявлено {len(low_voltages)} точек недопустимого снижения напряжения.')

            # проверка на наличие недопустимого повышения напряжения
            tn.SetSel('all_control & umax<vras & umax>0 & !sta')
            if tn.count:
                high_voltage = rm.df_from_table(table_name='node',
                                                fields='ny,'  # 'Ключ контроль,'
                                                       'vras,'  # 'Uрасч.(кВ),'
                                                       'umax,'  # 'Uнаиб.раб.(кВ)'
                                                       'otv_max',  # 'Uнаиб.раб.(кВ)'
                                                setsel='all_control & umax<vras & umax>0 & !sta')
                high_voltage.rename(columns={'ny': 's_key'}, inplace=True)
                high_voltage['comb_id'] = self.comb_id
                high_voltage['active_id'] = self.info_action['active_id']
                self.breach['high_u'] = pd.concat([self.breach['high_u'], high_voltage], axis=0, ignore_index=True)
                violation = True
                log_g_s.info(f'Выявлено {len(high_voltage)} точек недопустимого превышения напряжения.')

            if self.config['cb_save_i']:
                save_i = rm.df_from_table(table_name='vetv',
                                          fields='s_key,'  # 'Ключ контроль,'
                                                 'i_max,'  # 'Iрасч.(A),'
                                                 'i_dop_r,'  # 'Iддтн(A),'
                                                 'i_zag,'  # 'Iзагр.ддтн(%),'
                                                 'i_dop_r_av,'  # 'Iадтн(A),'
                                                 'i_zag_av',  # 'Iзагр.адтн(%),'
                                          setsel='all_control')
                save_i['comb_id'] = self.comb_id
                save_i['active_id'] = self.info_action['active_id']
                self.save_i_rm = pd.concat([self.save_i_rm, save_i], axis=0)

            # Таблица КОНТРОЛЬ - ОТКЛЮЧЕНИЕ
            if self.config['cb_tab_KO']:
                log_g_s.debug('Запись параметров УР в таблицу КО.')
                if len(self.control_I):
                    ci = rm.df_from_table(table_name='vetv',
                                          fields='index,i_max,i_zag,i_zag_av',
                                          setsel="all_control")
                    ci.set_index('index', inplace=True)
                    ci['i_max'] = ci['i_max'].round(0)
                    ci['i_zag'] = ci['i_zag'].round(0)
                    ci['i_zag_av'] = ci['i_zag_av'].round(0)
                    ci = ci.T
                    ci.index = pd.MultiIndex.from_product([[self.info_srs['Наименование СРС']],
                                                           [f'{self.comb_id}.'
                                                            f'{self.info_action["active_id"]}'],
                                                           ['I, А', 'I, % от ДДТН', 'I, % от АДТН']])
                    self.control_I = pd.concat([self.control_I, ci], axis=0)

                if len(self.control_U):
                    cu = rm.df_from_table(table_name='node',
                                          fields='index,vras,otv_min,otv_min_av',
                                          setsel="all_control")
                    cu.set_index('index', inplace=True)
                    cu['vras'] = cu['vras'].round(1)
                    cu['otv_min'] = cu['otv_min'].round(2)
                    cu['otv_min_av'] = cu['otv_min_av'].round(2)
                    cu = cu.T
                    cu.index = pd.MultiIndex.from_product([[self.info_srs['Наименование СРС']],
                                                           [f'{self.comb_id}.'
                                                            f'{self.info_action["active_id"]}'],
                                                           ['U, кВ', 'U, % от МДН', 'U, % от АДН']])
                    self.control_U = pd.concat([self.control_U, cu], axis=0)

        # Добавить рисунки.
        if self.config['results_RG2'] and (not self.config['pic_overloads'] or
                                           (self.config['pic_overloads'] and violation)):
            log_g_s.debug('Добавить рисунки.')
            pic_name_file = rm.save(folder_name=self.config['folder_result_calc'],
                                    file_name=f'{rm.name_base} '
                                              f'[{self.comb_id}_{self.info_action["active_id"]}] '
                                              f'рис {self.num_pic} {self.info_srs["Наименование СРС без()"]}')

            # Южный р-н. Зимний максимум нагрузки 2026 г (-32°C/ПЭВТ). Нормальная схема сети. Действия...Загрузка...
            # todo Действия...Загрузка...
            add_name = f' ({", ".join(rm.additional_name_list)})' if rm.additional_name_list else ""
            picture_name = (f'{self.name_pic[0]}{self.num_pic}{self.name_pic[1]} '
                            f'{rm.info_file["Сезон макс/мин"]} {rm.info_file["Год"]} г'
                            f'{add_name}. {self.info_srs["Наименование СРС"]}')
            pic_name_file = pic_name_file.replace(self.config['folder_result_calc'] + '\\', '')
            self.all_pic.loc[len(self.all_pic.index)] = (self.num_pic,
                                                         self.comb_id,
                                                         self.info_action['active_id'],
                                                         pic_name_file,
                                                         picture_name)
            self.num_pic += 1
        return None

    @staticmethod
    def find_double_repair_scheme(comb_df):
        """
        Функция поиска общего действия double_repair_scheme в ремонтируемых элементах comb.
        Добавляет в колонку double_repair_scheme общее действие из колонки double_repair_scheme_copy и возвращает его.
        :param comb_df:
        """
        double_repair_scheme = []
        if comb_df.loc[comb_df['status_repair'], 'double_repair_scheme_copy'].all():
            double_repair_scheme = comb_df.loc[comb_df['status_repair'], 'double_repair_scheme_copy'].to_list()
            double_repair_scheme = list(set(double_repair_scheme[0]) & set(double_repair_scheme[1]))
            for i in comb_df.index:
                if comb_df['status_repair'].iloc[i]:
                    comb_df['double_repair_scheme'].iloc[i] = double_repair_scheme
                else:
                    comb_df['double_repair_scheme'].iloc[i] = False
        return double_repair_scheme
