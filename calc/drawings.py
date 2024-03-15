__all__ = ['Drawings']

import logging
import os
import re

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Alignment, Border, PatternFill

from collection_func import create_table

log_draw = logging.getLogger(f'__main__.{__name__}')


class Drawings:
    """Сохранение смоделированной СМС в формате rg2, для последующей вставки в word"""
    _folder_path = ''
    _book_path = ''

    def __init__(self, name_drawing: str):
        assert bool(name_drawing)
        assert bool(isinstance(name_drawing, str))

        self._name_drawing = name_drawing

        self._num_drawing = int(re.findall('\[(\d+?)]',
                                           self._name_drawing)[0])
        # Для хранения имен файлов с рисунками и имен рисунков
        self.df_drawing = pd.DataFrame(columns=['pic_id',
                                                'comb_id',
                                                'active_id',
                                                'Наименование файла',
                                                'Наименование рисунка'])

    def draw(self,
             rm,
             folder_path: str,
             name_srs: str,
             comb_id: int,
             active_id: int):
        """
        Сохранение смоделированной СМС в формате rg2
        :param rm:
        :param folder_path: Имя папки для сохранения rg2
        :param name_srs:
        :param comb_id:
        :param active_id:
        """
        assert bool(rm)
        assert bool(isinstance(folder_path, str))
        assert bool(isinstance(name_srs, str))
        assert bool(isinstance(comb_id, int))
        assert bool(isinstance(active_id, int))

        log_draw.debug('Сохранение файла rg2 для рисунки.')

        self._folder_path = folder_path
        full_name_file = rm.save(folder_name=folder_path,
                                 file_name=f'[{comb_id}_{active_id}] '
                                           f'{rm.name_base} '
                                           f'рис_{self._num_drawing} '
                                           f'{name_srs}')

        # Южный р-н. Зимний максимум нагрузки 2026 г (-32°C ПЭВТ). Нормальная схема сети. Действия...Загрузка...
        # todo Действия...Загрузка...
        additional_name_str = ", ".join(rm.additional_name_list)
        additional_name = f' ({additional_name_str})' if rm.additional_name_list else ''

        cur_name_drawing = re.sub('\[\d+?]',
                                  str(self._num_drawing),
                                  self._name_drawing)

        cur_name_drawing = (f'{cur_name_drawing} '
                            f'{rm.info_file["Сезон макс/мин"]} '
                            f'{rm.info_file["Год"]} г'
                            f'{additional_name}. '
                            f'{name_srs}')

        name_file = os.path.basename(full_name_file)

        self.df_drawing.loc[len(self.df_drawing.index)] = (self._num_drawing,
                                                           comb_id,
                                                           active_id,
                                                           name_file,
                                                           cur_name_drawing)
        self._num_drawing += 1

    def add_to_xl(self, book_path: str):
        """
        Сохранить перечень режимов в файл xl
        :param book_path:
        """

        if not len(self.df_drawing):
            return

        self._book_path = book_path

        sheet_name_drawing = 'Рисунки'
        mode = 'a' if os.path.exists(book_path) else 'w'
        with pd.ExcelWriter(path=book_path,
                            mode=mode) as writer:
            self.df_drawing[['Наименование файла',
                             'Наименование рисунка']].to_excel(excel_writer=writer,
                                                               startrow=1,
                                                               index=False,
                                                               freeze_panes=(5, 1),
                                                               sheet_name=sheet_name_drawing)
        book = load_workbook(book_path)
        sheet = book[sheet_name_drawing]

        sheet.insert_rows(1, amount=3)

        sheet['A1'] = 'Формат листа (3 - А3, 4 - А4):'
        sheet['A2'] = 'Ориентация(1 - книжная, 0 - альбомная):'
        sheet['A3'] = 'Имя папки с файлами rg2:'
        sheet['B1'] = 3
        sheet['B2'] = 0
        sheet['B3'] = self._folder_path
        thins = Side(border_style='thin',
                     color='000000')
        for col in 'AB':
            sheet.column_dimensions[col].width = 100
            for r in '123':
                sheet[col + r].alignment = Alignment(horizontal='left')
                sheet[col + r].border = Border(thins, thins, thins, thins)
                sheet[col + r].fill = PatternFill(fill_type='solid',
                                                  fgColor='00B1E76E')

        create_table(sheet=sheet,
                     sheet_name=sheet_name_drawing,
                     point_start='A5')
        book.save(book_path)

    def add_macro(self, macro_path: str):
        """
        Скопировать макрос rbs в папку calc
        :param macro_path: Путь в файлу .rbs
        """
        assert bool(macro_path)
        assert bool(isinstance(macro_path, str))

        with open(macro_path) as macro:
            content_rbs = ''.join(macro.readlines())

        content_rbs = content_rbs.replace('папка с файлами', self._book_path)

        path = self._book_path.rsplit('.', 1)
        path = path[0] + '.rbs'

        with open(path, 'w') as macro_new:
            macro_new.write(content_rbs)

