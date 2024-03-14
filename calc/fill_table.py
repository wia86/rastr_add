__all__ = ['FillTable']

import logging
import os
import re
from collections import defaultdict

import pandas as pd
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from collection_func import s_key_vetv_in_tuple

log_fill_tb = logging.getLogger(f'__main__.{__name__}')


class FillTable:
    """Класс для заполнения таблиц контролируемые - отключаемые элементы в excel для каждой РМ."""
    _control_I = None
    _control_U = None
    _path_xl = None

    def __init__(self, rm, setsel: str):
        """
        Создается для каждого файла.
        :param rm: 
        :param setsel: Выборка в таблицах узлы и ветви для выбора контролируемых элементов
        """
        self._setsel = setsel
        log_fill_tb.debug('Инициализация таблицы "контролируемые - отключаемые" элементы.')

        # Ветви
        self._control_I = rm.df_from_table(table_name='vetv',
                                           fields='i_dop_r,i_dop_r_av,groupid,s_key,tip,ip,iq,key',
                                           setsel=self._setsel)
        if len(self._control_I):
            self._control_I.insert(0,
                                   'dname',
                                   self._control_I.s_key.apply(
                                       lambda x: rm.dt.t_name['vetv'][s_key_vetv_in_tuple(x)]))
            # Сортировка
            self._control_I['ip_unom'] = self._control_I.ip.apply(lambda x: rm.dt.ny_unom[x])
            self._control_I['iq_unom'] = self._control_I.iq.apply(lambda x: rm.dt.ny_unom[x])
            self._control_I['uhom'] = (self._control_I[['ip_unom', 'iq_unom']].max(axis=1) * 10000 +
                                       self._control_I[['ip_unom', 'iq_unom']].min(axis=1))
            self._control_I.sort_values(by=['tip', 'uhom', 'dname'],  # столбцы сортировки
                                        ascending=(False, False, True),  # обратный порядок
                                        inplace=True)
            self._control_I.drop(['ip_unom', 'iq_unom', 'uhom', 'tip', 'ip', 'iq', 's_key'],
                                 axis=1,
                                 inplace=True)

            self._control_I = self._control_I.round({'i_dop_r': 0,
                                                     'i_dop_r_av': 0})
            self._control_I.rename(columns={'i_dop_r': 'ДДТН, А',
                                            'i_dop_r_av': 'АДТН, А',
                                            'dname': 'Контролируемый элемент'}, inplace=True)
            self._control_I.set_index('key', inplace=True)
            self._control_I = self._control_I.T
            self._control_I.index = pd.MultiIndex.from_product([['-'],
                                                                ['-'],
                                                                self._control_I.index])

        # Узлы
        self._control_U = rm.df_from_table(table_name='node',
                                           fields='ny,umin,umin_av,uhom',
                                           setsel=self._setsel)
        if len(self._control_U):
            self._control_U.insert(0,
                                   'dname',
                                   self._control_U.ny.apply(lambda x: rm.dt.t_name['node'][x]))
            self._control_U.sort_values(by=['uhom', 'dname'],
                                        ascending=(False, True),
                                        inplace=True)
            self._control_U.drop(['uhom'], axis=1, inplace=True)
            self._control_U = self._control_U.round({'umin': 1,
                                                     'umin_av': 1})
            self._control_U.rename(columns={'umin': 'МДН, кВ',
                                            'umin_av': 'АДН, кВ',
                                            'dname': 'Контролируемый элемент'}, inplace=True)
            self._control_U.set_index('ny', inplace=True)
            self._control_U = self._control_U.T
            self._control_U.index = pd.MultiIndex.from_product([['-'],
                                                                ['-'],
                                                                self._control_U.index])

    def add_value(self, rm, name_srs: str, comb_id: int, active_id: int):
        """
        Добавить значения
        :param rm:
        :param name_srs:
        :param comb_id:
        :param active_id:
        """
        log_fill_tb.debug('Запись параметров УР в таблицу КО.')
        if len(self._control_I):
            ci = rm.df_from_table(table_name='vetv',
                                  fields='key,i_max,i_zag,i_zag_av',
                                  setsel=self._setsel)

            ci.set_index('key', inplace=True)
            ci = ci.round({'i_max': 0, 'i_zag': 0, 'i_zag_av': 0})
            ci = ci.T
            ci.index = pd.MultiIndex.from_product([[name_srs],
                                                   [f'{comb_id}.'
                                                    f'{active_id}'],
                                                   ['I, А',
                                                    'I, % от ДДТН',
                                                    'I, % от АДТН']])
            self._control_I = pd.concat([self._control_I, ci], axis=0)

        if len(self._control_U):
            cu = rm.df_from_table(table_name='node',
                                  fields='vras,otv_min,otv_min_av,ny',
                                  setsel=self._setsel)
            cu.set_index('ny', inplace=True)
            cu = cu.round({'vras': 1, 'otv_min': 2, 'otv_min_av': 2})
            cu = cu.T
            cu.index = pd.MultiIndex.from_product([[name_srs],
                                                   [f'{comb_id}.'
                                                    f'{active_id}'],
                                                   ['U, кВ',
                                                    'U, % от МДН',
                                                    'U, % от АДН']])
            self._control_U = pd.concat([self._control_U, cu], axis=0)

    def insert_tables_to_xl(self,
                            name_rm: str,
                            file_name: str,
                            file_count: int,
                            name_table: str,
                            path_xl: str):
        """
        Вставить таблицы в excel. Поправить оформление.
        :param name_rm: Имя файла без расширения
        :param file_name: Расшифрованное имя файла
        :param file_count: Порядковый номер текущего расчетного файла
        :param name_table: В формате 'Таблица [1] - ...'
        :param path_xl: Файл для сохранения таблиц в формате excel.
        """
        self._path_xl = path_xl
        if not (len(self._control_I) or len(self._control_U)):
            return

        num_tab = re.findall('\[(\d+?)]', name_table)[0]
        num_tab = int(num_tab) + file_count - 1

        name_table = f'{name_table} {name_rm}'

        file_name = re.sub('\[]', '', file_name)

        control_df_dict = {}  # имя листа: df
        if len(self._control_I):
            name_sheet = f'{num_tab}_I_{file_name}'[:30]
            control_df_dict[name_sheet] = self._control_I
        if len(self._control_U):
            name_sheet = f'{num_tab}_U_{file_name}'[:30]
            control_df_dict[name_sheet] = self._control_U

        # https://www.geeksforgeeks.org/how-to-write-pandas-dataframes-to-multiple-excel-sheets/
        if not os.path.exists(self._path_xl):
            Workbook().save(self._path_xl)

        with pd.ExcelWriter(path=self._path_xl, mode='a', engine='openpyxl') as writer:
            for name_sheet, df_control in control_df_dict.items():
                if '_I' in name_sheet:
                    # Поиск столбцов с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                    # https/www.geeksforgeeks.org/how-to-find-drop-duplicate-columns-in-a-pandas-dataframe/
                    df_control_head = df_control.iloc[:4].T  # включая groupid
                    duplicated_true = df_control_head.duplicated(keep=False)
                    groupid_true = df_control.loc['-', '-', 'groupid'] > 0
                    # Выборка в столбцах df_control для проверки
                    selection_columns = duplicated_true & groupid_true

                    # {номер:[перечень индексов столбцов с одинаковыми колонками]}
                    dict_equals = defaultdict(list)
                    if selection_columns.any():
                        df_control_head = df_control_head[selection_columns]
                        duplicated_unique = df_control_head.drop_duplicates()
                        for i in range(len(duplicated_unique)):
                            col_unique = duplicated_unique.iloc[i, :]
                            for ii in range(len(df_control_head)):
                                control_col = df_control_head.iloc[ii, :]
                                if col_unique.equals(control_col):
                                    dict_equals[str(i)].append(control_col.name)
                    # Объединить столбцы с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                    if dict_equals:
                        for cols in dict_equals.values():
                            df_control[cols[0]] = df_control[cols].max(axis=1)
                            df_control.drop(columns=cols[1:], inplace=True)
                    df_control.drop(index=('-', '-', 'groupid'), inplace=True)

                df_control.to_excel(excel_writer=writer,
                                    sheet_name=name_sheet,
                                    float_format='%.1f',
                                    header=False,
                                    startrow=1,
                                    freeze_panes=(2, 3),
                                    index=True)

        # Форматирование таблиц Отключение - Контроль
        wb = load_workbook(self._path_xl)
        if 'Sheet' in wb:
            del wb['Sheet']
        for n, name_sheet in enumerate(control_df_dict, 1):
            ws = wb[name_sheet]

            ws['A1'] = re.sub('\[\d+?]',
                              f'{num_tab}.{n}',
                              name_table)
            ws['A2'] = 'Наименование режима'
            ws['B2'] = 'Номер режима'
            ws['C2'] = 'Наименование параметра'
            thins = Side(border_style='thin', color='000000')
            max_column_lit = get_column_letter(ws.max_column)
            ws.merge_cells(f'A1:{max_column_lit}1')

            # Данные таблицы: заливка, рамка и формат
            for row in range(3, ws.max_row + 1):
                for col in range(4, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    val = ws.cell(row, col).value
                    test_val = ws.cell(row, 3).value

                    cell.border = Border(thins, thins, thins, thins)

                    if test_val in ['I, А', 'U, кВ']:
                        cell.font = Font(bold=True)
                    if (('I, %' in test_val) and (val >= 100)) or (('U, %' in test_val) and (val < 0)):
                        cell.fill = PatternFill(fill_type='solid', fgColor='00FF9900')

            # Высота колонок
            for litter, L in (('A', 35),
                              ('B', 6),
                              ('C', 17)):
                ws.column_dimensions[litter].width = L

            # Шапка:
            for col in range(4, ws.max_column + 1):
                address = f'{get_column_letter(col)}2'
                ws[address].alignment = Alignment(textRotation=90,
                                                  wrap_text=True,
                                                  horizontal='center',
                                                  vertical='center')
                ws.column_dimensions[get_column_letter(col)].width = 9
                ws[address].font = Font(bold=True)
                ws[address].border = Border(thins, thins, thins, thins)
            # Левый ряд:
            for row in range(1, ws.max_row + 1):
                alignment = Alignment(wrap_text=True,
                                      horizontal='center',
                                      vertical='center')
                ws[f'A{row}'].alignment = alignment
                ws[f'B{row}'].alignment = alignment
                ws[f'C{row}'].alignment = alignment
            ws.row_dimensions[2].height = 145
        wb.save(self._path_xl)

    def insert_word(self):
        """Вставить таблицы из excel в word."""
        log_fill_tb.info('Вставка таблицы К-О в word.')

        excel = win32com.client.Dispatch('Excel.Application')
        excel.DisplayAlerts = False  # Не показывать всплывающие окна
        excel.Visible = False
        try:
            book = excel.Workbooks.Open(self._path_xl)
        except:
            log_fill_tb.error('Ошибка вставки таблиц в word')
            return

        word = win32com.client.Dispatch('Word.Application')
        word.Visible = False
        word.ScreenUpdating = False
        doc = word.Documents.Add()

        doc.PageSetup.PageWidth = 29.7 * 28.35
        doc.PageSetup.PageHeight = 42.0 * 28.35
        doc.PageSetup.Orientation = 1  # 1 книжная или 0 альбомная

        cursor = word.Selection
        cursor.Font.Size = 12
        cursor.Font.Name = 'Times New Roman'
        cursor.EndKey(Unit=6)  # перейти в конец текста

        for i in range(1, book.Worksheets.Count + 1):
            sheet = book.Worksheets(i)
            cursor.TypeText(Text=sheet.Cells(1, 1).value)
            cursor.TypeParagraph()
            range_copy = sheet.UsedRange.address.replace('$', '').replace('A1', 'A2')
            sheet.Range(range_copy).Copy()
            # cursor.PasteExcelTable(LinkedToExcel=False, WordFormatting=False, RTF=False)
            cursor.PasteAndFormat(Type=13)  # 13 Вставить в виде рисунка.
            # разрыв:7 страницы с новой строки, 0-в той же строке,1 и 8 колонки,
            # 2-5 раздела со след стр,6 и 9-11 перенос на новую стр
            cursor.InsertBreak(Type=0)

        word.ScreenUpdating = True
        path_word = self._path_xl.rsplit('.', 1)[0]
        doc.SaveAs2(FileName=f'{path_word}.docx')  # FileFormat=16 .docx
        book.Close()
        doc.Close()
        excel.Quit()
        word.Quit()
