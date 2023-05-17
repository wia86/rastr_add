# установка модулей:
# Qt Designer для работы с файлами.ui
# pip freeze > requirements.txt
# pip install -r requirements.txt
# exe приложение:
# pyinstaller --onefile --noconsole main.py
import win32com.client  # установить pywin32
# Excel.Application https://memotut.com/en/150745ae0cc17cb5c866/
from abc import ABC
from Rastr_Method import RastrMethod
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
from typing import Union  # Any
# from urllib.request import urlopen
import sys
import shutil
from itertools import combinations
from PyQt5 import QtWidgets
from datetime import datetime
import time
import os
import re
import configparser  # создать ini файл
# import random
import logging
# import webbrowser
from tkinter import messagebox as mb
import numpy as np
import pandas as pd
from tabulate import tabulate
import yaml
from qt_choice import Ui_choice  # pyuic5 qt_choice.ui -o qt_choice.py # Запустить строку в Terminal
from qt_set import Ui_Settings  # pyuic5 qt_set.ui -o qt_set.py
from qt_cor import Ui_cor  # pyuic5 qt_cor.ui -o qt_cor.py
from qt_calc_ur import Ui_calc_ur  # pyuic5 qt_calc_ur.ui -o qt_calc_ur.py
from qt_calc_ur_set import Ui_calc_ur_set  # pyuic5 qt_calc_ur_set.ui -o qt_calc_ur_set.py
from collections import namedtuple, defaultdict, Counter


# Если не работает терминал, то в  PowerShell ввести:
# Set-ExecutionPolicy -Scope CurrentUser -ExecutionPolicy RemoteSigned –Force


class Window:
    """ Класс с общими методами для QT. """

    @staticmethod
    def check_status(set_checkbox_element: tuple):
        """
        По состоянию CheckBox изменить состояние видимости соответствующего элемента.
        :param set_checkbox_element: картеж картежей (checkbox, element)
        """
        for CB, element in set_checkbox_element:
            if CB.isChecked():
                element.show()
            else:
                element.hide()

    @staticmethod
    def hide_show(hide_window: tuple, show_window: tuple):
        """ Изменить состояние видимости окон."""
        for element in hide_window:
            element.hide()
        for element in show_window:
            element.show()

    def choice_file(self, directory: str, filter_: str = 'All Files(*)'):
        """
        Выбор файла.
        """
        fileName_choose, _ = QtWidgets.QFileDialog.getOpenFileName(self, directory=directory,
                                                                   filter=filter_)  # "All Files(*);Text Files(*.txt)"
        if fileName_choose:
            log.info(f"GUI. Выбран файл: {fileName_choose}")
            return fileName_choose

    def choice_folder(self, directory: str):
        """
        Выбор папки.
        """
        folder_Name_choose = QtWidgets.QFileDialog.getExistingDirectory(self, directory=directory)
        if folder_Name_choose:
            return folder_Name_choose

    def save_file(self, directory: str, filter_: str = ''):
        """
        Сохранение файла.
        """
        fileName_choose, _ = QtWidgets.QFileDialog.getSaveFileName(self, directory=directory, filter=filter_)
        if fileName_choose:
            log.info(f"GUI. Для сохранения выбран файл: {fileName_choose}, {_}")
            return fileName_choose

    def choice(self, type_choice: str, insert, directory=None):
        """
        Функция выбора папки или файла.
        :param type_choice: 'file', 'folder'
        :param insert: объект QT 'QPlainTextEdit' или 'QLineEdit' для вставки пути выбранного файла.
        :param directory: объект QT 'QPlainTextEdit' c начальной папкой для поиска.
        """
        name = ''
        if type_choice == 'file':
            name = self.choice_file(directory=directory.toPlainText().replace('*', ''))
        elif type_choice == 'folder':
            name = self.choice_folder(directory=directory.toPlainText().replace('*', ''))

        if name:
            name = name.replace('/', '\\')
            if insert.__class__.__name__ == 'QPlainTextEdit':
                insert.setPlainText(name)
            elif insert.__class__.__name__ == 'QLineEdit':
                insert.setText(name)


class MainChoiceWindow(QtWidgets.QMainWindow, Ui_choice, Window):
    """
    Окно главного меню.
    """

    def __init__(self):
        super(MainChoiceWindow, self).__init__()
        self.setupUi(self)
        self.settings.clicked.connect(lambda: gui_set.show())
        self.correction.clicked.connect(lambda: self.hide_show((gui_choice_window,), (gui_edit,)))
        self.calc_ur.clicked.connect(lambda: self.hide_show((gui_choice_window,), (gui_calc_ur,)))


class CalcWindow(QtWidgets.QMainWindow, Ui_calc_ur, Window):
    """
    Окно задания и запуска УР.
    """

    def __init__(self):
        super(CalcWindow, self).__init__()
        self.setupUi(self)
        self.task_calc = {}
        self.b_set.clicked.connect(lambda: gui_calc_ur_set.show())
        self.b_main_choice.clicked.connect(lambda: self.hide_show((gui_calc_ur,), (gui_choice_window,)))

        # Скрыть параметры при старте.
        self.check_status_visibility = (
            (self.cb_filter, self.gb_filter),
            (self.cb_cor_txt, self.te_cor_txt),
            (self.cb_import_model, self.gb_import_model),
            (self.cb_disable_comb, self.gb_disable_comb),
            (self.cb_disable_excel, self.gb_disable_excel),
            (self.cb_control, self.gb_control),
            (self.cb_tab_KO, self.gb_tab_KO),
            (self.cb_results_pic, self.gb_results_pic),
        )
        self.check_status(self.check_status_visibility)

        # CB показать / скрыть параметры.
        for CB, _ in self.check_status_visibility:
            CB.clicked.connect(lambda: self.check_status(self.check_status_visibility))

        # Функциональные кнопки

        self.b_task_save.clicked.connect(self.task_save_yaml)
        self.b_task_load.clicked.connect(self.task_load_yaml)

        self.b_choice_path_folder.clicked.connect(lambda: self.choice(type_choice='folder',
                                                                      insert=self.te_path_initial_models,
                                                                      directory=self.te_path_initial_models))
        self.b_choice_path_file.clicked.connect(lambda: self.choice(type_choice='file',
                                                                    insert=self.te_path_initial_models,
                                                                    directory=self.te_path_initial_models))
        self.b_choice_XL.clicked.connect(lambda: self.choice(type_choice='file', insert=self.te_XL_path,
                                                             directory=self.te_path_initial_models))
        self.b_choice_path_import_folder.clicked.connect(lambda: self.choice(type_choice='folder',
                                                                             insert=self.te_path_import_rg2,
                                                                             directory=self.te_path_initial_models))
        self.b_choice_path_import_file.clicked.connect(lambda: self.choice(type_choice='file',
                                                                           insert=self.te_path_import_rg2,
                                                                           directory=self.te_path_initial_models))

        self.run_calc_rg2.clicked.connect(lambda: self.start())
        self.te_path_initial_models.setPlainText(GeneralSettings.read_ini(section='save_form_folder_calc', key="path"))
        # Подсказки
        self.le_control_field.setToolTip("Например, 'sel'. Если '*' - то контролировать все ветви и узлы")
        self.te_path_initial_models.setToolTip("Для расчета файлов во всех вложенных папках нужно в конце поставить *")

    def task_save_yaml(self):
        name_file_save = self.save_file(directory=self.te_path_initial_models.toPlainText(),
                                        filter_="YAML Files (*.yaml)")
        if name_file_save:
            self.fill_task_calc()
            with open(name_file_save, 'w') as f:
                yaml.dump(data=self.task_calc, stream=f, default_flow_style=False, sort_keys=False)

    def task_load_yaml(self):
        name_file_load = self.choice_file(directory=self.te_path_initial_models.toPlainText().replace('*', ''),
                                          filter_="YAML Files (*.yaml)")
        if not name_file_load:
            return
        with open(name_file_load) as f:
            task_yaml = yaml.safe_load(f)
        if not task_yaml:
            return

        # Окно запуска расчета.
        self.te_path_initial_models.setPlainText(task_yaml["calc_folder"])
        # Выборка файлов.
        self.cb_filter.setChecked(task_yaml["Filter_file"])  # QCheckBox
        self.sb_count_file.setValue(task_yaml["file_count_max"])  # QSpainBox
        self.le_condition_file_years.setText(task_yaml["calc_criterion"]["years"])  # QLineEdit text()
        self.le_condition_file_season.setCurrentText(task_yaml["calc_criterion"]["season"])  # QComboBox
        self.le_condition_file_max_min.setCurrentText(task_yaml["calc_criterion"]["max_min"])
        self.le_condition_file_add_name.setText(task_yaml["calc_criterion"]["add_name"])
        # Корректировка в txt.
        self.cb_cor_txt.setChecked(task_yaml["cor_rm"]['add'])
        self.te_cor_txt.setPlainText(task_yaml["cor_rm"]['txt'])
        # Импорт ИД для расчетов УР из моделей.
        self.cb_import_model.setChecked(task_yaml['CB_Import_Rg2'])
        self.te_path_import_rg2.setPlainText(task_yaml["Import_file"])
        self.te_import_rg2.setPlainText(task_yaml['txt_Import_Rg2'])
        # Расчет всех возможных сочетаний. Отключаемые элементы.
        self.cb_disable_comb.setChecked(task_yaml['cb_disable_comb'])
        self.cb_n1.setChecked(task_yaml['SRS']['n-1'])
        self.cb_n2.setChecked(task_yaml['SRS']['n-2'])
        self.cb_n3.setChecked(task_yaml['SRS']['n-3'])

        self.cb_auto_disable.setChecked(task_yaml['cb_auto_disable'])
        self.le_auto_disable_choice.setText(task_yaml['auto_disable_choice'])

        self.cb_comb_field.setChecked(task_yaml['cb_comb_field'])
        self.le_comb_field.setText(task_yaml['comb_field'])

        self.cb_filter_comb.setChecked(task_yaml['filter_comb'])
        self.le_filter_comb_val.setText(task_yaml['filter_comb_val'])
        # Импорт перечня расчетных сочетаний из EXCEL
        self.cb_disable_excel.setChecked(task_yaml['cb_disable_excel'])
        self.te_XL_path.setPlainText(task_yaml['srs_XL_path'])
        self.le_XL_sheets.setText(task_yaml['srs_XL_sheets'])
        # Расчет всех возможных сочетаний. Контролируемые элементы.
        self.cb_control.setChecked(task_yaml['cb_control'])
        self.cb_control_field.setChecked(task_yaml['cb_control_field'])
        self.le_control_field.setText(task_yaml['le_control_field'])
        self.cb_Imax.setChecked(task_yaml['cb_Imax'])
        # Результаты в EXCEL: таблицы контролируемые - отключаемые элементы
        self.cb_tab_KO.setChecked(task_yaml['cb_tab_KO'])
        self.te_tab_KO_info.setPlainText(task_yaml['te_tab_KO_info'])
        # Результаты в RG2
        self.cb_results_pic.setChecked(task_yaml['results_RG2'])
        self.cb_pic_overloads.setChecked(task_yaml['pic_overloads'])
        self.te_name_pic.setPlainText(task_yaml['name_pic'])
        # TODO настройки
        self.check_status(self.check_status_visibility)

    def start(self):
        """
        Запуск расчета моделей
        """
        GeneralSettings.write_ini(section='save_form_folder_calc', key="path",
                                  value=self.te_path_initial_models.toPlainText())
        self.fill_task_calc()
        global cm
        cm = CalcModel(self.task_calc)
        cm.run_calc()

    def fill_task_calc(self):
        self.task_calc = {
            # Окно запуска расчета.
            'calc_folder': self.te_path_initial_models.toPlainText().strip(),
            # Выборка файлов.
            'Filter_file': self.cb_filter.isChecked(),  # QCheckBox
            'file_count_max': self.sb_count_file.value(),  # QSpainBox
            'calc_criterion': {'years': self.le_condition_file_years.text(),  # QLineEdit text()
                               'season': self.le_condition_file_season.currentText(),  # QComboBox
                               'max_min': self.le_condition_file_max_min.currentText(),
                               'add_name': self.le_condition_file_add_name.text()},
            # Корректировка в txt.
            'cor_rm': {'add': self.cb_cor_txt.isChecked(),
                       'txt': self.te_cor_txt.toPlainText()},
            # Импорт ИД для расчетов УР из моделей.
            'CB_Import_Rg2': self.cb_import_model.isChecked(),
            'Import_file': self.te_path_import_rg2.toPlainText(),
            'txt_Import_Rg2': self.te_import_rg2.toPlainText(),
            # Расчет всех возможных сочетаний. Отключаемые элементы.
            'cb_disable_comb': self.cb_disable_comb.isChecked(),
            'SRS': {'n-1': self.cb_n1.isChecked(),
                    'n-2': self.cb_n2.isChecked(),
                    'n-3': self.cb_n3.isChecked()},

            'cb_auto_disable': self.cb_auto_disable.isChecked(),
            'auto_disable_choice': self.le_auto_disable_choice.text(),

            'cb_comb_field': self.cb_comb_field.isChecked(),
            "comb_field": self.le_comb_field.text(),

            'filter_comb': self.cb_filter_comb.isChecked(),
            'filter_comb_val': self.le_filter_comb_val.text(),
            # Импорт перечня расчетных сочетаний из EXCEL
            'cb_disable_excel': self.cb_disable_excel.isChecked(),
            'srs_XL_path': self.te_XL_path.toPlainText(),
            'srs_XL_sheets': self.le_XL_sheets.text(),
            # Расчет всех возможных сочетаний. Контролируемые элементы.
            'cb_control': self.cb_control.isChecked(),
            'cb_control_field': self.cb_control_field.isChecked(),
            'le_control_field': self.le_control_field.text(),
            'cb_Imax': self.cb_Imax.isChecked(),

            # Результаты в EXCEL: таблицы контролируемые - отключаемые элементы
            'cb_tab_KO': self.cb_tab_KO.isChecked(),
            'te_tab_KO_info': self.te_tab_KO_info.toPlainText(),

            # Результаты в RG2
            'results_RG2': self.cb_results_pic.isChecked(),
            'pic_overloads': self.cb_pic_overloads.isChecked(),
            'name_pic': self.te_name_pic.toPlainText(),
            # TODO настройки
        }
        """
        Заполнить task_calc задание взяв данные с формы QT.
        """


class CalcSetWindow(QtWidgets.QMainWindow, Ui_calc_ur_set, Window):
    """
    Окно основных настроек расчета УР.
    """

    def __init__(self):
        super(CalcSetWindow, self).__init__()
        self.setupUi(self)
        self.load_ini_ur()
        self.b_save.clicked.connect(lambda: self.save_ini_ur())

    def load_ini_ur(self):
        """Загрузить, создать или перезаписать файл .ini """
        if os.path.exists(GeneralSettings.ini):
            config = configparser.ConfigParser()
            config.read(GeneralSettings.ini)
            try:
                self.cb_gost.setChecked(eval(config['CalcSetWindow']["gost"]))
                self.cb_skrm.setChecked(eval(config['CalcSetWindow']["skrm"]))
                self.cb_avr.setChecked(eval(config['CalcSetWindow']["avr"]))
                self.cb_add_disabling_repair.setChecked(eval(config['CalcSetWindow']["add_disabling_repair"]))
                self.cb_pa.setChecked(eval(config['CalcSetWindow']["pa"]))
            except LookupError:
                log.error(f'файл {GeneralSettings.ini} не читается, перезаписан')
                self.save_ini_ur()
        else:
            log.info(f'создан файл {GeneralSettings.ini}')
            self.save_ini_ur()

    def save_ini_ur(self):
        config = configparser.ConfigParser()
        config.read(GeneralSettings.ini)
        config['CalcSetWindow'] = {
            "gost": self.cb_gost.isChecked(),
            "skrm": self.cb_skrm.isChecked(),
            "avr": self.cb_avr.isChecked(),
            "add_disabling_repair": self.cb_add_disabling_repair.isChecked(),
            "pa": self.cb_pa.isChecked()}
        with open(GeneralSettings.ini, 'w') as configfile:
            config.write(configfile)


class SetWindow(QtWidgets.QMainWindow, Ui_Settings, Window):
    """
    Окно общих настроек.
    """

    def __init__(self):
        super(SetWindow, self).__init__()
        self.setupUi(self)
        self.load_ini()
        self.set_save.clicked.connect(lambda: self.save_ini())

    def load_ini(self):
        """Загрузить, создать или перезаписать файл .ini """
        if os.path.exists(GeneralSettings.ini):
            config = configparser.ConfigParser()
            config.read(GeneralSettings.ini)
            try:
                self.LE_path.setText(config['DEFAULT']["folder RastrWin3"])
                self.LE_rg2.setText(config['DEFAULT']["шаблон rg2"])
                self.LE_rst.setText(config['DEFAULT']["шаблон rst"])
                self.LE_sch.setText(config['DEFAULT']["шаблон sch"])
                self.LE_amt.setText(config['DEFAULT']["шаблон amt"])
                self.LE_trn.setText(config['DEFAULT']["шаблон trn"])
            except LookupError:
                log.error(f'файл {GeneralSettings.ini} не читается, перезаписан')
                self.save_ini()
        else:
            log.info(f'создан файл {GeneralSettings.ini}')
            self.save_ini()

    def save_ini(self):
        config = configparser.ConfigParser()
        config.read(GeneralSettings.ini)
        config['DEFAULT'] = {
            "folder RastrWin3": self.LE_path.text(),
            "шаблон rg2": self.LE_rg2.text(),
            "шаблон rst": self.LE_rst.text(),
            "шаблон sch": self.LE_sch.text(),
            "шаблон amt": self.LE_amt.text(),
            "шаблон trn": self.LE_trn.text()}
        with open(GeneralSettings.ini, 'w') as configfile:
            config.write(configfile)


class EditWindow(QtWidgets.QMainWindow, Ui_cor, Window):
    """
    Окно корректировки моделей.
    """

    def __init__(self):
        super(EditWindow, self).__init__()  # *args, **kwargs
        self.setupUi(self)
        self.task_ui = {}
        self.check_import = (
            (self.CB_N, 'узлы'),
            (self.CB_V, 'ветви'),
            (self.CB_G, 'генераторы'),
            (self.CB_A, 'районы'),
            (self.CB_A2, 'территории'),
            (self.CB_D, 'объединения'),
            (self.CB_PQ, 'PQ'),
            (self.CB_IT, 'I(T)'),)

        # Скрыть параметры при старте.
        self.check_status_visibility = (
            (self.CB_KFilter_file, self.GB_sel_file),
            (self.CB_cor_b, self.TE_cor_b),
            (self.CB_ImpRg2, self.sel_import),
            (self.CB_import_val_XL, self.GB_import_val_XL),
            (self.CB_kontrol_rg2, self.GB_control),
            (self.CB_Filtr_N, self.FFN),
            (self.CB_Filtr_V, self.FFV),
            (self.CB_Filtr_G, self.FFG),
            (self.CB_Filtr_A, self.FFA),
            (self.CB_Filtr_A2, self.FFA2),
            (self.CB_Filtr_D, self.FFD),
            (self.CB_Filtr_PQ, self.FFPQ),
            (self.CB_Filtr_IT, self.FFIT),
            (self.CB_printXL, self.GB_prinr_XL),
            (self.CB_print_tab_log, self.GB_sel_tabl),
            (self.CB_print_parametr, self.TA_parametr_vibor),
            (self.CB_print_balance_Q, self.balance_Q_vibor),)
        self.check_status(self.check_status_visibility)

        # CB показать / скрыть параметры.
        for CB, element in self.check_status_visibility:
            CB.clicked.connect(lambda: self.check_status(self.check_status_visibility))
        # CB показать список импортируемых моделей.
        for CB, _ in self.check_import:
            CB.clicked.connect(lambda: self.import_name_table())

        # Функциональные кнопки
        self.task_save.clicked.connect(self.task_save_yaml)
        self.task_load.clicked.connect(self.task_load_yaml)
        self.choice_from_folder.clicked.connect(lambda: self.choice(type_choice='folder', insert=self.T_IzFolder,
                                                                    directory=self.T_IzFolder))
        self.choice_from_file.clicked.connect(lambda: self.choice(type_choice='file', insert=self.T_IzFolder,
                                                                  directory=self.T_IzFolder))
        self.choice_in_folder.clicked.connect(lambda: self.choice(type_choice='folder', insert=self.T_InFolder,
                                                                  directory=self.T_IzFolder))
        self.choice_XL.clicked.connect(lambda: self.choice(type_choice='file', insert=self.T_PQN_XL_File,
                                                           directory=self.T_IzFolder))
        self.choice_N.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_N,
                                                          directory=self.T_IzFolder))
        self.choice_V.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_V,
                                                          directory=self.T_IzFolder))
        self.choice_G.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_G,
                                                          directory=self.T_IzFolder))
        self.choice_A.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_A,
                                                          directory=self.T_IzFolder))
        self.choice_A2.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_A2,
                                                           directory=self.T_IzFolder))
        self.choice_D.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_D,
                                                          directory=self.T_IzFolder))
        self.choice_PQ.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_PQ,
                                                           directory=self.T_IzFolder))
        self.choice_IT.clicked.connect(lambda: self.choice(type_choice='file', insert=self.file_IT,
                                                           directory=self.T_IzFolder))

        self.run_krg2.clicked.connect(lambda: self.start())
        self.b_main_choice.clicked.connect(lambda: self.hide_show((gui_edit,), (gui_choice_window,)))
        # Подсказки
        self.T_IzFolder.setToolTip("Для корректировки файлов во всех вложенных папках нужно в конце поставить *")
        # Загрузить из .ini начальный путь для T_IzFolder
        self.T_IzFolder.setPlainText(GeneralSettings.read_ini(section='save_form_folder_edit', key="path"))

    def import_name_table(self):
        """
        Сформировать строку имени CB_ImpRg2 для наглядности выбранных вкладок.
        """
        add_str = 'Импорт из файлa (.rg2)'
        for CB, name in self.check_import:
            if CB.isChecked():
                add_str += f', {name}'
        self.CB_ImpRg2.setText(add_str)

    def task_save_yaml(self):
        name_file_save = self.save_file(directory=self.T_IzFolder.toPlainText(), filter_="YAML Files (*.yaml)")
        if name_file_save:
            self.fill_task_ui()
            with open(name_file_save, 'w') as f:
                yaml.dump(data=self.task_ui, stream=f, default_flow_style=False, sort_keys=False)

    def task_load_yaml(self):
        name_file_load = self.choice_file(directory=self.T_IzFolder.toPlainText().replace('*', ''),
                                          filter_="YAML Files (*.yaml)")
        if not name_file_load:
            return
        with open(name_file_load) as f:
            task_yaml = yaml.safe_load(f)
        if not task_yaml:
            return

        self.T_IzFolder.setPlainText(task_yaml["KIzFolder"])
        self.T_InFolder.setPlainText(task_yaml["KInFolder"])

        self.CB_KFilter_file.setChecked(task_yaml["KFilter_file"])  # QCheckBox
        self.D_count_file.setValue(task_yaml["max_file_count"])  # QSpainBox
        self.condition_file_years.setText(task_yaml["cor_criterion_start"]["years"])  # QLineEdit text()
        self.condition_file_season.setCurrentText(task_yaml["cor_criterion_start"]["season"])  # QComboBox
        self.condition_file_max_min.setCurrentText(task_yaml["cor_criterion_start"]["max_min"])
        self.condition_file_add_name.setText(task_yaml["cor_criterion_start"]["add_name"])

        self.CB_cor_b.setChecked(task_yaml["cor_beginning_qt"]['add'])
        self.TE_cor_b.setPlainText(task_yaml["cor_beginning_qt"]['txt'])

        self.CB_import_val_XL.setChecked(task_yaml["import_val_XL"])
        self.T_PQN_XL_File.setPlainText(task_yaml["excel_cor_file"])
        self.T_PQN_Sheets.setText(task_yaml["excel_cor_sheet"])

        self.CB_kontrol_rg2.setChecked(task_yaml["checking_parameters_rg2"])
        self.CB_U.setChecked(task_yaml["control_rg2_task"]['node'])
        self.CB_I.setChecked(task_yaml["control_rg2_task"]['vetv'])
        self.CB_gen.setChecked(task_yaml["control_rg2_task"]['Gen'])
        self.kontrol_rg2_Sel.setText(task_yaml["control_rg2_task"]['sel_node'])

        self.CB_printXL.setChecked(task_yaml["printXL"])
        self.CB_print_sech.setChecked(task_yaml['set_printXL']["sechen"]['add'])
        self.CB_print_area.setChecked(task_yaml['set_printXL']["area"]['add'])
        self.CB_print_area2.setChecked(task_yaml['set_printXL']["area2"]['add'])
        self.CB_print_darea.setChecked(task_yaml['set_printXL']["darea"]['add'])
        for key in task_yaml['set_printXL']:
            if key not in ["sechen", "area", "area2", "darea"]:
                self.CB_print_tab_log.setChecked(task_yaml['set_printXL'][key]['add'])
                self.print_tab_log_ar_set.setText(task_yaml['set_printXL'][key]["sel"])
                self.print_tab_log_ar_cols.setText(task_yaml['set_printXL'][key]['par'])
                self.print_tab_log_rows.setText(task_yaml['set_printXL'][key]['rows'])
                self.print_tab_log_cols.setText(task_yaml['set_printXL'][key]['columns'])
                self.print_tab_log_vals.setText(task_yaml['set_printXL'][key]['values'])
                break

        self.CB_print_parametr.setChecked(task_yaml['print_parameters']['add'])
        self.TA_parametr_vibor.setPlainText(task_yaml['print_parameters']['sel'])

        self.CB_print_balance_Q.setChecked(task_yaml['print_balance_q']['add'])
        self.balance_Q_vibor.setText(task_yaml['print_balance_q']['sel'])

        self.CB_ImpRg2.setChecked(task_yaml['CB_ImpRg2'])
        self.CB_ImpRg2.setText(task_yaml['CB_ImpRg2_name'])

        dict_ = task_yaml['Imp_add']['node']
        self.CB_N.setChecked(dict_['add'])
        self.file_N.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_N.setChecked(dict_['selection'])
        self.Filtr_god_N.setText(dict_["years"])
        self.Filtr_sez_N.setCurrentText(dict_["season"])
        self.Filtr_max_min_N.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_N.setText(dict_["add_name"])
        self.tab_N.setText(dict_['tables'])
        self.param_N.setText(dict_['param'])
        self.sel_N.setText(dict_['sel'])
        self.tip_N.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['vetv']
        self.CB_V.setChecked(dict_['add'])
        self.file_V.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_V.setChecked(dict_['selection'])
        self.Filtr_god_V.setText(dict_["years"])
        self.Filtr_sez_V.setCurrentText(dict_["season"])
        self.Filtr_max_min_V.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_V.setText(dict_["add_name"])
        self.tab_V.setText(dict_['tables'])
        self.param_V.setText(dict_['param'])
        self.sel_V.setText(dict_['sel'])
        self.tip_V.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['gen']
        self.CB_G.setChecked(dict_['add'])
        self.file_G.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_G.setChecked(dict_['selection'])
        self.Filtr_god_G.setText(dict_["years"])
        self.Filtr_sez_G.setCurrentText(dict_["season"])
        self.Filtr_max_min_G.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_G.setText(dict_["add_name"])
        self.tab_G.setText(dict_['tables'])
        self.param_G.setText(dict_['param'])
        self.sel_G.setText(dict_['sel'])
        self.tip_G.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['area']
        self.CB_A.setChecked(dict_['add'])
        self.file_A.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_A.setChecked(dict_['selection'])
        self.Filtr_god_A.setText(dict_["years"])
        self.Filtr_sez_A.setCurrentText(dict_["season"])
        self.Filtr_max_min_A.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_A.setText(dict_["add_name"])
        self.tab_A.setText(dict_['tables'])
        self.param_A.setText(dict_['param'])
        self.sel_A.setText(dict_['sel'])
        self.tip_A.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['area2']
        self.CB_A2.setChecked(dict_['add'])
        self.file_A2.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_A2.setChecked(dict_['selection'])
        self.Filtr_god_A2.setText(dict_["years"])
        self.Filtr_sez_A2.setCurrentText(dict_["season"])
        self.Filtr_max_min_A2.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_A2.setText(dict_["add_name"])
        self.tab_A2.setText(dict_['tables'])
        self.param_A2.setText(dict_['param'])
        self.sel_A2.setText(dict_['sel'])
        self.tip_A2.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['darea']
        self.CB_D.setChecked(dict_['add'])
        if 'selection' in dict_:
            self.CB_Filtr_D.setChecked(dict_['selection'])
        self.file_D.setText(dict_['import_file_name'])
        self.Filtr_god_D.setText(dict_["years"])
        self.Filtr_sez_D.setCurrentText(dict_["season"])
        self.Filtr_max_min_D.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_D.setText(dict_["add_name"])
        self.tab_D.setText(dict_['tables'])
        self.param_D.setText(dict_['param'])
        self.sel_D.setText(dict_['sel'])
        self.tip_D.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['PQ']
        self.CB_PQ.setChecked(dict_['add'])
        self.file_PQ.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_PQ.setChecked(dict_['selection'])
        self.Filtr_god_PQ.setText(dict_["years"])
        self.Filtr_sez_PQ.setCurrentText(dict_["season"])
        self.Filtr_max_min_PQ.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_PQ.setText(dict_["add_name"])
        self.tab_PQ.setText(dict_['tables'])
        self.param_PQ.setText(dict_['param'])
        self.sel_PQ.setText(dict_['sel'])
        self.tip_PQ.setCurrentText(dict_['calc'])

        dict_ = task_yaml['Imp_add']['IT']
        self.CB_IT.setChecked(dict_['add'])
        self.file_IT.setText(dict_['import_file_name'])
        if 'selection' in dict_:
            self.CB_Filtr_IT.setChecked(dict_['selection'])
        self.Filtr_god_IT.setText(dict_["years"])
        self.Filtr_sez_IT.setCurrentText(dict_["season"])
        self.Filtr_max_min_IT.setCurrentText(dict_["max_min"])
        self.Filtr_dop_name_IT.setText(dict_["add_name"])
        self.tab_IT.setText(dict_['tables'])
        self.param_IT.setText(dict_['param'])
        self.sel_IT.setText(dict_['sel'])
        self.tip_IT.setCurrentText(dict_['calc'])

        self.check_status(self.check_status_visibility)

    def start(self):
        """
        Запуск корректировки моделей.
        """
        GeneralSettings.write_ini(section='save_form_folder_edit', key="path", value=self.T_IzFolder.toPlainText())
        if self.print_tab_log_ar_tab.text() in ['area', 'area2', 'darea', 'sechen']:
            raise ValueError('В поле таблица на выбор нельзя задавать таблицы: area, area2, darea, sechen.')
        self.fill_task_ui()
        global em
        em = EditModel(self.task_ui)
        self.gui_import()
        em.run_cor()

    def gui_import(self):
        """
        Добавление в ImportFromModel данных с формы.
        """
        if self.CB_ImpRg2.isChecked():
            for tables in self.task_ui['Imp_add']:
                if self.task_ui['Imp_add'][tables]['add']:
                    criterion_start = {}
                    if self.task_ui['Imp_add'][tables]['selection']:
                        criterion_start = {"years": self.task_ui['Imp_add'][tables]['years'],
                                           "season": self.task_ui['Imp_add'][tables]['season'],
                                           "max_min": self.task_ui['Imp_add'][tables]['max_min'],
                                           "add_name": self.task_ui['Imp_add'][tables]['add_name']}

                    ifm = ImportFromModel(import_file_name=self.task_ui['Imp_add'][tables]['import_file_name'],
                                          criterion_start=criterion_start,
                                          tables=self.task_ui['Imp_add'][tables]['tables'],
                                          param=self.task_ui['Imp_add'][tables]['param'],
                                          sel=self.task_ui['Imp_add'][tables]['sel'],
                                          calc=self.task_ui['Imp_add'][tables]['calc'])
                    ImportFromModel.set_import_model.append(ifm)

    def fill_task_ui(self):
        """
        Заполнить task_ui задание взяв данные с формы QT.
        """
        self.task_ui = {
            "KIzFolder": self.T_IzFolder.toPlainText(),  # QPlainTextEdit
            "KInFolder": self.T_InFolder.toPlainText(),
            # Выборка файлов.
            "KFilter_file": self.CB_KFilter_file.isChecked(),  # QCheckBox
            "max_file_count": self.D_count_file.value(),  # QSpainBox
            "cor_criterion_start": {"years": self.condition_file_years.text(),  # QLineEdit text()
                                    "season": self.condition_file_season.currentText(),  # QComboBox
                                    "max_min": self.condition_file_max_min.currentText(),
                                    "add_name": self.condition_file_add_name.text()},
            # Корректировка в начале.
            "cor_beginning_qt": {'add': self.CB_cor_b.isChecked(),
                                 'txt': self.TE_cor_b.toPlainText()},
            # Задание из 'EXCEL'
            "import_val_XL": self.CB_import_val_XL.isChecked(),
            "excel_cor_file": self.T_PQN_XL_File.toPlainText(),
            "excel_cor_sheet": self.T_PQN_Sheets.text(),

            # Расчет режима и контроль параметров режима
            "checking_parameters_rg2": self.CB_kontrol_rg2.isChecked(),
            "control_rg2_task": {'node': self.CB_U.isChecked(),
                                 'vetv': self.CB_I.isChecked(),
                                 'Gen': self.CB_gen.isChecked(),
                                 'sel_node': self.kontrol_rg2_Sel.text()},
            # Выводить данные из моделей в XL
            "printXL": self.CB_printXL.isChecked(),
            "set_printXL": {
                "sechen": {'add': self.CB_print_sech.isChecked(),
                           "sel": 'ns>0',
                           'par': '',  # "ns,name,pmin,pmax,psech",
                           "rows": "ns,name",  # поля строк в сводной
                           "columns": "Год,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля столбцов в сводной
                           "values": "psech,pmax,difference_p"},
                "area": {'add': self.CB_print_area.isChecked(),
                         "sel": 'na>0',
                         'par': '',  # 'na,name,no,pg,pn,pn_sum,dp,pop,set_pop,qn_sum,pg_max,pg_min,poq,qn,qg,dev_pop',
                         "rows": "na,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                         "columns": "Год",  # поля столбцов в сводной
                         "values": "pop,difference_p"},
                "area2": {'add': self.CB_print_area2.isChecked(),
                          "sel": 'npa>0',
                          'par': '',  # 'npa,name,pg,pn,dp,pop,vnp,qg,qn,dq,poq,vnq,pn_sum,qn_sum,set_pop,dev_pop',
                          "rows": "npa,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                          "columns": "Год",  # поля столбцов в сводной
                          "values": "pop,difference_p"},
                "darea": {'add': self.CB_print_darea.isChecked(),
                          "sel": 'no>0',
                          'par': '',  # 'no,name,pg,pp,pvn,qn_sum,pnr_sum,pn_sum,set_pop,qvn,qp,qg,dev_pop',
                          "rows": "no,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                          "columns": "Год",  # поля столбцов в сводной
                          "values": "pp,difference_p"},
                self.print_tab_log_ar_tab.text(): {'add': self.CB_print_tab_log.isChecked(),
                                                   "sel": self.print_tab_log_ar_set.text(),
                                                   'par': self.print_tab_log_ar_cols.text(),
                                                   "rows": self.print_tab_log_rows.text(),  # поля строк в сводной
                                                   "columns": self.print_tab_log_cols.text(),  # поля столбцов в сводной
                                                   "values": self.print_tab_log_vals.text()}},  # поля значений в свод
            "print_parameters": {'add': self.CB_print_parametr.isChecked(),
                                 "sel": self.TA_parametr_vibor.toPlainText()},
            "print_balance_q": {'add': self.CB_print_balance_Q.isChecked(),
                                "sel": self.balance_Q_vibor.text()},
            # только для UI
            'CB_ImpRg2_name': self.CB_ImpRg2.text(),
            'CB_ImpRg2': self.CB_ImpRg2.isChecked(),
            'Imp_add': {
                'node': {'add': self.CB_N.isChecked(),
                         'import_file_name': self.file_N.text(),
                         "selection": self.CB_Filtr_N.isChecked(),
                         "years": self.Filtr_god_N.text(),
                         "season": self.Filtr_sez_N.currentText(),
                         "max_min": self.Filtr_max_min_N.currentText(),
                         "add_name": self.Filtr_dop_name_N.text(),
                         'tables': self.tab_N.text(),
                         'param': self.param_N.text(),
                         'sel': self.sel_N.text(),
                         'calc': self.tip_N.currentText(), },
                'vetv': {'add': self.CB_V.isChecked(),
                         'import_file_name': self.file_V.text(),
                         "selection": self.CB_Filtr_V.isChecked(),
                         "years": self.Filtr_god_V.text(),
                         "season": self.Filtr_sez_V.currentText(),
                         "max_min": self.Filtr_max_min_V.currentText(),
                         "add_name": self.Filtr_dop_name_V.text(),
                         'tables': self.tab_V.text(),
                         'param': self.param_V.text(),
                         'sel': self.sel_V.text(),
                         'calc': self.tip_V.currentText(), },
                'gen': {'add': self.CB_G.isChecked(),
                        'import_file_name': self.file_G.text(),
                        "selection": self.CB_Filtr_G.isChecked(),
                        "years": self.Filtr_god_G.text(),
                        "season": self.Filtr_sez_G.currentText(),
                        "max_min": self.Filtr_max_min_G.currentText(),
                        "add_name": self.Filtr_dop_name_G.text(),
                        'tables': self.tab_G.text(),
                        'param': self.param_G.text(),
                        'sel': self.sel_G.text(),
                        'calc': self.tip_G.currentText(), },
                'area': {'add': self.CB_A.isChecked(),
                         'import_file_name': self.file_A.text(),
                         "selection": self.CB_Filtr_A.isChecked(),
                         "years": self.Filtr_god_A.text(),
                         "season": self.Filtr_sez_A.currentText(),
                         "max_min": self.Filtr_max_min_A.currentText(),
                         "add_name": self.Filtr_dop_name_A.text(),
                         'tables': self.tab_A.text(),
                         'param': self.param_A.text(),
                         'sel': self.sel_A.text(),
                         'calc': self.tip_A.currentText(), },
                'area2': {'add': self.CB_A2.isChecked(),
                          'import_file_name': self.file_A2.text(),
                          "selection": self.CB_Filtr_A2.isChecked(),
                          "years": self.Filtr_god_A2.text(),
                          "season": self.Filtr_sez_A2.currentText(),
                          "max_min": self.Filtr_max_min_A2.currentText(),
                          "add_name": self.Filtr_dop_name_A2.text(),
                          'tables': self.tab_A2.text(),
                          'param': self.param_A2.text(),
                          'sel': self.sel_A2.text(),
                          'calc': self.tip_A2.currentText(), },
                'darea': {'add': self.CB_D.isChecked(),
                          'import_file_name': self.file_D.text(),
                          "selection": self.CB_Filtr_D.isChecked(),
                          "years": self.Filtr_god_D.text(),
                          "season": self.Filtr_sez_D.currentText(),
                          "max_min": self.Filtr_max_min_D.currentText(),
                          "add_name": self.Filtr_dop_name_D.text(),
                          'tables': self.tab_D.text(),
                          'param': self.param_D.text(),
                          'sel': self.sel_D.text(),
                          'calc': self.tip_D.currentText(), },
                'PQ': {'add': self.CB_PQ.isChecked(),
                       'import_file_name': self.file_PQ.text(),
                       "selection": self.CB_Filtr_PQ.isChecked(),
                       "years": self.Filtr_god_PQ.text(),
                       "season": self.Filtr_sez_PQ.currentText(),
                       "max_min": self.Filtr_max_min_PQ.currentText(),
                       "add_name": self.Filtr_dop_name_PQ.text(),
                       'tables': self.tab_PQ.text(),
                       'param': self.param_PQ.text(),
                       'sel': self.sel_PQ.text(),
                       'calc': self.tip_PQ.currentText(), },
                'IT': {'add': self.CB_IT.isChecked(),
                       'import_file_name': self.file_IT.text(),
                       "selection": self.CB_Filtr_IT.isChecked(),
                       "years": self.Filtr_god_IT.text(),
                       "season": self.Filtr_sez_IT.currentText(),
                       "max_min": self.Filtr_max_min_IT.currentText(),
                       "add_name": self.Filtr_dop_name_IT.text(),
                       'tables': self.tab_IT.text(),
                       'param': self.param_IT.text(),
                       'sel': self.sel_IT.text(),
                       'calc': self.tip_IT.currentText(), },
            }
        }


class GeneralSettings(ABC):
    """
    Для хранения общих настроек.
    """
    # коллекция настроек, которые хранятся в ini файле
    set_save = {}
    ini = 'settings.ini'
    log_file = 'log_file.log'

    # @abstractmethod
    def __init__(self):
        # коллекция для хранения информации о расчете
        self.set_info = {"calc_val": {1: "ЗАМЕНИТЬ", 2: "ПРИБАВИТЬ", 3: "ВЫЧЕСТЬ", 0: "УМНОЖИТЬ"},
                         'collapse': '', 'end_info': ''}
        self.all_read_ini()

        self.file_count = 0  # Счетчик расчетных файлов.

        self.now = datetime.now()
        self.time_start = time.time()
        self.now_start = self.now.strftime("%d-%m-%Y %H:%M:%S")

    def all_read_ini(self):
        # Прочитать ini файл
        if os.path.exists(self.ini):
            config = configparser.ConfigParser()
            config.read(self.ini)
            try:
                for key in config['DEFAULT']:
                    self.set_save[key] = config['DEFAULT'][key]
                for key in config['CalcSetWindow']:
                    if config['CalcSetWindow'][key] in ['True', 'False']:
                        self.set_save[key] = config['CalcSetWindow'].getboolean(key)
                    else:
                        self.set_save[key] = config['CalcSetWindow'][key]
            except LookupError:
                raise LookupError('файл settings.ini не читается')
        else:
            raise LookupError("Отсутствует файл settings.ini")

    def the_end(self):  # по завершению
        execution_time = time.strftime("%H:%M:%S", time.gmtime(time.time() - self.time_start))
        self.set_info['end_info'] = (
            f"РАСЧЕТ ЗАКОНЧЕН! \nНачало расчета {self.now_start}, конец {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
            f" \nЗатрачено: {execution_time} (файлов: {self.file_count}).")
        log.info(self.set_info['end_info'])

    @staticmethod
    def read_ini(section: str, key: str):
        """
        Прочитать в файле settings.ini значение в разделе section по ключу key.
        """
        if os.path.exists(GeneralSettings.ini):
            config = configparser.ConfigParser()
            config.read(GeneralSettings.ini)
            try:
                return config[section][key]
            except LookupError:
                log.error(f'В файле {GeneralSettings.ini!r} не найден разделе {section!r} или ключ {key!r}.')

    @staticmethod
    def write_ini(section: str, key: str, value):
        """
        Записать в файл settings.ini значение value в раздел section по ключу key.
        """
        config = configparser.ConfigParser()
        config.read(GeneralSettings.ini)
        config[section] = {key: value}
        with open(GeneralSettings.ini, 'w') as configfile:
            config.write(configfile)

    @staticmethod
    def split_task_action(txt: str) -> Union[list | bool]:
        """
        Разделить строку по запятым, если запятая не внутри [] {}
        :param txt: [15114,15011,0:sta=1],[15011,15012:sta=0]{15114,15011:sta==1},[15105,15131:sta=1],6
        :return: [[15114,15011,0:sta=1],[15011,15012:sta=0]{15114,15011:sta==1},[15105,15131:sta=1],6]
        или  False
        """
        if not txt:
            return False
        # вычленить значения в [ ] и { }
        actions = re.findall(re.compile(r"\[(.+?)]"), txt)
        conditions = re.findall(re.compile(r"\{(.+?)}"), txt)

        # заменить значения в [ ] и { } на act_cond_{n}
        dict_key = {}  # замена, действие
        for n, action in enumerate(actions + conditions):
            dict_key[f'act_cond_{n}'] = action
            txt = txt.replace(action, f'act_cond_{n}')

        #  заменить act_cond_{n} на значения в [ ] и { }
        result = []
        for part in txt.split(','):
            for key in dict_key:
                if key in part:
                    part = part.replace(key, dict_key[key])
            result.append(part)
        return result

    @staticmethod
    def read_title(txt: str) -> tuple:
        """
        Разделить строку типа 'Рисунок [1] - Южный'
        :param txt:
        :return: (1, ['Рисунок ', ' - Южный'])
        """
        txt = txt.strip()
        num = txt[txt.find('[') + 1: txt.find(']')]
        txt = txt.split(f'[{num}]')
        num = int(num) if num.isdigit() else 1
        return num, txt


class CalcModel(GeneralSettings):
    """
    Расчет нормативных возмущений.
    """
    def __init__(self, task_calc):
        super(CalcModel, self).__init__()
        self.all_read_ini()
        self.number_comb = 0
        self.task_calc = task_calc
        self.all_folder = False  # Не перебирать вложенные папки
        self.set_comb = {}  # {количество отключений: контроль ДТН, 1:"ДДТН",2:"АДТН"}
        # self.auto_shunt = {}

        self.control_I = None
        self.control_U = None
        self.restore_only_state = True

        # DF для хранения токовых перегрузок и недопустимого снижения U

        self.srs_xl = pd.DataFrame()  # Перечень отключений их excel
        self.overloads_all = pd.DataFrame()  # общий
        self.overloads_srs = pd.DataFrame()  # СРС перегрузки
        self.info_srs = pd.Series(dtype='object')  # СРС
        self.info_action = pd.Series(dtype='object')  # действие ПА

        self.book_path: str = ''  # Путь к файлу excel.
        self.pa = None  # Объект Automation
        self.task_full_name = ''  # Путь к файлу задания rg2.

        # Для хранения имен файлов с рисунками и имен рисунков
        self.df_picture = pd.DataFrame(dtype='str', columns=['Наименование файла',
                                                             'Наименование рисунка'])
        self.num_pic, self.name_pic = list(GeneralSettings.read_title(self.task_calc['name_pic']))

    def run_calc(self):
        """
        Запуск расчета нормативных возмущений (НВ) в РМ.
        """
        test_run('calc')
        log.info('Запуск расчета нормативных возмущений (НВ) в расчетной модели (РМ).')
        if "*" in self.task_calc["calc_folder"]:
            self.task_calc["calc_folder"] = self.task_calc["calc_folder"].replace('*', '')
            self.all_folder = True

        if not os.path.exists(self.task_calc["calc_folder"]):
            raise ValueError(f'Не найден путь: {self.task_calc["calc_folder"]}.')

        # папка для сохранения результатов
        self.task_calc['folder_result_calc'] = self.task_calc["calc_folder"] + r"\result"
        if os.path.isfile(self.task_calc["calc_folder"]):
            self.task_calc['folder_result_calc'] = os.path.dirname(self.task_calc["calc_folder"]) + r"\result"
        if not os.path.exists(self.task_calc['folder_result_calc']):
            os.mkdir(self.task_calc['folder_result_calc'])  # создать папку result

        self.task_calc['name_time'] = f"{self.task_calc['folder_result_calc']}" \
                                      f"\\{datetime.now().strftime('%d-%m-%Y %H-%M-%S')}"

        if self.task_calc['cb_disable_excel']:
            self.srs_xl = pd.read_excel(self.task_calc['srs_XL_path'], sheet_name=self.task_calc['srs_XL_sheets'])

            self.srs_xl = self.srs_xl[self.srs_xl['Статус'] != '-']
            self.srs_xl.drop(columns=['Примечание', 'Статус'], inplace=True)
            self.srs_xl.dropna(how='all', axis=0, inplace=True)
            self.srs_xl.dropna(how='all', axis=1, inplace=True)
            for col in self.srs_xl.columns:
                self.srs_xl[col] = self.srs_xl[col].str.split('#').str[0]
            self.srs_xl.fillna(0, inplace=True)

        # Цикл, если несколько файлов задания.
        if self.task_calc['CB_Import_Rg2'] and os.path.isdir(self.task_calc["Import_file"]):
            task_files = os.listdir(self.task_calc["Import_file"])
            task_files = list(filter(lambda x: x.endswith('.rg2'), task_files))
            for task_file in task_files:  # цикл по файлам '.rg2' в папке
                self.task_full_name = os.path.join(self.task_calc["Import_file"], task_file)
                log.info(f'Текущий файл задания: {self.task_full_name}')
                self.run_calc_task()
        else:
            if self.task_calc['CB_Import_Rg2']:
                self.task_full_name = self.task_calc['Import_file']
                self.run_calc_task()
            else:
                self.run_calc_task()

        self.the_end()
        notepad_path = f'{self.task_calc["name_time"]} протокол расчета РМ.log'
        shutil.copyfile(self.log_file, notepad_path)
        with open(self.task_calc['name_time'] + ' задание на расчет РМ.yaml', 'w') as f:
            yaml.dump(data=self.task_calc, stream=f, default_flow_style=False, sort_keys=False)
        # webbrowser.open(notepad_path)  #  Открыть блокнотом лог-файл.
        mb.showinfo("Инфо", self.set_info['end_info'])

    @staticmethod
    def gen_comb_xl(rm, df: pd.DataFrame) -> pd.DataFrame:
        """
        Генератор комбинаций из XL
        :param rm:
        :param df:
        :return:  комбинацию comb_xl
        """
        for _, row in df.iterrows():  # todo может можно заменить на Itertuples
            raise ValueError(' todo добавить s_key')
            comb_xl = pd.DataFrame(columns=['table',
                                            'index',
                                            'dname',
                                            'status_repair',
                                            'key',  # todo добавить s_key
                                            'repair_scheme',
                                            'disable_scheme'])

            if row['Ключ откл.']:
                table = rm.name_table_from_key(row['Ключ откл.'])
                index = rm.index(table_name=table, key_str=row['Ключ откл.'])
                if table and index >= 0:
                    scheme = rm.rastr.tables(table).cols.item("disable_scheme").Z(index)
                    if 'Схема при отключении' in row and row['Схема при отключении']:
                        scheme += ',' + row['Схема при отключении']
                    comb_xl.loc[len(comb_xl.index)] = [table,  # 'table'
                                                       index,  # 'index'
                                                       row['Отключение'],  # 'dname'
                                                       False,  # 'status_repair'
                                                       row['Ключ откл.'],  # 'key'
                                                       0,  # 'repair_scheme'
                                                       scheme]  # 'disable_scheme'
                else:
                    log.info(f'Задание комбинаций их XL: {row["Отключение"]!r} не найден ключ {row["Ключ откл."]!r}')
                    continue

            if row['Ключ рем.1']:
                table = rm.name_table_from_key(row['Ключ рем.1'])
                index = rm.index(table_name=table, key_str=row['Ключ рем.1'])
                if table and index >= 0:
                    scheme = rm.rastr.tables(table).cols.item("repair_scheme").Z(index)
                    if 'Ремонтная схема' in row and row['Ремонтная схема']:
                        scheme += ',' + row['Ремонтная схема']
                    comb_xl.loc[len(comb_xl.index)] = [table,  # 'table'
                                                       index,  # 'index'
                                                       row['Ремонт 1'],  # 'dname'
                                                       True,  # 'status_repair'
                                                       row['Ключ рем.1'],  # 'key'
                                                       0,  # 'repair_scheme'
                                                       scheme]  # 'disable_scheme'
                else:
                    log.info(f'Задание комбинаций их XL: {row["Ремонт 1"]!r} не найден ключ {row["Ключ рем.1"]!r}')
                    continue

            if row['Ключ рем.2']:
                table = rm.name_table_from_key(row['Ключ рем.2'])
                index = rm.index(table_name=table, key_str=row['Ключ рем.2'])
                if table and index >= 0:
                    comb_xl.loc[len(comb_xl.index)] = [table,  # 'table'
                                                       index,  # 'index'
                                                       row['Ремонт 2'],  # 'dname'
                                                       True,  # 'status_repair'
                                                       row['Ключ рем.2'],  # 'key'
                                                       0,  # 'repair_scheme'
                                                       rm.rastr.tables(table).cols.item("repair_scheme").Z(index)]
                else:
                    log.info(f'Задание комбинаций их XL: {row["Ремонт 2"]!r} не найден ключ {row["Ключ рем.2"]!r}')
                    continue
            yield comb_xl

    def run_calc_task(self):
        """
        Запуск расчета с текущим файлом импорта задания или без него.
        """
        xlApp = None

        if os.path.isdir(self.task_calc["calc_folder"]):
            if self.all_folder:  # с вложенными папками
                for address, dir_, file_ in os.walk(self.task_calc["calc_folder"]):
                    self.for_file(folder_calc=address)
            else:  # без вложенных папок
                self.for_file(folder_calc=self.task_calc["calc_folder"])

        elif os.path.isfile(self.task_calc["calc_folder"]):
            rm = RastrModel(self.task_calc["calc_folder"])
            if not rm.code_name_rg2:
                raise ValueError(f'Имя файла {self.task_calc["calc_folder"]!r} не подходит.')
            self.calc_file(rm=rm)

        # Сохранить в Excel таблицы перегрузки.
        if len(self.overloads_all):
            # https://www.geeksforgeeks.org/how-to-write-pandas-dataframes-to-multiple-excel-sheets/
            log.debug(f'Запись перегрузок в excel ({len(self.overloads_all)}строк)')
            mode = 'a' if os.path.exists(self.book_path) else 'w'
            with pd.ExcelWriter(path=self.book_path, mode=mode) as writer:
                for col in ["Отключение", "Ремонт 1", "Ремонт 2", "Доп. имя"]:
                    for col_df in self.overloads_all.columns:
                        if col in col_df:
                            self.overloads_all.fillna(value={col_df: 0}, inplace=True)
                            self.overloads_all.loc[self.overloads_all[col_df] == 0, col_df] = '-'
                self.overloads_all.rename(columns={'dname': 'Контролируемые элементы'}, inplace=True)
                self.overloads_all.to_excel(excel_writer=writer,
                                            float_format="%.2f",
                                            index_label='Номер сочетания.подсочетания',
                                            freeze_panes=(1, 1),
                                            sheet_name='Перегрузки')
        # Сохранить в Excel таблицы перегрузки.
        sheet_name_pic = 'Рисунки'
        if len(self.df_picture):
            with pd.ExcelWriter(path=self.book_path,
                                mode='a' if os.path.exists(self.book_path) else 'w') as writer:
                self.df_picture.to_excel(excel_writer=writer,
                                         startrow=1,
                                         index=False,
                                         freeze_panes=(5, 1),
                                         sheet_name=sheet_name_pic)

            book = load_workbook(self.book_path)
            sheet_pic = book[sheet_name_pic]

            sheet_pic.insert_rows(1, amount=3)

            sheet_pic['A1'] = 'Формат листа (3 - А3, 4 - А4):'
            sheet_pic['A2'] = 'Ориентация(1 - книжная, 0 - альбомная):'
            sheet_pic['A3'] = 'Имя папки с файлами rg2:'
            sheet_pic['B1'] = 3
            sheet_pic['B2'] = 1
            sheet_pic['B3'] = self.task_calc['folder_result_calc']
            thins = Side(border_style="thin", color="000000")
            for col in ['A', 'B']:
                sheet_pic.column_dimensions[col].width = 100
                for r in ['1', '2', '3']:
                    sheet_pic[col + r].alignment = Alignment(horizontal='left')
                    sheet_pic[col + r].border = Border(thins, thins, thins, thins)
                    sheet_pic[col + r].fill = PatternFill(fill_type='solid', fgColor="00B1E76E")
            PrintXL.create_table(sheet=sheet_pic,
                                 sheet_name=sheet_name_pic,
                                 point_start='A5')
            book.save(self.book_path)

        # Сводная
        if len(self.overloads_all):
            log.info(f'Формируется сводная таблица ({self.book_path}).')
            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.ScreenUpdating = False  # Обновление экрана
            try:
                book = xlApp.Workbooks.Open(self.book_path)
            except Exception:
                raise Exception(f'Ошибка при открытии файла {self.book_path=}')
            try:
                sheet = book.sheets['Перегрузки']
            except Exception:
                raise Exception(f'Не найден лист Перегрузки')
            # Создать объект таблица из всего диапазона листа.
            tabl_overload = sheet.ListObjects.Add(SourceType=1, Source=sheet.Range(sheet.UsedRange.address))
            tabl_overload.Name = "Таблица_Перегрузки"
            pt_cache = book.PivotCaches().add(1, tabl_overload)  # Создать КЭШ xlDatabase, ListObjects

            task_pivot = []
            task_pivot_i = namedtuple('task_pivot',
                                      ['sheet_name', 'pivot_table_name', 'data_field'])
            # todo сводная если только режимы которые не моделируются
            if "i_max" in self.overloads_all.columns:
                task_pivot.append(task_pivot_i('Сводная_I', "Свод_I",
                                               dict(i_max="Iрасч.,A",
                                                    i_dop_r="Iддтн,A",
                                                    i_zag="Iзагр. ддтн,%",
                                                    i_dop_r_av="Iадтн,A",
                                                    i_zag_av="Iзагр. адтн,%")))
            if "umin" in self.overloads_all.columns:
                task_pivot.append(task_pivot_i('Сводная_Umin', "Свод_Umin",
                                               dict(vras="Uрасч.,кВ",
                                                    umin="Uмдн,кВ",
                                                    otv_min="Uмдн,%",
                                                    umin_av="Uaдн,кВ",
                                                    otv_min_av="Uaдн,%")))
            if "umax" in self.overloads_all.columns:
                task_pivot.append(task_pivot_i('Сводная_Umax', "Свод_Umax",
                                               dict(vras="Uрасч.,кВ",
                                                    umax="Uнаиб.раб.,кВ",
                                                    otv_max="Uнаиб.раб.,%")))

            RowFields = [col for col in ["Контролируемые элементы", "Отключение", "Ремонт 1", "Ремонт 2"]
                         if col in self.overloads_all.columns]

            ColumnFields = ["Год", "Сезон макс/мин"] + [col for col in self.overloads_all.columns if "Доп. имя" in col]

            for task in task_pivot:
                sheet_pivot = book.Sheets.Add()
                sheet_pivot.Name = task.sheet_name

                pt = pt_cache.CreatePivotTable(TableDestination=task.sheet_name + "!R1C1",
                                               TableName=task.pivot_table_name)
                pt.ManualUpdate = True  # True не обновить сводную
                pt.AddFields(RowFields=RowFields,
                             ColumnFields=ColumnFields,
                             PageFields=["Имя файла", 'Кол. откл. эл.', 'Конец', 'Наименование СРС', 'Контроль ДТН',
                                         'Темп.(°C)'],
                             AddToTable=False)
                for field_df, field_pt in task.data_field.items():
                    pt.AddDataField(Field=pt.PivotFields(field_df),
                                    Caption=field_pt,
                                    Function=-4136)  # xlMax -4136 xlSum -4157
                    pt.PivotFields(field_pt).NumberFormat = "0"

                pt.PivotFields("Контролируемые элементы").ShowDetail = True  # группировка
                pt.RowAxisLayout(1)  # 1 xlTabularRow показывать в табличной форме!!!!
                pt.DataPivotField.Orientation = 1  # xlRowField = 1 "Значения" в столбцах или строках xlColumnField
                pt.RowGrand = False  # Удалить строку общих итогов
                pt.ColumnGrand = False  # Удалить столбец общих итогов
                pt.MergeLabels = True  # Объединять одинаковые ячейки
                pt.HasAutoFormat = False  # Не обновлять ширину при обновлении
                pt.NullString = "--"  # Заменять пустые ячейки
                pt.PreserveFormatting = False  # Сохранять формат ячеек при обновлении
                pt.ShowDrillIndicators = False  # Показывать кнопки свертывания
                for row in RowFields + ColumnFields:
                    pt.PivotFields(row).Subtotals = [False, False, False, False, False, False, False, False, False,
                                                     False, False, False]  # промежуточные итоги и фильтры
                field = list(task.data_field)[2]
                pt.PivotFields(field).Orientation = 3  # xlPageField = 3
                pt.PivotFields(field).CurrentPage = "(All)"  #
                if len(task_pivot) > 1:
                    pt.PivotFields(field).PivotItems("(blank)").Visible = False
                pt.TableStyle2 = ""  # стиль
                pt.ColumnRange.ColumnWidth = 10  # ширина строк
                pt.RowRange.ColumnWidth = 20
                pt.DataBodyRange.HorizontalAlignment = -4108  # xlCenter = -4108
                pt.TableRange1.WrapText = True  # перенос текста в ячейке
                for i in range(7, 13):
                    pt.TableRange1.Borders(i).LineStyle = 1  # лево
                # Условное форматирование
                for i in range(3, len(task.data_field) + 1, 2):
                    dpz = pt.DataBodyRange.Rows(i).Cells(1)
                    dpz.FormatConditions.AddColorScale(2)  # ColorScaleType:=2
                    dpz.FormatConditions(dpz.FormatConditions.count).SetFirstPriority()
                    dpz.FormatConditions(1).ColorScaleCriteria(1).Type = 0  # xlConditionValueNumber = 0
                    if list(task.data_field)[2] == 'i_zag':
                        dpz.FormatConditions(1).ColorScaleCriteria(1).Value = 100
                    else:
                        dpz.FormatConditions(1).ColorScaleCriteria(1).Value = 0
                    dpz.FormatConditions(1).ColorScaleCriteria(1).FormatColor.ThemeColor = 1  # xlThemeColorDark1 = 1
                    dpz.FormatConditions(1).ColorScaleCriteria(1).FormatColor.TintAndShade = 0
                    dpz.FormatConditions(1).ColorScaleCriteria(2).Type = 2  # xlConditionValueHighestValue = 2
                    dpz.FormatConditions(1).ColorScaleCriteria(2).FormatColor.ThemeColor = 3 + i  # номер темы
                    dpz.FormatConditions(1).ColorScaleCriteria(2).FormatColor.TintAndShade = -0.249977111117893
                    dpz.FormatConditions(1).ScopeType = 2  # xlDataFieldScope = 2 применить ко всем значениям поля
                    pass
                pt.ManualUpdate = False  # обновить сводную
            book.Save()
            book.Close()
        else:
            log.info('Отклонений параметров режима от допустимых значений не выявлено.')

        # Вставить таблицы К-О в word.
        if self.task_calc['cb_tab_KO']:
            log.info('Вставить таблицы К-О в word.')

            xlApp = win32com.client.Dispatch("Excel.Application")
            xlApp.Visible = False
            book = xlApp.Workbooks.Open(self.book_path)

            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            word.ScreenUpdating = False
            doc = word.Documents.Add()  # doc = word.Documents.Open(r"I:\file.docx")

            doc.PageSetup.PageWidth = 29.7 * 28.35  # CentimetersToPoints( format_list_i (2) ) 1 см = 28,35
            doc.PageSetup.PageHeight = 42.0 * 28.35  # CentimetersToPoints( format_list_i (1) )
            doc.PageSetup.Orientation = 1  # 1 книжная или 0 альбомная

            cursor = word.Selection
            cursor.Font.Size = 12
            cursor.Font.Name = "Times New Roman"
            cursor.EndKey(Unit=6)  # перейти в конец текста

            for i in range(1, book.Worksheets.Count + 1):
                if book.Worksheets(i).name[:1].isnumeric():
                    sheet = book.Worksheets(i)
                    cursor.TypeText(Text=sheet.Cells(1, 1).value)
                    cursor.TypeParagraph()

                    sheet.Range(sheet.UsedRange.address.replace('$', '').replace('A1', 'A2')).Copy()
                    # cursor.PasteExcelTable(LinkedToExcel=False, WordFormatting=False, RTF=False)
                    cursor.PasteAndFormat(Type=13)  # 13 Вставить в виде рисунка.
                    cursor.InsertBreak(Type=0)
                    # разрыв:7 страницы с новой строки, 0-в той же строке,1 и 8 колонки,
                    # 2-5 раздела со след стр,6 и 9-11 перенос на новую стр

            word.ScreenUpdating = True
            doc.SaveAs2(FileName=self.task_calc['name_time'] + ' таблицы К-О.docx')  # FileFormat=16 .docx
            doc.Close()

        if xlApp:
            xlApp.Visible = True
            xlApp.ScreenUpdating = True  # обновление экрана

    def for_file(self, folder_calc: str):
        """
        Цикл по файлам.
        :param folder_calc:
        """
        files_calc = os.listdir(folder_calc)  # список всех файлов в папке
        rm_files = list(filter(lambda x: x.endswith('.rg2'), files_calc))

        for rastr_file in rm_files:  # цикл по файлам '.rg2' в папке
            if self.task_calc["Filter_file"] and self.file_count == self.task_calc["file_count_max"]:
                break  # Если включен фильтр файлов проверяем количество расчетных файлов.
            full_name = os.path.join(folder_calc, rastr_file)
            rm = RastrModel(full_name)
            # если включен фильтр файлов и имя стандартизовано
            if not rm.code_name_rg2:
                log.info(f'Имя файла {full_name} не подходит.')
                continue
            if self.task_calc["Filter_file"]:
                if not rm.test_name(condition=self.task_calc["calc_criterion"],
                                    info=f'Имя файла {full_name} не подходит.'):
                    continue  # пропускаем, если не соответствует фильтру
            self.calc_file(rm)

    def calc_file(self, rm):
        """
        Рассчитать РМ.
        """
        self.file_count += 1
        self.book_path = self.task_calc['name_time'] + ' результаты расчетов.xlsx'
        log.info("\n\n")
        rm.load()
        log.info(f"Расчетная температура: {rm.temperature}")
        rm.rastr.CalcIdop(rm.temperature, 0.0, "")
        if self.task_calc['cor_rm']['add']:
            rm.cor_rm_from_txt(self.task_calc['cor_rm']['txt'])

        # Импорт из РМ c ИД.
        if self.task_full_name:  # :task_full_name: полный путь к текущему файлу задания
            # "таблица: node, vetv; тип: 2; поле: disable_scheme, automation; выборка: sel"
            for row in self.task_calc['txt_Import_Rg2'].split('\n'):
                row = row.replace(' ', '').split('#')[0]  # удалить текст после '#'
                if row:
                    rm.txt_import_rm(type_import='файл', description=f'({self.task_full_name});{row}')

        # Подготовка.
        rm.voltage_fix_frame()
        # if self.set_save['skrm']:
        #     self.auto_shunt = rm.auto_shunt_rec(selection='')

        # Добавить поле index в таблицы.
        rm.fill_field_index('vetv,node,Generator')

        # Поля для сортировки ветвей и др.
        rm.add_fields_in_table(name_tables='vetv', fields='temp,temp1', type_fields=1)

        # Поля для контроля напряжений
        rm.add_fields_in_table(name_tables='node', fields='umin_av', type_fields=1)
        rm.add_fields_in_table(name_tables='node', fields='otv_min', type_fields=1,
                               prop=((5, 'if(sta=0) (-vras+umin)/umin*100:0'),),
                               replace=True)
        rm.add_fields_in_table(name_tables='node', fields='otv_min_av', type_fields=1,
                               prop=((5, 'if(sta=0) (-vras+umin_av)/umin_av*100:0'),),
                               replace=True)
        rm.add_fields_in_table(name_tables='node', fields='otv_max', type_fields=1,
                               prop=((5, 'if(sta=0) (vras-umax)/umax*100:0'),))
        # Поля для загрузки ветвей
        rm.add_fields_in_table(name_tables='vetv', fields='i_zag_av', type_fields=1,
                               prop=((5, 'if(ktr!=0) zag_it_av:zag_i_av'), (12, 1000),))

        # Поля для автоматики, что бы не было ошибок
        rm.add_fields_in_table(name_tables='vetv,node,Generator',
                               fields='repair_scheme,double_repair_scheme,disable_scheme,dname', type_fields=2)
        rm.add_fields_in_table(name_tables='vetv,node',
                               fields='automation', type_fields=2)
        # Поля с ключами таблиц
        rm.add_fields_in_table(name_tables='vetv', fields='key', type_fields=2,
                               prop=((5, '"ip="+str(ip)+"&iq="+str(iq)+"&np="+str(np)'),))
        rm.add_fields_in_table(name_tables='node', fields='key', type_fields=2,
                               prop=((5, '"ny="+str(ny)'),))
        rm.add_fields_in_table(name_tables='Generator', fields='key', type_fields=2,
                               prop=((5, '"Num="+str(Num)'),))

        # Сохранить текущее состояние РМ
        rm.save_value_fields()

        # if self.set_save["pa"]:
        self.pa = Automation(rm)

        # Контролируемые элементы сети.
        if self.task_calc['cb_control']:
            log.debug('Инициализация контролируемых элементов сети.')
            # all_control для отметки всех контролируемых узлов и ветвей (авто и field)
            rm.add_fields_in_table(name_tables='vetv,node', fields='all_control', type_fields=3)

            if self.task_calc['cb_control_field']:

                if '*' in self.task_calc['le_control_field']:
                    rm.rastr.Tables("node").cols.item("all_control").Calc("1")
                    rm.rastr.Tables("vetv").cols.item("all_control").Calc("1")
                else:
                    # Добавит поле отметки отключений, если их нет в какой-то таблице.
                    rm.add_fields_in_table(name_tables='vetv,node',
                                           fields=self.task_calc['le_control_field'],
                                           type_fields=3)
                    for table_name in ['vetv', 'node']:
                        rm.group_cor(tabl=table_name,
                                     param="all_control",
                                     selection=self.task_calc['le_control_field'],
                                     formula='1')

                    # all_control_groupid для отметки всех контролируемых ветвей и ветвей с теми же groupid
                    log.debug('Добавление в контролируемые элементы ветвей по groupid.')
                    if not self.task_calc['cb_tab_KO']:
                        rm.add_fields_in_table(name_tables='vetv', fields='all_control_groupid', type_fields=3)
                        rm.rastr.tables('vetv').cols.item("all_control_groupid").calc("all_control")

                        for gr in set(rm.df_from_table(table_name='vetv',
                                                       fields='groupid',
                                                       setsel='all_control & groupid>0')['groupid']):
                            rm.group_cor(tabl='vetv',
                                         param="all_control",
                                         selection=f"groupid={gr}",
                                         formula=1)

            if self.task_calc['cb_tab_KO']:
                log.debug('Инициализация таблицы "контролируемые - отключаемые" элементы.')
                rm.rastr.tables('vetv').cols.item("temp").calc('ip.uhom')
                rm.rastr.tables('vetv').cols.item("temp1").calc('iq.uhom')
                self.control_I = rm.df_from_table(table_name='vetv',
                                                  fields='index,dname,temp,temp1,i_dop_r,i_dop_r_av,groupid,key,tip',
                                                  # ip, iq, np, name
                                                  setsel="all_control")
                if len(self.control_I):
                    self.control_I['uhom'] = (self.control_I[['temp', 'temp1']].max(axis=1) * 10000 +
                                              self.control_I[['temp', 'temp1']].min(axis=1))
                    self.control_I.sort_values(by=['tip', 'uhom', 'dname'],  # столбцы сортировки
                                               ascending=(False, False, True),  # обратный порядок
                                               inplace=True)  # изменить df
                    self.control_I.drop(['temp', 'temp1', 'uhom', 'tip'], axis=1, inplace=True)
                    self.control_I['i_dop_r'] = self.control_I['i_dop_r'].round(0).astype(int)
                    self.control_I['i_dop_r_av'] = self.control_I['i_dop_r_av'].round(0).astype(int)
                    self.control_I.rename(columns={'i_dop_r': 'ДДТН, А',
                                                   'i_dop_r_av': 'АДТН, А',
                                                   'dname': 'Контролируемый элемент'}, inplace=True)
                    self.control_I.set_index('index', inplace=True)
                    self.control_I = self.control_I.T
                    self.control_I.index = pd.MultiIndex.from_product([['-'], ['-'], self.control_I.index])

                self.control_U = rm.df_from_table(table_name='node',
                                                  fields='index,dname,umin,umin_av,uhom',  # ,ny,umax
                                                  setsel="all_control")
                if len(self.control_U):
                    self.control_U.sort_values(by=['uhom', 'dname'],  # столбцы сортировки
                                               ascending=(False, True),  # обратный порядок
                                               inplace=True)  # изменить df
                    self.control_U.drop(['uhom'], axis=1, inplace=True)
                    self.control_U['umin'] = self.control_U['umin'].round(1)
                    self.control_U['umin_av'] = self.control_U['umin_av'].round(1)
                    self.control_U.rename(columns={'umin': 'МДН, кВ',
                                                   'umin_av': 'АДН, кВ',
                                                   'dname': 'Контролируемый элемент'}, inplace=True)
                    self.control_U.set_index('index', inplace=True)
                    self.control_U = self.control_U.T
                    self.control_U.index = pd.MultiIndex.from_product([['-'], ['-'], self.control_U.index])

        # Нормальная схема сети
        self.info_srs = pd.Series(dtype='object')  # СРС
        self.info_srs['Наименование СРС'] = 'Нормальная схема сети.'
        self.info_srs["Наименование СРС без()"] = 'Нормальная схема сети'
        self.info_srs['Номер СРС'] = self.number_comb
        self.info_srs['Кол. откл. эл.'] = 0
        self.info_srs['Контроль ДТН'] = 'ДДТН'
        log.info(f"Сочетание {self.number_comb}: {self.info_srs['Наименование СРС']}")
        self.do_action(rm)

        # Отключаемые элементы сети.
        if self.task_calc['cb_disable_comb']:
            # self.set_comb[0] = 'ДДТН'
            # Выбор количества одновременно отключаемых элементов
            if self.task_calc['SRS']['n-1']:
                self.set_comb[1] = 'ДДТН'
            if self.task_calc['SRS']['n-2']:
                self.set_comb[2] = 'ДДТН'
                if 0 < rm.code_name_rg2 < 4 and self.set_save['gost']:
                    self.set_comb[2] = 'AДТН'
            if self.task_calc['SRS']['n-3']:
                if self.set_save['gost']:
                    if rm.code_name_rg2 > 3:
                        self.set_comb[3] = 'АДТН'
                else:
                    self.set_comb[3] = 'ДДТН'
            log.info(f'Расчетные СРС: {self.set_comb}.')

            # В поле all_disable складываем элементы авто отмеченные и отмеченные в поле comb_field
            rm.add_fields_in_table(name_tables='vetv,node,Generator', fields='all_disable', type_fields=3)

            if self.task_calc['cb_auto_disable']:
                # Выбор отключаемых элементов автоматически из выборки в таблице узлы
                # Отметка в таблицах ветви и узлы нужное поле
                rm.network_analysis(field='all_disable',
                                    selection_node_for_disable=self.task_calc['auto_disable_choice'])

            # Выбор отключаемых элементов из отмеченных в поле comb_field
            if self.task_calc['cb_comb_field']:
                # Добавит поле отметки отключений, если их нет в какой-то таблице
                rm.add_fields_in_table(name_tables='vetv,node,Generator', fields=self.task_calc['comb_field'],
                                       type_fields=3)
                for table_name in ['vetv', 'node', 'Generator']:
                    rm.group_cor(tabl=table_name,
                                 param="all_disable",
                                 selection=self.task_calc['comb_field'],
                                 formula='1')

            # Создать df отключаемых узлов и ветвей и генераторов. Сортировка.
            columns_pa = ',repair_scheme,disable_scheme,double_repair_scheme'
            # Генераторы
            disable_df_gen = rm.df_from_table(table_name='Generator',
                                              fields='index,key,Num' + columns_pa,  # ,Num,NodeState,Node
                                              setsel="all_disable")
            disable_df_gen['table'] = 'Generator'
            disable_df_gen.rename(columns={'Num': 's_key'}, inplace=True)
            # Узлы
            disable_df_node = rm.df_from_table(table_name='node',
                                               fields='index,name,uhom,key,ny' + columns_pa,
                                               setsel="all_disable")
            # disable_df_node.index = self.disable_df_node['index']

            disable_df_node['table'] = 'node'
            disable_df_node.sort_values(by=['uhom', 'name'],  # столбцы сортировки
                                        ascending=(False, True),  # обратный порядок
                                        inplace=True)  # изменить df
            disable_df_node.drop(['name'], axis=1, inplace=True)
            disable_df_node.rename(columns={'ny': 's_key'}, inplace=True)
            # Ветви
            disable_df_vetv = rm.df_from_table(table_name='vetv',
                                               fields='index,name,key,temp,temp1,tip,ip,iq,np' + columns_pa,
                                               setsel="all_disable")
            disable_df_vetv['table'] = 'vetv'
            disable_df_vetv['uhom'] = disable_df_vetv[['temp', 'temp1']].max(axis=1) * 10000 + \
                                      disable_df_vetv[['temp', 'temp1']].min(axis=1)
            disable_df_vetv.sort_values(by=['tip', 'uhom', 'name'],  # столбцы сортировки
                                        ascending=(False, False, True),  # обратный порядок
                                        inplace=True)  # изменить df
            disable_df_vetv['s_key'] = None
            for i in disable_df_vetv.index:
                ip = disable_df_vetv['ip'].iloc[i]
                iq = disable_df_vetv['iq'].iloc[i]
                np_ = disable_df_vetv['np'].iloc[i]
                disable_df_vetv['s_key'].iloc[i] = (ip, iq, np_) if np_ else (ip, iq)

            disable_df_vetv.drop(['temp', 'temp1', 'tip', 'name', 'ip', 'iq', 'np'], axis=1, inplace=True)

            log.info(f'Количество отключаемых элементов сети:'
                     f' ветвей - {len(disable_df_vetv.axes[0])},'
                     f' узлов - {len(disable_df_node.axes[0])},'
                     f' генераторов - {len(disable_df_gen.axes[0])}.')

            disable_df_all = pd.concat([disable_df_vetv, disable_df_node, disable_df_gen])

            # удалить пробелы и значения после #
            # disable_df_all.loc[disable_df_all['dname'] == '', 'dname'] = \
            #     disable_df_all.loc[disable_df_all['dname'] == '', 'name']
            # disable_df_all['dname'] = disable_df_all['dname'].str.replace('  ', ' ').str.split('(').str[0]
            # disable_df_all['dname'] = disable_df_all['dname'].str.split(',').str[0].str.strip()

            for col in ['disable_scheme', 'repair_scheme', 'double_repair_scheme']:
                disable_df_all[col] = disable_df_all[col].str.replace(' ', '').str.split('#').str[0]
                disable_df_all[col] = disable_df_all[col].apply(GeneralSettings.split_task_action)

            # Цикл по всем возможным сочетаниям отключений
            for n_, self.info_srs['Контроль ДТН'] in self.set_comb.items():  # Цикл н-1 н-2 н-3.
                if n_ > len(disable_df_all):
                    break
                log.info(f"Количество отключаемых элементов в комбинации: {n_} ({self.info_srs['Контроль ДТН']}).")
                if n_ == 1:
                    disable_all = disable_df_all.copy()
                else:
                    disable_all = \
                        disable_df_all[(disable_df_all['uhom'] > 300) | (disable_df_all['table'] != 'node')]
                disable_all.drop(['uhom'], axis=1, inplace=True)
                name_columns = list(disable_all.columns)
                disable_all = tuple(disable_all.itertuples(index=False, name=None))

                for comb in combinations(disable_all, r=n_):  # Цикл по комбинациям.
                    log.debug(f'Комбинация элементов {comb}')
                    comb_df = pd.DataFrame(data=comb, columns=name_columns)
                    comb_df['double_repair_scheme_copy'] = comb_df['double_repair_scheme']
                    unique_set_actions = []

                    # Под i понимаем номер отключаемого элемента, остальные в ремонте.
                    # Если -1, то ремонт всех элементов.
                    i_min = 0 if len(comb_df) == 3 else -1
                    for i in range(n_ - 1, i_min - 1, -1):  # От последнего по первого элемента или -1.

                        comb_df['status_repair'] = True  # Истина, если элемент в ремонте. Ложь отключен.
                        if i != -1:
                            comb_df.loc[i, 'status_repair'] = False

                        # Если в ремонте 2 элемента.
                        double_repair = True if (n_ == 2 and i == -1) or (n_ == 3) else False
                        if self.info_srs['Контроль ДТН'] == "AДТН" and double_repair and n_ == 2:
                            continue  # Не расчетный случай.

                        double_repair_scheme = []
                        comb_df['double_repair_scheme'] = False
                        if double_repair:

                            # Оставить только общие значения.
                            if comb_df.loc[comb_df['status_repair'], 'double_repair_scheme_copy'].all():
                                double_repair_scheme = comb_df.loc[comb_df['status_repair'], 'double_repair_scheme_copy'
                                                                   ].to_list()
                                double_repair_scheme = list(set(double_repair_scheme[0]) & set(double_repair_scheme[1]))
                                if double_repair_scheme:
                                    comb_df.loc[comb_df['status_repair'], 'double_repair_scheme'] = comb_df.loc[
                                        comb_df['status_repair'], 'double_repair_scheme'].apply(
                                        lambda x: double_repair_scheme)

                        # Если нет дополнительных изменений сети, то всего 1 сочетание.
                        if not comb_df[['disable_scheme', 'repair_scheme', 'double_repair_scheme']].any().any():
                            self.calc_comb(rm, comb_df)
                            break

                        # Суммировать текущий набор изменений сети в set и проверить на уникальность.
                        set_actions = set()
                        for _, row in comb_df.iterrows():  # todo может можно заменить на Itertuples
                            if row['status_repair']:
                                if double_repair_scheme:
                                    set_actions.add(tuple(double_repair_scheme))
                                else:
                                    if row['repair_scheme']:
                                        set_actions.add(tuple(row['repair_scheme']))
                            else:
                                if row['disable_scheme']:
                                    set_actions.add(tuple(row['disable_scheme']))

                        if set_actions not in unique_set_actions:
                            unique_set_actions.append(set_actions)
                            self.calc_comb(rm, comb_df)

        # Отключаемые элементы сети по excel.
        if self.task_calc['cb_disable_excel']:
            if self.srs_xl.empty:
                raise ValueError(f'Таблица отключений из xl отсутствует.')
            # self.srs_xl.fillna(0, inplace=True)
            comb_xl = self.gen_comb_xl(rm, self.srs_xl)
            for comb in comb_xl:
                self.info_srs['Контроль ДТН'] = 'ДДТН'
                if self.set_save['gost']:
                    if comb.shape[0] == 3 or (comb.shape[0] == 2 and rm.code_name_rg2 in [1, 2, 3]):
                        self.info_srs['Контроль ДТН'] = 'АДТН'
                    if rm.code_name_rg2 in [1, 2, 3] and (comb.shape[0] == 3 or
                                                          (comb.shape[0] == 2 and all(comb['status_repair']))):
                        log.info(f'Сочетание отклонено по ГОСТ: ')
                        log.info(tabulate(comb, headers='keys', tablefmt='psql'))
                        continue
                self.calc_comb(rm, comb)

        # Прибавить info_file к overloads_srs.
        if not self.overloads_srs.empty:
            self.overloads_srs.index = self.overloads_srs['Номер СРС'].astype(str) + '.' + \
                                       self.overloads_srs['Номер подсочетания'].astype(str) + '_' + \
                                       self.overloads_srs.index.astype(str)
            self.overloads_all = pd.concat([self.overloads_all,
                                            self.overloads_srs.apply(lambda x: rm.info_file,
                                                                     axis=1).join(self.overloads_srs)])
            self.overloads_srs.drop(self.overloads_srs.index, inplace=True)

        # Вывод таблиц К-О в excel
        if self.task_calc['cb_tab_KO'] and (len(self.control_I) or len(self.control_U)):
            name_sheet = f'{self.file_count}_{rm.info_file["Имя файла"]}'.replace('[', '').replace(']', '')[:28]
            control_df_dict = {}
            if len(self.control_I):
                control_df_dict[name_sheet + '{I}'] = self.control_I
                self.control_I = None
            if len(self.control_U):
                control_df_dict[name_sheet + '{U}'] = self.control_U
                self.control_U = None
            # https://www.geeksforgeeks.org/how-to-write-pandas-dataframes-to-multiple-excel-sheets/

            # mode = 'a' if os.path.exists(self.book_path) else 'w'
            if not os.path.exists(self.book_path):
                Workbook().save(self.book_path)

            with pd.ExcelWriter(path=self.book_path, mode='a', engine="openpyxl") as writer:
                for name_sheet, df_control in control_df_dict.items():
                    # Поиск столбцов с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                    # https/www.geeksforgeeks.org/how-to-find-drop-duplicate-columns-in-a-pandas-dataframe/
                    df_control_head = df_control.iloc[:4].T  # включая groupid
                    duplicated_true = df_control_head.duplicated(keep=False)
                    groupid_true = df_control.loc['-', '-', 'groupid'] > 0
                    selection_columns = duplicated_true & groupid_true  # выборка в столбцах df_control для проверки

                    dict_equals = defaultdict(list)  # {номер:[перечень индексов столбцов с одинаковыми колонками]}
                    if selection_columns.any():
                        df_control_head = df_control_head[selection_columns]
                        duplicated_unique = df_control_head.drop_duplicates()
                        for i in range(len(duplicated_unique)):
                            col_unique = duplicated_unique.iloc[i, :]
                            for ii in range(len(df_control_head)):
                                control_col = df_control_head.iloc[ii, :]
                                if col_unique.equals(control_col):
                                    dict_equals[str(i)].append(int(control_col.name))
                    # Объединить столбцы с одинаковыми dname; ДДТН, А; АДТН, А; groupid
                    if dict_equals:
                        for cols in dict_equals.values():
                            df_control[cols[0]] = df_control[cols].max(axis=1)
                            df_control.drop(columns=cols[1:], inplace=True)

                    df_control.to_excel(excel_writer=writer,
                                        sheet_name=name_sheet,
                                        header=False,
                                        startrow=1,
                                        freeze_panes=(2, 3),
                                        index=True)

            # Форматирование таблиц Отключение - Контроль
            wb = load_workbook(self.book_path)
            for name_sheet in control_df_dict:
                ws = wb[name_sheet]
                num_tab, name_tab = GeneralSettings.read_title(self.task_calc['te_tab_KO_info'])
                ws['A1'] = f'{name_tab[0]}{num_tab + self.file_count - 1}{name_tab[1]} {rm.name_rm}'
                ws['A2'] = 'Наименование режима'
                ws['B2'] = 'Номер режима'
                ws['C2'] = 'Наименование параметра'
                # ws.merge_cells('A2:B4')
                thins = Side(border_style="thin", color="000000")
                max_column_lit = get_column_letter(ws.max_column)
                ws.merge_cells(f'A1:{max_column_lit}1')

                # Данные
                for row in range(3, ws.max_row + 1):
                    for col in range(4, ws.max_column + 1):
                        ws.cell(row, col).border = Border(thins, thins, thins, thins)
                        if ws.cell(row, 3).value in ['I, А', 'U, кВ']:
                            ws.cell(row, col).font = Font(bold=True)
                        if 'I, %' in ws.cell(row, 3).value:
                            if ws.cell(row, col).value >= 100:
                                ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor="00FF9900")
                        if 'U, %' in ws.cell(row, 3).value:
                            if ws.cell(row, col).value > 0:
                                ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor="00FF9900")
                # Колонки
                for litter, L in {'A': 35, 'B': 6, 'C': 17}.items():
                    ws.column_dimensions[litter].width = L
                for n in range(4, ws.max_column + 1):
                    ws[f'{get_column_letter(n)}2'].alignment = Alignment(textRotation=90, wrap_text=True,
                                                                         horizontal="center", vertical="center")
                    ws.column_dimensions[get_column_letter(n)].width = 9
                    ws[f'{get_column_letter(n)}2'].font = Font(bold=True)
                    ws[f'{get_column_letter(n)}2'].border = Border(thins, thins, thins, thins)
                # Строки
                ws.row_dimensions[5].hidden = True  # Скрыть
                ws.row_dimensions[6].hidden = True
                for n in range(1, ws.max_row + 1):
                    ws[f'A{n}'].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws[f'B{n}'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    ws[f'C{n}'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 145
            wb.save(self.book_path)

    def calc_comb(self, rm, comb: pd.DataFrame):
        """
        Смоделировать отключение элементов в комбинации.
        :param rm:
        :param comb:
        :return:
        """
        if self.restore_only_state:
            for name_table in rm.data_save_sta:
                rm.rastr.tables(name_table).ReadSafeArray(2, rm.data_columns_sta[name_table],
                                                          rm.data_save_sta[name_table])
            log.debug('Состояние элементов сети восстановлено.')
        else:
            for name_table in rm.data_save:
                rm.rastr.tables(name_table).ReadSafeArray(2, rm.data_columns[name_table],
                                                          rm.data_save[name_table])
            self.restore_only_state = True
            log.debug('Состояние элементов сети и параметров восстановлено.')

        comb.sort_values(by='status_repair', inplace=True)
        # Для добавления в 'Наименование СРС' данных о disable_scheme, double_repair_scheme и repair_scheme
        comb['scheme_info'] = ''
        log.debug(tabulate(comb, headers='keys', tablefmt='psql'))

        repair2_one = True  # Для выполнения действия с двойным отключением на 2-м элементе.

        for i in comb.index:
            if not rm.sta(table_name=comb.loc[i, 'table'],
                          ndx=comb.loc[i, 'index']):  # отключить элемент
                log.info(f'Комбинация отклонена: элемент {rm.t_name[comb.loc[i, "table"]][comb.loc[i, "s_key"]]!r}'
                         f' отключен в исходной РМ.')
                return False
            scheme_info = ''

            # Ремонтная схема
            if comb.loc[i, 'status_repair']:
                if comb.loc[i, 'double_repair_scheme']:
                    if repair2_one:
                        repair2_one = False
                    else:
                        scheme_info = self.perform_action(rm, comb.loc[i, 'double_repair_scheme'])
                else:
                    if comb.loc[i, 'repair_scheme']:
                        scheme_info = self.perform_action(rm, comb.loc[i, 'repair_scheme'])

            # Схема при отключении
            if (not comb.loc[i, 'status_repair']) and comb.loc[i, 'disable_scheme']:
                scheme_info = self.perform_action(rm, comb.loc[i, 'disable_scheme'])

            if scheme_info:
                comb.loc[i, 'scheme_info'] = f' ({scheme_info})'
        log.debug('Элементы сети из сочетания отключены.')

        # Имя сочетания
        self.info_srs.drop(labels=['Отключение', 'Ключ откл.', 'Ремонт 1', 'Ключ рем.1', 'Ремонт 2', 'Ключ рем.2'],
                           inplace=True, errors='ignore')
        dname = rm.t_name[comb["table"].iloc[0]][comb["s_key"].iloc[0]]
        if comb.iloc[0]["status_repair"]:
            name_srs = 'Ремонт '
            self.info_srs['Ремонт 1'] = dname + comb['scheme_info'].iloc[0]
            self.info_srs['Ключ рем.1'] = comb["s_key"].iloc[0]
        else:
            name_srs = 'Отключение '
            self.info_srs['Отключение'] = dname + comb['scheme_info'].iloc[0]
            self.info_srs['Ключ откл.'] = comb["s_key"].iloc[0]

        name_srs += dname + comb['scheme_info'].iloc[0]

        if len(comb) > 1:
            dname = rm.t_name[comb["table"].iloc[1]][comb["s_key"].iloc[1]]
            name_srs += ' при ремонте' if 'Откл' in name_srs else ' и'
            name_srs += f' {dname}{comb["scheme_info"].iloc[1]}'
            self.info_srs['Ремонт 1'] = dname + comb["scheme_info"].iloc[1]
            self.info_srs['Ключ рем.1'] = comb["s_key"].iloc[1]
        if len(comb) == 3:
            dname = rm.t_name[comb["table"].iloc[2]][comb["s_key"].iloc[2]]
            name_srs += f', {dname}{comb["scheme_info"].iloc[2]}'
            self.info_srs['Ремонт 2'] = dname + comb["scheme_info"].iloc[2]
            self.info_srs['Ключ рем.2'] = comb["s_key"].iloc[2]
        self.info_srs['Наименование СРС без()'] = re.sub(r'\(.+?\)', '', name_srs)
        name_srs += '.'

        self.info_srs['Наименование СРС'] = name_srs
        self.info_srs['Номер СРС'] = self.number_comb
        self.info_srs['Кол. откл. эл.'] = comb.shape[0]
        log.info(f"Сочетание {self.number_comb}: {name_srs}")

        self.do_action(rm)

    def perform_action(self, rm, task_action: list) -> str:
        """
        Выполнить действия, записанные в поле repair_scheme, disable_scheme.
        :param task_action: list("[1,2:sta=1]", "2")
        :param rm:
        :return: Наименование внесенных изменений в расчетное НВ.
        """
        info = []
        for task_action_i in task_action:
            if task_action_i.replace('.', '').isdigit():
                names, actions = self.pa.scheme_description(task_action_i)
                for i in range(len(actions)):
                    name = rm.cor_rm_from_txt(actions[i])
                    if name:
                        if names[i]:
                            name = names[i]
                        info.append(name)
            else:
                info.append(rm.cor_rm_from_txt(task_action_i))

        all_info = ', '.join(info) if info else ''
        self.test_not_only_sta(all_info)
        return all_info

    def test_not_only_sta(self, txt):
        """
        Проверка на наличие изменений в сети кроме состояния.
        :param txt: Строка сформированная group_cor
        """
        for i in ['нагрузки', 'генерации', 'ktr', 'pn', 'qn', 'pg', 'qg', 'vzd', 'bsh', 'P']:
            # список параметров сверять по group_cor, data_columns
            if i in txt:
                self.restore_only_state = False
                break

    def do_action(self, rm):
        """
        Цикл по действиям для ввода режима в область допустимых значений.
        :param rm:
        :return:
        """

        self.info_action = pd.Series(dtype='object')
        self.info_action['Номер подсочетания'] = 0

        # Если False - значит есть ПА, True - конец расчета сочетания (перегрузку нечем ликвидировать или отсутствует).

        # Цикл по действиям (ПА или ОП)
        while True:
            self.info_action['Конец'] = True

            self.do_control(rm)

            # прибавить info_srs и info_action к overloads_srs
            self.info_action['Номер подсочетания'] += 1
            self.info_action.drop(labels=['АРВ', 'СКРМ', 'Действие'], inplace=True, errors='ignore')

            if self.info_action['Конец']:
                break
        self.number_comb += 1  # код комбинации

    def do_control(self, rm):
        """
        Проверка параметров режима.
        :return:  Наполняет overloads_srs
        """
        log.debug(f'Проверка параметров УР.')
        test_rgm = rm.rgm('do_control')
        if self.set_save['avr']:
            self.info_action['АРВ'] = rm.node_include()
            if 'Восстановлено' in self.info_action['АРВ']:
                test_rgm = rm.rgm('Перерасчет после действия АВР.')
        # if self.set_save['skrm']:
        #     self.info_action['СКРМ'] = rm.auto_shunt_cor(all_auto_shunt=self.auto_shunt)
        #     if self.info_action['СКРМ']:
        #         test_rgm = rm.rgm('do_control')

        if not test_rgm:
            overloads = pd.DataFrame({'dname': ['Режим не моделируется'], 'i_zag': [-1], 'otv_min': [-1]})
        else:
            overloads = pd.DataFrame()
            # проверка на наличие перегрузок ветвей (ЛЭП, трансформаторов, выключателей)
            if self.info_srs['Контроль ДТН'] == 'АДТН':
                selection_v = 'all_control & i_zag_av > 0.1004'
                selection_n = 'all_control & vras<umin_av & !sta'
            else:
                selection_v = 'all_control & i_zag > 0.1004'
                selection_n = 'all_control & vras<umin & !sta'

            tv = rm.rastr.tables('vetv')
            tv.SetSel(selection_v)
            if tv.count:
                overloads = rm.df_from_table(table_name='vetv',
                                             fields='dname,'  # 'Контролируемые элементы,'
                                                    'key,'  # 'Ключ контроль,'
                                                    'txt_zag,'  # 'txt_zag,' 
                                                    'i_max,'  # 'Iрасч.(A),'
                                                    'i_dop_r,'  # 'Iддтн(A),'
                                                    'i_zag,'  # 'Iзагр.ддтн(%),'
                                                    'i_dop_r_av,'  # 'Iадтн(A),'
                                                    'i_zag_av',  # 'Iзагр.адтн(%),'
                                             setsel=selection_v)
            # проверка на наличие недопустимого снижение напряжения
            # todo округлить i_max
            tn = rm.rastr.tables('node')
            tn.SetSel(selection_n)
            if tn.count:
                overloads = pd.concat([overloads, rm.df_from_table(table_name='node',
                                                                   fields='dname,'  # 'Контролируемые элементы,'
                                                                          'key,'  # 'Ключ контроль,'
                                                                   # 'txt_zag,'  # 'txt_zag,'
                                                                   # todo сделать что бы в txt_zag были значения узлов?
                                                                          'vras,'  # 'Uрасч.(кВ),'
                                                                          'umin,'  # 'Uмин.доп.(кВ),'
                                                                          'umin_av,'  # 'U ав.доп.(кВ),'
                                                                          'otv_min,'
                                                                   # отклонение vras от 'Uмин.доп.' (%)
                                                                          'otv_min_av',
                                                                   # отклонение vras от 'U ав.доп.' (%)
                                                                   setsel=selection_n)])

            # проверка на наличие недопустимого повышения напряжения
            tn.SetSel('all_control & umax<vras & umax>0 & !sta')
            if tn.count:
                overloads = pd.concat([overloads, rm.df_from_table(table_name='node',
                                                                   fields='dname,'  # 'Контролируемые элементы,'
                                                                          'key,'  # 'Ключ контроль,'
                                                                          'vras,'  # 'Uрасч.(кВ),'
                                                                          'umax,'  # 'Uнаиб.раб.(кВ)'
                                                                          'otv_max',  # 'Uнаиб.раб.(кВ)'
                                                                   setsel='all_control & umax<vras & umax>0 & !sta')])
            # Таблица КОНТРОЛЬ - ОТКЛЮЧЕНИЕ
            if self.task_calc['cb_tab_KO']:
                log.debug('Запись параметров УР в таблицу КО.')
                if len(self.control_I):
                    ci = rm.df_from_table(table_name='vetv',
                                          fields='index,i_max,i_zag,i_zag_av',
                                          setsel="all_control")
                    ci.set_index('index', inplace=True)
                    ci['i_max'] = ci['i_max'].round(0)
                    ci['i_zag'] = ci['i_zag'].round(0)
                    ci['i_zag_av'] = ci['i_zag_av'].round(0)
                    ci = ci.T
                    ci.index = pd.MultiIndex.from_product([[self.info_srs['Наименование СРС']],
                                                           [f'{self.number_comb}.'
                                                            f'{self.info_action["Номер подсочетания"]}'],
                                                           ['I, А', 'I, % от ДДТН', 'I, % от АДТН']])
                    self.control_I = pd.concat([self.control_I, ci], axis=0)

                if len(self.control_U):
                    cu = rm.df_from_table(table_name='node',
                                          fields='index,vras,otv_min,otv_min_av',
                                          setsel="all_control")
                    cu.set_index('index', inplace=True)
                    cu['vras'] = cu['vras'].round(1)
                    cu['otv_min'] = cu['otv_min'].round(2)
                    cu['otv_min_av'] = cu['otv_min_av'].round(2)
                    cu = cu.T
                    cu.index = pd.MultiIndex.from_product([[self.info_srs['Наименование СРС']],
                                                           [f'{self.number_comb}.'
                                                            f'{self.info_action["Номер подсочетания"]}'],
                                                           ['U, кВ', 'U, % от МДН', 'I, % от АДН']])
                    self.control_U = pd.concat([self.control_U, cu], axis=0)

        num = len(overloads)
        log.debug(f'Выявлено {num} отклонений от допустимых значений.')
        # Добавить рисунки.
        if self.task_calc['results_RG2'] and (not self.task_calc['pic_overloads'] or
                                              (self.task_calc['pic_overloads'] and num)):
            log.debug('Добавить рисунки.')
            pic_name_file = rm.save(folder_name=self.task_calc['folder_result_calc'],
                                    file_name=f'{rm.name_base} '
                                              f'[{self.number_comb}_{self.info_action["Номер подсочетания"]}] '
                                              f'рис {self.num_pic} {self.info_srs["Наименование СРС без()"]}')

            # Южный р-н. Зимний максимум нагрузки 2026 г (-32°C/ПЭВТ). Нормальная схема сети. Действия...Загрузка...
            # todo Действия...Загрузка...
            add_name = f' ({", ".join(rm.additional_name_list)})' if rm.additional_name_list else ""
            picture_name = f'{self.name_pic[0]}{self.num_pic}{self.name_pic[1]} {rm.season_name} {rm.god} г' \
                           f'{add_name}. {self.info_srs["Наименование СРС"]}'
            pic_name_file = pic_name_file.replace(self.task_calc['folder_result_calc'] + '\\', '')
            self.df_picture.loc[len(self.df_picture.index)] = (pic_name_file, picture_name,)

            if num:
                overloads['num_pic'] = self.num_pic
            self.num_pic += 1

        if num:
            overloads.index = range(num)
            self.overloads_srs = pd.concat([self.overloads_srs,
                                            overloads.apply(lambda x: pd.concat([self.info_srs, self.info_action]),
                                                            axis=1).join(other=overloads)])


class EditModel(GeneralSettings):
    """
    Коррекция файлов.
    """

    def __init__(self, task):
        super(EditModel, self).__init__()
        self.all_read_ini()
        self.print_xl = None
        self.cor_xl = None
        self.task = task
        self.rastr_files = None
        self.all_folder = False  # Не перебирать вложенные папки

    def run_cor(self):
        """
        Запуск корректировки моделей.
        """
        test_run('edit')
        log.info('\n!!! Запуск корректировки РМ !!!\n')
        self.task["KIzFolder"] = self.task["KIzFolder"].strip()
        if "*" in self.task["KIzFolder"]:
            self.task["KIzFolder"] = self.task["KIzFolder"].replace('*', '')
            self.all_folder = True

        if not os.path.exists(self.task["KIzFolder"]):
            raise ValueError(f'Не найден путь: {self.task["KIzFolder"]}.')

        self.task['folder_result'] = self.task["KIzFolder"] + r"\result"
        if os.path.isfile(self.task["KIzFolder"]):
            self.task['folder_result'] = os.path.dirname(self.task["KIzFolder"]) + r"\result"

        self.task["KInFolder"] = self.task["KInFolder"].strip()
        # папка для сохранения result и KInFolder
        if self.task["KInFolder"] and not os.path.exists(self.task["KInFolder"]):
            if os.path.isdir(self.task["KIzFolder"]):
                log.info("Создана папка: " + self.task["KInFolder"])
                os.makedirs(self.task["KInFolder"])  # создать папку
                self.task['folder_result'] = self.task["KInFolder"] + r"\result"
            else:
                self.task['folder_result'] = os.path.dirname(self.task["KIzFolder"]) + r"\result"

        if not os.path.exists(self.task['folder_result']):
            os.mkdir(self.task['folder_result'])  # создать папку result

        self.task['name_time'] = f"{self.task['folder_result']}\\{datetime.now().strftime('%d-%m-%Y %H-%M-%S')}"

        if "import_val_XL" in self.task:
            if self.task["import_val_XL"]:  # Задать параметры узла по значениям в таблице excel (имя книги, имя листа)
                self.cor_xl = CorXL(excel_file_name=self.task["excel_cor_file"],
                                    sheets=self.task["excel_cor_sheet"])
                self.cor_xl.init_export_model()

        if os.path.isdir(self.task["KIzFolder"]):  # корр файлы в папке
            if self.all_folder:  # с вложенными папками
                for address, dirs, files in os.walk(self.task["KIzFolder"]):
                    in_dir = ''
                    if self.task["KInFolder"]:
                        in_dir = address.replace(self.task["KIzFolder"], self.task["KInFolder"])
                        if not os.path.exists(in_dir):
                            os.makedirs(in_dir)

                    self.for_file_in_dir(from_dir=address, in_dir=in_dir)

            else:  # без вложенных папок
                self.for_file_in_dir(from_dir=self.task["KIzFolder"], in_dir=self.task["KInFolder"])

        elif os.path.isfile(self.task["KIzFolder"]):  # корр файл
            rm = RastrModel(full_name=self.task["KIzFolder"])
            log.info("\n\n")
            rm.load()

            self.cor_file(rm)
            if self.task["KInFolder"]:
                if os.path.isdir(self.task["KInFolder"]):
                    rm.save(file_name=self.task["KInFolder"],
                            folder_name=rm.name_base)
                else:  # if os.path.isfile(self.task["KInFolder"]):
                    rm.save(full_name_new=self.task["KInFolder"])

        # для нескольких запусков через GUI
        if ImportFromModel.set_import_model:
            ImportFromModel.set_import_model = []

        if self.print_xl:
            self.print_xl.finish()

        self.the_end()
        if self.set_info['collapse']:
            self.set_info['end_info'] += f"\nВНИМАНИЕ! Развалились модели:\n[{self.set_info['collapse']}].\n"

        notepad_path = self.task['name_time'] + ' протокол коррекции файлов.log'
        shutil.copyfile(self.log_file, notepad_path)
        with open(self.task['name_time'] + ' задание на корректировку.yaml', 'w') as f:
            yaml.dump(data=self.task, stream=f, default_flow_style=False, sort_keys=False)
        # webbrowser.open(notepad_path)  #  Открыть блокнотом лог-файл.
        mb.showinfo("Инфо", self.set_info['end_info'])

    def for_file_in_dir(self, from_dir: str, in_dir: str):
        files = os.listdir(from_dir)  # список всех файлов в папке
        self.rastr_files = list(filter(lambda x: x.endswith('.rg2') | x.endswith('.rst'), files))

        for rastr_file in self.rastr_files:  # цикл по файлам .rg2 .rst в папке KIzFolder
            if self.task["KFilter_file"] and self.file_count == self.task["max_file_count"]:
                break  # Если включен фильтр файлов проверяем количество расчетных файлов.
            full_name = os.path.join(from_dir, rastr_file)

            rm = RastrModel(full_name)
            # если включен фильтр файлов и имя стандартизовано
            if self.task["KFilter_file"] and rm.code_name_rg2:
                if not rm.test_name(condition=self.task["cor_criterion_start"], info='Цикл по файлам.'):
                    continue  # пропускаем если не соответствует фильтру
            log.info("\n\n")
            rm.load()
            self.cor_file(rm)
            if self.task["KInFolder"]:
                rm.save(full_name_new=os.path.join(in_dir, rastr_file))

    def cor_file(self, rm):
        """Корректировать файл rm"""
        self.file_count += 1

        # Импорт моделей
        if ImportFromModel.set_import_model:
            for im in ImportFromModel.set_import_model:
                im.import_data_in_rm(rm)

        if self.task['cor_beginning_qt']['add']:
            log.info("\t*** Корректировка моделей в текстовом формате ***")
            rm.cor_rm_from_txt(self.task['cor_beginning_qt']['txt'])
            log.info("\t*** Конец выполнения корректировки моделей в текстовом формате ***")

        # Задать параметры по значениям в таблице excel
        if "import_val_XL" in self.task:
            if self.task["import_val_XL"]:
                self.cor_xl.run_xl(rm)

        if 'checking_parameters_rg2' in self.task:
            if self.task['checking_parameters_rg2']:
                if not rm.checking_parameters_rg2(self.task['control_rg2_task']):  # расчет и контроль параметров режима
                    self.set_info['collapse'] += rm.name_base + ', '

        if 'printXL' in self.task:
            if self.task['printXL']:
                if not type(self.print_xl) == PrintXL:
                    self.print_xl = PrintXL(self.task)
                self.print_xl.add_val(rm)


class RastrModel(RastrMethod):
    """
    Для хранения параметров текущего расчетного файла.
    """

    def __init__(self, full_name: str):
        super(RastrModel, self).__init__()
        self.full_name = full_name
        self.dir = os.path.dirname(full_name)
        self.Name = os.path.basename(full_name)  # вернуть имя с расширением "2020 зим макс.rg2"
        self.name_base, self.type_file = self.Name.rsplit(sep='.', maxsplit=1)
        self.pattern = GeneralSettings.set_save["шаблон " + self.type_file]
        self.code_name_rg2 = 0  # 0 не распознан, 1 зим макс 2 зим мин 3 ПЭВТ 4 лет макс 5 лет мин 6 паводок
        self.all_auto_shunt = {}
        self.temperature: float = 0
        self.rastr = None
        self.name_list = ["-", "-", "-"]
        self.additional_name = None
        self.additional_name_list = None
        self.season_name: str = ''
        self.god: str = ''
        self.name_rm: str = self.Name
        self.info_file = pd.Series(dtype='object')  # имя файла

        # Для хранения исходной схемы и параметров сети
        self.data_save = None
        self.data_columns = None
        self.data_save_sta = None
        self.data_columns_sta = None
        self.t_sta = {}  # {имя таблицы: {(ip, iq, np): 0 или 1}}
        self.t_name = {}  # {имя таблицы: {ny: имя}}
        self.t_i = {}  # {имя таблицы: {(ip, iq, np): индекс}}
        for tab_name in ['node', 'vetv', 'Generator']:
            self.t_sta[tab_name] = {}
            self.t_name[tab_name] = {}
            self.t_i[tab_name] = {}
        self.ny_join_vetv = defaultdict(list)  # {ny: все присоединенные ветви}

        # self.ny_pqng = defaultdict(tuple)  # {ny: (pn, qn, pg, qn)} - все с pn pg > 0 | qn pg > 0 | pg > 0 | qg > 0
        self.v_gr = {}   # {(ip, iq, np): groupid} - все c groupid > 0
        self.v_rxb = {}   # {(ip, iq, np): (r, x, b)} - все

        # "^(20[1-9][0-9])\s(лет\w?|зим\w?|паводок)\s?(макс|мин)?"
        match = re.search(re.compile(r"^(20[1-9][0-9])\s(лет\w*|зим\w*|паводок)\s?(макс\w*|мин\w*)?"), self.name_base)
        if match:
            if match.re.groups == 3:
                self.name_list = [match[1], match[2], match[3]]
                if not self.name_list[2]:
                    self.name_list = "-"
                if self.name_list[1] == "паводок":
                    self.code_name_rg2 = 6
                    self.season_name = "Паводок"
                if "зим" in self.name_list[1] and "макс" in self.name_list[2]:
                    self.code_name_rg2 = 1
                    self.season_name = "Зимний максимум нагрузки"
                if "зим" in self.name_list[1] and "мин" in self.name_list[2]:
                    self.code_name_rg2 = 2
                    self.season_name = "Зимний минимум нагрузки"
                if "лет" in self.name_list[1] and "макс" in self.name_list[2]:
                    self.code_name_rg2 = 4
                    self.season_name = "Летний максимум нагрузки"
                if "лет" in self.name_list[1] and "мин" in self.name_list[2]:
                    self.code_name_rg2 = 5
                    self.season_name = "Летний минимум нагрузки"

        self.god = self.name_list[0]
        if self.code_name_rg2:
            if self.code_name_rg2 in [4, 5] and ("ПЭВТ" in self.name_base):
                self.code_name_rg2 = 3
        self.name_rm = f'{self.season_name} {self.god} г.'

        # поиск в строке значения в ()
        match = re.search(re.compile(r"\((.+)\)"), self.name_base)
        if match:
            self.additional_name = match[1]
            self.additional_name_list = self.additional_name.split(";")

        if "°C" in self.name_base:
            match = re.search(re.compile(r"(-?\d+((,|\.)\d*)?)\s?°C"), self.name_base)  # -45.,14 °C
            if match:
                self.temperature = float(match[1].replace(',', '.'))  # число
                self.name_rm += f' Расчетная температура {self.temperature} °C.'

        self.info_file['Имя файла'] = self.name_base
        self.info_file['Год'] = self.god
        self.info_file['Сезон макс/мин'] = self.season_name
        self.info_file['Темп.(°C)'] = self.temperature
        if self.additional_name_list:
            for i, additional_name in enumerate(self.additional_name_list, 1):
                self.info_file['Доп. имя' + str(i)] = additional_name

    def save_value_fields(self):
        """
        Сохранить значения изменяемых полей в исходной схеме сети и считать некоторые данные.
        """
        log.info('Сохранение значений исходных параметров сети.')
        
        self.data_save_sta = {'vetv': None, 'node': None, 'Generator': None}
        self.data_columns_sta = {'vetv': 'ip,iq,np,sta', 
                                 'node': 'ny,sta',
                                 'Generator': 'Num,sta'}
        for name_tab in self.data_save_sta:
            self.data_save_sta[name_tab] = \
                self.rastr.tables(name_tab).writesafearray(self.data_columns_sta[name_tab], "000")
            
        self.data_save = {'vetv': None, 'node': None, 'Generator': None}
        self.data_columns = {'vetv': 'ip,iq,np,sta,ktr',  # ,r,x,b
                             'node': 'ny,sta,pn,qn,pg,qg,vzd,bsh',
                             'Generator': 'Num,sta,P'}
        for name_tab in self.data_save:
            self.data_save[name_tab] = \
                self.rastr.tables(name_tab).writesafearray(self.data_columns[name_tab], "000")

        # Узлы
        for ny, sta, pn, qn, pg, qg, vzd, bsh in self.data_save['node']:
            self.t_sta['node'][ny] = sta
            # if pn or qn or pg:
            #     self.ny_pqng[ny] = (pn, qn, pg, qg)
        t = self.rastr.tables('node').writesafearray("ny,name,dname,index", "000")
        for ny, name, dname, index in t:
            self.t_i['node'][ny] = index
            if dname:
                self.t_name['node'][ny] = dname
            else:
                self.t_name['node'][ny] = name if name else f'Узел {ny}'

        # Ветви
        for ip, iq, np_, sta, ktr in self.data_save['vetv']:  # , r, x, b
            s_key = (ip, iq, np_) if np_ else (ip, iq)
            self.t_sta['vetv'][s_key] = sta
            self.ny_join_vetv[ip].append(s_key)
            self.ny_join_vetv[iq].append(s_key)

        t = self.rastr.tables('vetv').writesafearray("ip,iq,np,dname,groupid,r,x,b,index", "000")
        for ip, iq, np_, dname, groupid, r, x, b, index in t:
            s_key = (ip, iq, np_) if np_ else (ip, iq)
            self.t_i['vetv'][s_key] = index

            if dname:
                self.t_name['vetv'][s_key] = dname
            else:
                self.t_name['vetv'][s_key] = f'{self.t_name["node"][ip]} - {self.t_name["node"][iq]}'

            if groupid:
                self.v_gr[s_key] = groupid
            self.v_rxb[s_key] = (r, x, b)

        # Генераторы
        for Num, sta, P in self.data_save['Generator']:
            self.t_sta['Generator'][Num] = sta

        t = self.rastr.tables('Generator').writesafearray("Num,index,Name,Node", "000")
        for Num, index, Name, Node in t:
            self.t_i['Generator'][Num] = index
            if Name:
                self.t_name['Generator'][Num] = Name
            else:
                self.t_name['Generator'][Num] = f'генератор номер {Num} в узле {self.t_name["node"][Node]}'

    def network_analysis(self, disable_on: bool = True, field: str = 'disable',
                         selection_node_for_disable: str = ''):
        """
        Анализ графа сети.
        :param disable_on: Отметить отключаемые элементы в поле field узлов и ветвей.
        :param field: Поле для отметки.
        :param selection_node_for_disable: Выборка в таблице узлы, для выбора отключаемых элементов.
        например, района, территории, нагрузочной группы для расчета или "" - все узлы.
        """
        log.info('Анализ графа сети с заполнением поля transit в таблице узлов и ветвей.')
        if self.t_sta['node']:
            self.save_value_fields()

        all_ny = [ny for ny, _ in self.data_save_sta['node']]
        log.debug(f'В РМ {len(all_ny)} узлов.')

        vetv = self.rastr.tables('vetv')
        vetv.setsel('sta=0')
        data_v = vetv.writesafearray('ip,iq,np,groupid,tip,pl_ip,groupid', "000")
        # Все включенные ветви РМ
        all_vetv_sta0 = [(ip, iq, np_) for ip, iq, np_, groupid, tip, pl_ip, groupid in data_v]
        all_ny_in_v = []  # [Все ip iq в таблице ветви, если ny встречается 1 раз, то это тупик.]
        ny_end = set()  # Все тупиковые узлы

        # Поиск узлов с одной отходящей включенной ветвью - это тупик.
        ny_all_vetv = defaultdict(list)  # {ny: все примыкающие включенные ветви}

        for ip, iq, np_ in all_vetv_sta0:
            all_ny_in_v.append(ip)
            all_ny_in_v.append(iq)
            ny_all_vetv[ip].append((ip, iq, np_))
            ny_all_vetv[iq].append((ip, iq, np_))

        for k, v in Counter(all_ny_in_v).items():  # {Номер узла: количество отходящих ветвей}
            if v == 1:
                ny_end.add(k)
        log.debug(f'В РМ {len(ny_end)} тупиков.')

        # Найти остальные узлы тупиковых цепочек.
        all_v_end = set()  # Все тупиковые ветви
        ny_end2 = set()  # Вспомогательный набор
        for ny in ny_end:
            ny_next = ny
            # Поиск в цикле следующего узла цепочки, если его нет, то ny_next равен 0.
            while ny_next > 0:  # ny_next следующий проверяемый узел
                ny_source = ny_next
                v_not_end = []  # [записываем не тупиковые ветви узла]
                for i in ny_all_vetv[ny_source]:
                    if i not in all_v_end:
                        v_not_end.append(i)
                if len(v_not_end) == 1:
                    all_v_end.add(v_not_end[0])
                    ip, iq, np_ = v_not_end[0]
                    ny_next = iq if ip == ny_source else ip
                    ny_end2.add(ny_source)
                else:
                    ny_next = 0
        ny_end = ny_end | ny_end2
        log.debug(f'В РМ {len(ny_end)} тупиковых узлов.')

        # Определить транзитные и узловые узлы
        all_ny_transit = []  # [Все узлы РМ входящие в транзиты]
        all_ny_nodal = {}  # {ny: количество примыкающих не тупиковых ветвей.}
        for ny in all_ny:
            if ny not in ny_end:
                v_not_end = 0  # записываем не тупиковые ветви узла
                for i in ny_all_vetv[ny]:
                    if i not in all_v_end:
                        v_not_end += 1
                if v_not_end > 2:
                    all_ny_nodal[ny] = v_not_end
                else:
                    all_ny_transit.append(ny)
        # Заполнить номера транзитов.
        num_transit = 0
        transit_num_all_ny = defaultdict(list)  # {номер транзита: все входящие узлы}
        transit_num_all_v_end = defaultdict(list)  # {номер транзита: крайние ветви транзита (ip, iq, np)}
        ny_use = set()

        for ny in all_ny_transit:
            if ny in ny_use:
                continue
            ny_use.add(ny)
            num_transit += 1
            transit_num_all_ny[num_transit].append(ny)
            for ip, iq, np_ in ny_all_vetv[ny]:
                v_end_transit = (ip, iq, np_)
                ny_next = iq if ip == ny else ip
                while ny_next:
                    ny_source = ny_next
                    ny_next = 0
                    if ny_source not in all_ny_nodal:
                        ny_use.add(ny_source)
                        transit_num_all_ny[num_transit].append(ny_source)
                        for ip1, iq1, np_1 in ny_all_vetv[ny_source]:
                            if (ip1, iq1, np_1) in all_v_end:
                                continue
                            for i in [ip1, iq1]:
                                if i not in ny_use:
                                    ny_next = i
                                    v_end_transit = (ip1, iq1, np_1)
                                    break
                            if ny_next:
                                break
                    else:
                        transit_num_all_v_end[num_transit].append(v_end_transit)
                        # log.debug((num_transit, v_end_transit))

        log.debug(f'В РМ {num_transit} групп транзитных узлов.')

        # Внести номера транзитов в таблицу узлы растра
        all_ny_transit = []  # [(транзитные узлы, номер транзита,)]
        ny__num_transit = {}  # {номер узла: номер транзита}
        for num in transit_num_all_ny:
            for ny in transit_num_all_ny[num]:
                all_ny_transit.append((ny, num,))
                ny__num_transit[ny] = num
        all_ny_transit = all_ny_transit + [(ny, -(all_ny_nodal[ny]),) for ny in all_ny_nodal]
        self.rastr.tables('node').ReadSafeArray(2, 'ny,transit', all_ny_transit)

        # Внести номера транзитов в таблицу ветви растра
        all_transit_one = []  # [(ip, iq, np_) всех транзитных ветвей состоящих из 1 элемента.]
        all_v_transit = []  # [(ip, iq, np_, num) все транзитные ветви]
        for i in all_vetv_sta0:
            if i in all_v_end:
                continue
            ip, iq, np_ = i
            num = 0
            if ip in ny__num_transit:
                num = ny__num_transit[ip]
            elif iq in ny__num_transit:
                num = ny__num_transit[iq]
            if num:
                all_v_transit.append((ip, iq, np_, num,))
            else:
                num_transit += 1
                all_transit_one.append((ip, iq, np_))
                all_v_transit.append((ip, iq, np_, num_transit,))
        vetv.ReadSafeArray(2, 'ip,iq,np,transit', all_v_transit)

        if disable_on:
            # Отключаемы узлы
            node = self.rastr.tables('node')
            node.setsel(selection_node_for_disable + '&transit<-3')  # 4-х и более отходящих транзитов
            node.cols.item(field).calc(1)
            log.info(f'{len(node)} отключаемых узлов')
            # Отключаемы ветви
            node.setsel(selection_node_for_disable)
            sel_ny = node.writesafearray('ny', "000")
            sel_ny = [x[0] for x in sel_ny]
            all_v_disable = []  # Все отключаемые ветви
            transit_use = []  # Уже добавленные в отключения номера транзиты
            v__gr = {(ip, iq, np_): groupid for ip, iq, np_, groupid, tip, pl_ip, groupid in data_v}
            v__pl = {(ip, iq, np_): pl_ip for ip, iq, np_, groupid, tip, pl_ip, groupid in data_v}
            v__tip = {(ip, iq, np_): tip for ip, iq, np_, groupid, tip, pl_ip, groupid in data_v}
            node.setsel('')
            ny__un = {ny: uhom for ny, uhom in self.rastr.tables('node').writesafearray('ny,uhom', "000")}
            for ny in sel_ny:
                if ny not in ny_end:
                    for v in ny_all_vetv[ny]:  # Цикл по прилегающим ветвям
                        if v not in all_v_end:  # Без тупиков
                            if v in all_transit_one:  # todo поверить all_transit_one
                                all_v_disable.append(v)
                            else:
                                ip, iq, np_ = v
                                ny_transit = ip if ip in ny__num_transit else 0
                                if not ny_transit:
                                    ny_transit = iq if iq in ny__num_transit else 0

                                if ny_transit:
                                    num_transit = ny__num_transit[ny_transit]
                                    if num_transit in transit_use:
                                        continue
                                    transit_use.append(num_transit)
                                    # Сравнить groupid концов транзита, если одинаковый, то отключаем конец
                                    # с большей суммой напряжений ip и ip
                                    # log.debug(transit_num_all_v_end[num_transit])
                                    ip1, iq1, np_1 = transit_num_all_v_end[num_transit][0]
                                    ip2, iq2, np_2 = transit_num_all_v_end[num_transit][1]
                                    if v__gr[(ip1, iq1, np_1)] == v__gr[(ip2, iq2, np_2)] and v__gr[(ip1, iq1, np_1)]:
                                        # В случае АТ, нужно отключать обмотку ВН
                                        if (ny__un[ip1] + ny__un[iq1]) > (ny__un[ip2] + ny__un[iq2]):
                                            all_v_disable.append((ip1, iq1, np_1))
                                        else:
                                            all_v_disable.append((ip2, iq2, np_2))
                                    else:
                                        # Отключаем оба конца. Если разница P < 1, то любой конец.
                                        # Положительное направление в центр транзита
                                        p1 = v__pl[(ip1, iq1, np_1)]  # Поток от начала к концу со знаком -
                                        if ip1 in all_ny_nodal:
                                            p1 = -p1
                                        p2 = v__pl[(ip2, iq2, np_2)]
                                        if ip2 in all_ny_nodal:
                                            p2 = -p2

                                        if abs(p1 + p2) > 1:
                                            all_v_disable.append((ip2, iq2, np_2))
                                            all_v_disable.append((ip1, iq1, np_1))
                                        else:
                                            if v__tip[(ip1, iq1, np_1)] == 2:  # выключатель
                                                all_v_disable.append((ip2, iq2, np_2))
                                            else:
                                                all_v_disable.append((ip1, iq1, np_1))
            # todo в all_v_disable  есть дубликаты ?
            # todo опционо убрать выключатели
            all_v_disable = tuple(set([(ip, iq, np_, 1) for ip, iq, np_ in all_v_disable]))

            if all_v_disable:
                log.info(f'{len(all_v_disable)} отключаемых ветвей')
                vetv.ReadSafeArray(2, 'ip,iq,np,' + field, all_v_disable)

    def index(self, table_name: str, key_int: Union[int | tuple] = 0,  key_str: str = '') -> int:
        """
        Возвращает номер строки в таблице по ключу в одном из форматов.
        При наличии t_i индекс берется из них.
        :param table_name: 'vetv' ...
        :param key_int: например  10 или (1, 2, 0)
        :param key_str: например 'ny=10' или 'ip=1&iq=2&np=3'
        :return: index
        """
        if not table_name:
            raise ValueError(f'Ошибка в задании {table_name=}.')
        if key_int:
            if table_name in ['node', 'vetv', 'Generator'] and key_int in self.t_i[table_name]:
                return self.t_i[table_name][key_int]
            else:
                t = self.rastr.tables(table_name)

                if table_name == 'vetv':
                    np_ = key_int[2] if len(key_int) == 3 else 0
                    t.setsel(f'ip={key_int[0]}&iq={key_int[1]}&np={np_}')
                else:
                    t.setsel(f'{t.Key}={key_int}')
                i = t.FindNextSel(-1)
                if i > -1:
                    log.warning(f'В таблице{table_name} не найдена строка по ключу {key_int} ')
                return i
        if key_str:
            t = self.rastr.tables(table_name)
            t.setsel(key_str)
            i = t.FindNextSel(-1)
            if i > -1:
                log.warning(f'В таблице{table_name} не найдена строка по ключу {key_int} ')
            return i

    @staticmethod
    def name_table_from_key(task_key: str):
        """
        По ключу строки (нр ny=1) определяет имя таблицы.
        :return: table:str 'node' or False если не найдено.
        """
        for key_tables in RastrMethod.KEY_TABLES:
            if key_tables in task_key:
                return RastrMethod.KEY_TABLES[key_tables]
        return False

    def replace_links(self, formula: str) -> str:
        """
        Функция заменяет в формуле ссылки на значения в таблицах rastr, на соответствующие значения.
        :param formula: '(10.5+15,16,2:r)*ip.uhom'
        :return: formula: '(10.5+z)*ip.uhom'
        """
        # formula = formula.replace(' ', '')
        formula_list = re.split('\*|/|\^|\+|-|\(|\)|==|!=|&|\||not|>|<|<=|=<|>=|=>', formula)
        for formula_i in formula_list:
            if ':' in formula_i:
                if any([txt in formula_i for txt in ['years', 'season', 'max_min', 'add_name']]):
                    continue
                sel_all, field = formula_i.split(':')
                name_table, sel = self.recognize_key(sel_all)
                self.rgm(f'для определения значения {formula}')
                index = self.index(table_name=name_table, key_str=sel)
                if index > -1:
                    new_val = self.rastr.tables(name_table).cols.Item(field).ZS(index)
                    formula = formula.replace(formula_i, new_val)
                else:
                    raise ValueError(f'В таблице {name_table} отсутствует {sel}')
        return formula

    def sta(self, table_name: str, ndx: int = 0, key_int: Union[int | tuple] = 0) -> bool:
        """
        Отключить ветвь(группу ветвей, если groupid!=0), узел (с примыкающими ветвями) или генератор.
        Отключаемый элемент определяется по ndx или key_int.
        :param table_name:
        :param ndx:
        :param key_int:
        :return: False если элемент отключен в исходном состоянии.
        """
        if not ndx:
            ndx = self.index(table_name=table_name, key_int=key_int)

        rtable = self.rastr.tables(table_name)

        sta_test = self.t_i[table_name].get(key_int)

        if table_name in ['node', 'vetv', 'Generator'] and sta_test:
            if sta_test == 1:
                return False
        else:
            if rtable.cols.item('sta').Z(ndx) == 1:
                return False

        if table_name == 'vetv':
            if self.v_gr and key_int:
                groupid = self.v_gr.get(key_int)
                if groupid:
                    rtable.setsel(f'groupid={groupid}')
                    rtable.cols.item('sta').Calc(1)
                    return True
            else:
                groupid = rtable.cols.item('groupid').Z(ndx)
                if groupid:
                    rtable.setsel(f'groupid={groupid}')
                    rtable.cols.item('sta').Calc(1)
                    return True
        # elif table == 'node':
        #     self.sta_node_with_branches(ny=rtable.cols.item('ny').Z(ndx), sta=1)
        rtable.cols.item('sta').SetZ(ndx, 1)
        return True

    def test_name(self, condition: dict, info: str = "") -> bool:
        """
         Проверка имени файла на соответствие условию condition.
        :param condition:
        {"years":"2020,2023...2025","season": "лет,зим,паводок","max_min":"макс","add_name":"-41С;МДП:ТЭ-У"}
        :param info: для вывода в протокол
        :return: True если удовлетворяет
        """
        if not condition:
            return True
        if not (any(condition.values())):  # условие пустое
            return True
        # Проверка года
        if 'years' in condition:
            if condition['years']:
                if not int(self.god) in str_yeas_in_list(str(condition['years'])):
                    log.info(f"{info} {self.Name!r}. Год не проходит по условию: {condition['years']!r}")
                    return False
        # Проверка "зим" "лет" "паводок"
        if 'season' in condition:
            if condition['season']:
                if not self.name_list[1] in condition['season'].replace(' ', '').split(","):
                    log.info(f'{info} {self.Name!r}. Сезон не проходит по условию: {condition["season"]!r}')
                    return False
        # Проверка "макс" "мин"
        if 'max_min' in condition:
            if condition['max_min']:
                if self.name_list[2] not in condition['max_min'].replace(' ', '').split(","):
                    log.info(f'{info} {self.Name!r}. Не проходит по условию: {condition["max_min"]!r}')
                    return False
        # Проверка доп имени, например (-41С;МДП:ТЭ-У)
        if 'add_name' in condition:
            if condition['add_name']:
                if condition['add_name'].strip():
                    for us in condition['add_name'].split(";"):
                        if us not in self.additional_name_list:
                            log.debug(f'{info} {self.Name}. Не проходит по условию: {us!r}')
                            return False
        return True

    def load(self):
        """
        Загрузить модель в Rastr
        """
        if not self.rastr:
            try:
                self.rastr = win32com.client.Dispatch("Astra.Rastr")
            except Exception:
                raise Exception('Com объект Astra.Rastr не найден')

        log.info(f"Загружен файл: {self.full_name}")
        self.rastr.Load(1, self.full_name, self.pattern)  # Загрузить или перезагрузить

    def downloading_additional_files(self, load_additional: list = None):
        """
        Загрузка в Rastr дополнительных файлов из папки с РМ.
        :param load_additional: ['amt','sch','trn']
        """
        for extension in load_additional:
            files = os.listdir(self.dir)
            names = list(filter(lambda x: x.endswith('.' + extension), files))
            if len(names) > 0:
                self.rastr.Load(1, f'{self.dir}\\{names[0]}', GeneralSettings.set_save[f"шаблон {extension}"])
                log.info(f"Загружен файл: {names[0]}")
            else:
                raise ValueError(f'Файл с расширением {extension!r} не найден в папке {self.dir}')

    def save(self, full_name_new: str = '', file_name: str = '', folder_name: str = ''):
        """
        Сохранить файл. Указать полное имя или имя файла (без расширения) с папкой.
        """
        if not full_name_new:
            if file_name and folder_name:
                full_name_new = folder_name + '\\' + re.sub(r'\\|\/|\:|\?|<|>|\||\.', '', file_name)
                full_name_new = full_name_new[:252]
                full_name_new += '.' + self.type_file

        self.rastr.Save(full_name_new, self.pattern)
        log.info("Файл сохранен: " + full_name_new)
        return full_name_new

    def checking_parameters_rg2(self, dict_task: dict):
        """  контроль  dict_task = {'node': True, 'vetv': True, 'Gen': True, 'section': True,
             'area': True, 'area2': True, 'darea': True, 'sel_node': "na>0"}  """
        if not self.rgm("checking_parameters_rg2"):
            return False

        log.info(f"Расчет загрузки ветвей для температуры {self.temperature}.")
        self.rastr.CalcIdop(self.temperature, 0.0, "")

        node = self.rastr.tables("node")
        branch = self.rastr.tables("vetv")
        # Также проверяется наличие узлов без ветвей, ветвей без узлов начала или конца, генераторов без узлов.
        all_ny = set([x[0] for x in node.writesafearray("ny", "000")])
        all_ip = set([x[0] for x in branch.writesafearray("ip", "000")])
        all_iq = set([x[0] for x in branch.writesafearray("iq", "000")])
        all_iq_ip = all_ip.union(all_iq)

        # Узлы без ветвей.
        all_ny_not_branches = all_ny - all_iq_ip
        if all_ny_not_branches:
            log.error(f'В таблице node узлы без ветвей: {all_ny_not_branches}')
        # Ветви без узлов.
        all_ip_iq_not_node = all_iq_ip - all_ny
        if all_ip_iq_not_node:
            log.error(f'В таблице vetv есть ссылка на узлы которых нет в таблице node: {all_ip_iq_not_node}')
        # Генераторы без узлов.
        generator = self.rastr.tables("Generator")
        if generator.size:
            all_gen_ny = set([x[0] for x in generator.writesafearray("Node", "000")])
            all_gen_not_node = all_gen_ny - all_ny
            if all_gen_not_node:
                log.error(f'В таблице Generator есть ссылка на узлы которых нет в таблице node: {all_gen_not_node}')

        if dict_task["sel_node"]:
            self.add_fields_in_table(name_tables='node', fields='sel1', type_fields=3)
            node.cols.item("sel1").calc(0)
            node.setsel(dict_task["sel_node"])
            node.cols.item("sel1").calc(1)

        # Напряжения
        if dict_task["node"]:
            log.info("\tКонтроль напряжений.")
            self.voltage_nominal(choice=(dict_task["sel_node"] + '&uhom>30'))
            self.voltage_deviation(choice=dict_task["sel_node"])
            self.voltage_fine(choice=dict_task["sel_node"])
            self.voltage_error(choice=dict_task["sel_node"])

        # Токи
        if dict_task['vetv']:
            # Контроль токовой загрузки

            if dict_task["sel_node"]:
                sel_vetv = "i_zag>=0.1&(ip.sel1|iq.sel1)"
                presence_n_it = {'n_it': "(ip.sel1|iq.sel1)&n_it>0",
                                 'n_it_av': "(ip.sel1|iq.sel1)&n_it_av>0"}
            else:
                sel_vetv = "i_zag>=0.1"
                presence_n_it = {'n_it': "n_it>0",
                                 'n_it_av': "n_it_av>0"}

            log.debug('Контроль токовой загрузки.')
            branch.setsel(sel_vetv)
            if branch.count:  # есть превышения
                j = branch.FindNextSel(-1)
                while j > -1:
                    log.info(f"\t\tВНИМАНИЕ ТОКИ! vetv:{branch.SelString(j)}, "
                             f"{branch.cols.item('name').ZS(j)} - {branch.cols.item('i_zag').ZS(j)} %")
                    j = branch.FindNextSel(j)

            log.debug('Проверка наличия n_it,n_it_av в таблице График_Iдоп_от_Т(graphikIT).')
            graph_it = self.rastr.tables("graphikIT")
            if graph_it.size:
                all_graph_it = set([x[0] for x in graph_it.writesafearray("Num", "000")])
                for field, sel_vetv_n_it in presence_n_it.items():
                    branch.setsel(sel_vetv_n_it)
                    for i in branch.writesafearray(field + ",name,ip,iq,np", "000"):
                        if i[0] > 0 and i[0] not in all_graph_it:
                            log.error(f"\t\tВНИМАНИЕ graphikIT! vetv: {i[1]} [{i[2]},{i[3]},{i[4]}] "
                                      f"{field}={i[0]} не найден в таблице График_Iдоп_от_Т")

        #  ГЕНЕРАТОРЫ
        if dict_task['Gen']:
            log.info("\tКонтроль генераторов")
            chart_pq = set([x[0] for x in self.rastr.tables("graphik2").writesafearray("Num", "000")])
            sel_gen = "!sta&Node.sel1" if dict_task["sel_node"] else "!sta"
            generator.setsel(sel_gen)
            col = {'Num': 0, 'Node': 1, 'Name': 2, 'Pmin': 3, 'Pmax': 4, 'P': 5, 'NumPQ': 6}
            if generator.count:
                for i in generator.writesafearray(','.join(col), "000"):
                    Pmin = i[col['Pmin']]
                    Pmax = i[col['Pmax']]
                    P = i[col['P']]
                    Name = i[col['Name']]
                    Num = i[col['Num']]
                    Node = i[col['Node']]
                    NumPQ = i[col['NumPQ']]
                    if P < Pmin and Pmin:
                        log.info(f"\t\tВНИМАНИЕ! ГЕНЕРАТОР: {Name}, {Num=},{Node=}, {P=} < {Pmin=}")
                    if P > Pmax and Pmax:
                        log.info(f"\t\tВНИМАНИЕ! ГЕНЕРАТОР: {Name}, {Num=},{Node=}, {P=} > {Pmax=}")
                    if NumPQ and self.rastr.tables("graphik2").size:
                        chart_pq = set([x[0] for x in self.rastr.tables("graphik2").writesafearray("Num", "000")])
                        if NumPQ not in chart_pq:
                            log.info(f"\t\tВНИМАНИЕ! ГЕНЕРАТОР: {Name}, {Num=},{Node=}, "
                                     f"{NumPQ=} не найден в таблице PQ-диаграммы (graphik2)")
        return True

    def cor_rm_from_txt(self, task_txt: str) -> str:
        """
        Корректировать модели по заданию в текстовом формате:
        Имя_функции[действие]{Условие_выполнения}#комментарии\n...
        Имя_функции по умолчанию = изм
        :param task_txt:
        :return: Информация
        """
        info = []

        task_rows = task_txt.split('\n')
        for task_row in task_rows:
            task_row = task_row.split('#')[0]  # удалить текст после '#'
            name_fun = task_row.split('[', 1)[0]  # Имя функции, стоит перед "[".
            name_fun = name_fun.replace(' ', '')
            if not name_fun:
                if '[' in task_row:
                    name_fun = 'изм'
                else:
                    continue  # К следующей строке.

            # Условие выполнения в фигурных скобках
            if '{' in task_row:
                if not self.conditions_test(task_row):
                    log.debug(f'Условие не выполняется: {task_row}')
                    continue  # К следующей строке.
                else:
                    log.debug(f'Условие выполняется: {task_row}')

            # Параметры функции в квадратных скобках
            function_parameters = []
            match = re.search(re.compile(r"\[(.+?)]"), task_row)
            if match:
                function_parameters = match[1].split(':', maxsplit=1)
            function_parameters += ['', '']
            info_i = self.txt_task_cor(name=name_fun,
                                       sel=function_parameters[0],
                                       value=function_parameters[1])
            if info_i:
                info.append(info_i)
        return ', '.join(info) if info else ''

    def conditions_test(self, conditions: str) -> bool:
        """
        В строке типа "years : 2026...2029& ny=1: vras>125|(not ny=1: na==2)" проверяет выполнение условий.
        Если в conditions имеются {}, то значения берутся внутри скобок
        :param conditions:
        :return:
        """
        if '{' in conditions:
            match = re.search(re.compile(r"\{(.+?)}"), conditions)
            if match:
                conditions = match[1].strip()
            else:
                raise ValueError(f'Ошибка в условии {conditions}')
        conditions_s = conditions
        conditions = self.replace_links(conditions)
        conditions_list = re.split('\*|/|\^|\+|-|\(|\)|==|!=|&|\||not|>|<|<=|=<|>=|=>', conditions)
        for condition in conditions_list:
            if ':' in condition:
                for key_txt in ['years', 'season', 'max_min', 'add_name']:
                    if not self.code_name_rg2:  # Если имя не стандартное, то True.
                        conditions = conditions.replace(condition, 'True')
                        continue
                    else:
                        if key_txt in condition:
                            par, value = condition.split(':')

                            if self.test_name(condition={par.replace(' ', ''): value.strip()}, info=condition):
                                conditions = conditions.replace(condition, 'True')
                            else:
                                conditions = conditions.replace(condition, 'False')
        if ':' in conditions:
            raise ValueError("Ошибка в условии: " + conditions)
        try:
            log.debug(f'conditions_test: {conditions}')
            return bool(eval(conditions))
        except Exception:
            raise ValueError(f'Ошибка у условии: {conditions_s!r}.')

    def txt_task_cor(self, name: str, sel: str = '', value: str = '') -> str:
        """
        Функция для выполнения задания в текстовом формате
        :param name: Имя функции.
        :param sel: Выборка, нр, 15145; 12,13.
        :param value: Значение, нр, name=Промплощадка: изм name; pg=qn*2+10.
        :return: информация
        """
        name = name.lower()
        if 'уд' in name:
            return self.cor(keys=sel, values='del', del_all=('*' in name), print_log=True)
        elif 'изм' in name:
            return self.cor(keys=sel, values=value, print_log=True)
        elif 'импорт' in name:
            self.txt_import_rm(type_import=sel, description=value)
        elif 'снять' in name:
            return self.cor(keys='(node); (vetv); (Generator)', values='sel=0', print_log=True)
        elif 'расчет' in name:
            self.rgm(txt='txt_task_cor')
            return 'выполнен расчет режима'
        elif 'добавить' in name:
            self.table_add_row(table=sel, tasks=value)
        elif 'текст' in name:
            self.txt_field_right(table_field=sel)
            return "\tИсправить пробелы, заменить английские буквы на русские."
        elif 'схн' in name:
            self.shn(choice=sel)
        elif 'сечение' in name:
            sel = sel.replace(' ', '')
            for i in ['ns:', 'psech:', 'выбор:', 'тип:']:
                if i not in sel:
                    raise ValueError(f'В задании "сечение": {sel!r} отсутствует ключ {i!r}')
            sel = sel.split(';')
            sd = {}
            for _ in sel:
                key, val = _.split(':')
                sd[key] = val
            self.loading_section(ns=sd['ns'], p_new=sd['psech'], type_correction=sd['тип'])
        elif 'напряжения' in name:
            self.voltage_nominal(choice=sel, edit=True)
            self.voltage_error(choice=sel, edit=True)

        elif 'отключения' in name:
            if sel == '-':
                self.network_analysis(disable_on=False)
            else:
                self.network_analysis(selection_node_for_disable=sel)
        elif 'скрм' in name:
            if 'скрм*' in name:
                self.all_auto_shunt = self.auto_shunt_rec(selection=sel)
            else:
                self.all_auto_shunt = self.auto_shunt_rec(selection=sel, only_auto_bsh=True)
            self.auto_shunt_cor(all_auto_shunt=self.all_auto_shunt)
        else:
            raise ValueError(f'Задание {name=} не распознано ({sel=}, {value=})')
        return ''

    def txt_import_rm(self, type_import: str, description: str):
        """
        Импорт данных из РМ.
        :param type_import: Если 'папка', то переносить данные из одноименных файлов в указанной папке,
         'файл' - из указанного файла
        :param description: "(I:\pop);таблица:node; тип:2; поле: pn,qn; выборка:"
        :return:
        """
        description_dict = {}
        path = re.search(re.compile(r"\((.+)\)"), description)[1]
        description_list = description.replace(path, '').replace(' ', '').split(';')
        dict_name = {'таблица': 'tables', 'тип': 'calc', 'поле': 'param', 'выборка': 'sel'}

        for i in description_list:
            if ':' in i:
                key, val = i.split(':')
                for x in dict_name:
                    if x in key:
                        description_dict[dict_name[x]] = val

        if type_import == 'папка':
            file_name = path + '\\' + self.Name
            if os.path.isfile(file_name):
                path = file_name
        if os.path.isfile(path):
            ifm = ImportFromModel(import_file_name=path, **description_dict)
            ifm.import_data_in_rm(rm=self)
        else:
            if type_import == 'файл':
                raise ValueError(f'Файл для импорта не найден {path}')
            if type_import == 'папка':
                log.error(f'Файл для импорта не найден {path}')

    def loading_section(self, ns: int, p_new: Union[float, str], type_correction: str = 'pg'):
        """
        Изменить переток мощности в сечении номер ns до величины p_new за счет изменения нагрузки('pn') или
        генерации ('qn') в отмеченных узлах и генераторах
        :param ns: номер сечения
        :param p_new:
        :param type_correction:  'pn' изменения нагрузки или 'pg' генерации
        """
        # --------------настройки----------
        choice = 'sel&!sta'
        max_cycle = 30  # максимальное количество циклов
        accuracy = 0.05  # процент, точность задания мощности сечения, но не превышает заданную
        dr_p_zad = 0.01  # величина реакции начальная

        log.info(f'\tИзменить переток мощности в сечении {ns=}: P={p_new}, выборка: {choice}, тип: {type_correction}.')
        if self.rastr.tables.Find("sechen") == -1:
            self.downloading_additional_files(['sch'])

        index_ns = self.index(table_name='sechen', key_int=ns)
        if index_ns == -1:
            raise ValueError(f'Сечение {ns=} отсутствует в файле сечений.')
        grline = self.rastr.Tables("grline")
        sechen = self.rastr.tables('sechen')
        name_ns = sechen.cols('name').ZS(index_ns)
        if p_new in ['pmax', 'pmin']:
            p_new = sechen.cols(p_new).Z(index_ns)
        try:
            p_new = float(p_new)
        except ValueError:
            raise ValueError(f'Заданная величина перетока мощности не распознано {p_new!r}')
        if not p_new:
            p_new = 0.01
        p_current = round(self.rastr.Calc("sum", "sechen", "psech", f"ns={ns}"), 2)
        log.info(f'\tТекущий переток мощности в сечении {name_ns!r}: {p_current}.')

        self.rastr.sensiv_start("")
        grline.SetSel(f'ns={ns}')
        index_grline = grline.FindNextSel(-1)
        while not index_grline == -1:
            self.rastr.sensiv_back(4, 1., grline.Cols("ip").Z(index_grline), grline.Cols("iq").Z(index_grline), 0)
            index_grline = grline.FindNextSel(index_grline)

        self.rastr.sensiv_write("")
        self.rastr.sensiv_end()

        node = self.rastr.tables("node")
        change_p = round(p_new - p_current, 2)
        db = 0
        # 'pn'
        p_sum = 0
        dr_p_sum = 0
        # 'pg'
        node_all = {}

        if type_correction == 'pn':  # изменение нагрузки

            choice_dr_p = f"!sta & abs(dr_p) > {dr_p_zad}"  # !sta вкл
            db = abs(self.rastr.Calc("sum", "node", "dr_p", choice_dr_p + "&dr_p>0"))
            db += abs(self.rastr.Calc("sum", "node", "dr_p", choice_dr_p + "&dr_p<0"))

        elif type_correction == 'pg':

            if self.rastr.Tables("Generator").cols.Find("sel") < 0:
                log.info('В таблицу Generator добавляется отсутствующее поле sel')
                self.rastr.Tables("Generator").Cols.Add('sel', 3)

            # Доотметить узлы и генераторы которые нужно корректировать
            # отметить генераторы у отмеченных узлов
            node.SetSel("sel")
            i = node.FindNextSel(-1)
            while i >= 0:
                self.group_cor(tabl="Generator", param="sel", selection=f"Node={node.cols('ny').ZS(i)}", formula="1")
                i = node.FindNextSel(i)
            # отметить узлы у отмеченных генераторов
            generators = self.rastr.tables("Generator")
            generators.SetSel("sel")
            i = generators.FindNextSel(-1)
            while i >= 0:
                self.group_cor(tabl="node", param="sel", selection=f"ny={generators.cols('Node').ZS(i)}", formula="1")
                i = generators.FindNextSel(i)
            choice_dr_p = f"tip>1 &!sta & abs(dr_p) > {dr_p_zad}"  # tip>1 ген   !sta вкл
            db = abs(self.rastr.Calc("sum", "node", "dr_p", choice_dr_p + "&dr_p>0"))
            db += abs(self.rastr.Calc("sum", "node", "dr_p", choice_dr_p + "&dr_p<0"))

            node.SetSel(choice)
            i = node.FindNextSel(-1)

            while i >= 0:
                nd = NodeGeneration(rastr=self.rastr, i=i)
                node_all[node.cols("ny").Z(i)] = nd
                i = node.FindNextSel(i)

        if db < dr_p_zad:
            log.error("Невозможно изменить мощность по сечению (с учетом отмеченных узлов и/или генераторов)")
            return False

        for cycle in range(max_cycle):

            p_current = round(self.rastr.Calc("sum", "sechen", "psech", f"ns={ns}"), 2)
            change_p = round(p_new - p_current, 2)
            log.debug(f'\t{cycle=}, {p_current=}, {p_new=}, {change_p=} МВт ({round(abs(change_p / p_new) * 100)} %)')

            if abs(change_p / p_new) * 100 < accuracy:
                if (p_current < p_new and p_new > 0) or (p_current > p_new and p_new < 0):
                    log.info(f'\tЗаданная точность достигнута P={p_current},'
                             f' отклонение {change_p}. {cycle + 1} итераций')
                    break

            # изменение нагрузки
            if type_correction == 'pn':
                node.SetSel(choice)
                i = node.FindNextSel(-1)
                while not i == -1:
                    p_sum += node.cols("pn").Z(i)
                    dr_p_sum += node.cols("dr_p").Z(i)
                    i = node.FindNextSel(i)
                if not p_sum:
                    log.error('Изменение мощности сечения: сумма нагрузки узлов равна 0')
                    break
                if dr_p_sum < 0:
                    coefficient = 1 + (1 - (p_sum - change_p) / p_sum)
                else:
                    coefficient = (p_sum - change_p) / p_sum
                node.cols("pn").Calc(f"pn*({coefficient})")
                node.cols("qn").Calc(f"qn*({coefficient})")

            # изменение генерации
            elif type_correction == 'pg':
                NodeGeneration.change_p = change_p
                section_up_sum = 0
                section_down_sum = 0
                for nd in node_all:
                    if nd.use:
                        nd.reserve_p()
                        if change_p * nd.dr_p > 0:
                            if nd.reserve_p_up:
                                section_up_sum += nd.reserve_p_up
                                nd.up_pgen = True
                        elif change_p * nd.dr_p < 0:
                            if nd.reserve_p_down:
                                section_down_sum += nd.reserve_p_down
                                nd.up_pgen = False
                log.debug(f'')
                if not (section_up_sum and section_down_sum):
                    log.error(f'Не удалось добиться заданной точности в сечении')

                # на сколько МВт нужно снизить Р
                reduce_p = abs(section_down_sum / (section_down_sum + section_up_sum) * change_p)
                if section_down_sum < reduce_p:
                    reduce_p = section_down_sum
                # на сколько МВт нужно увеличить Р
                increase_p = abs(abs(change_p) - reduce_p)
                if section_up_sum < increase_p:
                    increase_p = section_up_sum

                if (section_down_sum + section_up_sum) < change_p:
                    log.info("Генерации не хватает")
                # Коэффициент на сколько нужно умножить резерв Рген и прибавить к резерву Рген, для снижения генерации
                koef_p_down = 0
                # Коэффициент: на сколько нужно умножить резерв Рген и прибавить его к резерву Рген,
                # для увеличения генерации
                koef_p_up = 0
                if section_down_sum:
                    koef_p_down = 1 - (section_down_sum - reduce_p) / section_down_sum
                if section_up_sum:
                    koef_p_up = 1 - (section_up_sum - increase_p) / section_up_sum

                for nd in node_all:
                    if nd.use:
                        nd.change(koef_p_down=koef_p_down, koef_p_up=koef_p_up)

            self.rgm('loading_section')
        else:
            log.info(f'Заданная точность не достигнута P={p_current}, отклонение {change_p}.')

    def node_include(self) -> str:
        """
        Восстановление питания узлов путем включения выключателей (r<0.011 & x<0.011).
        :return: информация о включенных узлах
        """
        # self.ny_join_vetv
        log.debug('Восстановление питания отключенных узлов.')
        node_info = ''
        node_all = set()
        node_include = set()

        for ny, sta, pn, qn, pg, qg in self.rastr.tables('node').writesafearray('ny,sta,pn,qn,pg,qg', "000"):
            if not self.t_sta['node'][ny] and sta and (pn or qn or pg or qg):
                node_all.add((ny, self.t_name["node"][ny]))

                for s_key in self.ny_join_vetv[ny]:
                    r, x, _ = self.v_rxb[s_key]
                    if r < 0.011 and x < 0.011:
                        ny_connectivity = s_key[0] if ny != s_key[0] else s_key[1]
                        ndx = self.t_i['node'][ny_connectivity]
                        if not self.rastr.tables('node').Cols("sta").Z(ndx):  # Питающий узел включен.
                            # Включить узел и ветвь
                            self.rastr.tables('node').Cols("sta").SetZ(self.t_i['node'][ny], False)
                            self.rastr.tables('vetv').Cols("sta").SetZ(self.t_i['vetv'][s_key], False)
                            node_include.add((ny, self.t_name["node"][ny]))
                            break

        if node_include:
            node_info = "Восстановлено питание узлов:"
            for ny, name in node_include:
                node_info += f' {name} ({ny}),'
            node_info = node_info.strip(',') + ". "

        node_not_include = node_all - node_include
        if node_not_include:
            node_info += "Узлы, оставшиеся без питания:"
            for ny, name in node_not_include:
                node_info += f' {name} ({ny}),'
            node_info = node_info.strip(',') + ". "

        if node_info:
            log.info('\tnode_include: ' + node_info)
        return node_info.strip()

class NodeGeneration:
    """Класс для хранения информации об узле для изменения мощности в сечении."""
    dr_p_koeff = 0  # если 1, то умножаем дополнительно на dr_p в этом случае больше загружаются
    # генераторы которые меньше влияют на изменение мощности в сечении

    no_pmin = True  # ' не учитывать Pmin
    abs_change_p = None  # todo что это?
    change_p = 0
    unbalance_p = 0

    def __init__(self, i: int, rastr):
        """
        :param i: Индекс в таблице узлы
        :param rastr: 
        """
        self.gen_available = False  # Узел с генераторами
        self.use = True
        self.up_pgen = True
        self.reserve_p_up = 0
        self.reserve_p_down = 0
        self.rastr = rastr
        self.i = i
        self.node_t = self.rastr.tables("node")
        gen_t = self.rastr.tables("Generator")
        self.ny = self.node_t.Cols("ny").Z(self.i)
        self.dr_p = self.node_t.Cols("dr_p").Z(self.i)
        self.gen_all = {}
        dr_p = self.node_t.Cols("dr_p").Z(self.i)
        self.name = self.node_t.Cols("name").ZS(self.i)
        txt = f'\t\tУзел {self.ny}: {self.name}'
        gen_t.SetSel(f"Node={self.ny}")
        if gen_t.count:
            gen_t.SetSel(f"Node={self.ny}&sel")  # все генераторы дб отмечены, если не отмечен то не используем
            i = gen_t.FindNextSel(-1)
            while i >= 0:  # ЦИКЛ ген
                self.gen_available = True  # узел с генераторами
                gen = Gen(rastr=self.rastr, i=i)
                self.gen_all[gen.Num] = gen
                i = gen_t.FindNextSel(i)
        else:
            self.pg_max = self.node_t.Cols("pg_max").Z(self.i)
            self.pg_min = self.node_t.Cols("pg_min").Z(self.i)

    def reserve_p(self):
        self.reserve_p_up = 0
        self.reserve_p_down = 0
        if self.gen_available:
            for gen in self.gen_all:
                if gen.use:
                    gen.reserve_p()
                    self.reserve_p_up += gen.reserve_p_up
                    self.reserve_p_down += gen.reserve_p_down
        else:
            if self.pg_max:
                self.reserve_p_up = self.pg_max - self.node_t.Cols("pg").Z(self.i)
            else:
                log.info(f"в узле {self.ny} {self.name} не задано поле pg_max")
            self.reserve_p_down = self.node_t.Cols("pg").Z(self.i)

    def change(self, koef_p_down: float = 0, koef_p_up: float = 0):
        # --------------настройки----------
        change_p = abs(NodeGeneration.abs_change_p)
        unbalance_p = NodeGeneration.unbalance_p
        pg_node = self.node_t.Cols("pg").Z(self.i)

        if self.up_pgen:
            deviation_pg = koef_p_up * self.reserve_p_up  # На сколько нужно изменить генерацию в узле
        else:
            deviation_pg = pg_node * koef_p_down

        if not deviation_pg:
            return False

        if unbalance_p > 0:
            if unbalance_p > deviation_pg:
                unbalance_p = unbalance_p - deviation_pg
                deviation_pg = 0
            if unbalance_p < deviation_pg:
                deviation_pg = deviation_pg - unbalance_p
                unbalance_p = 0

        if not self.gen_available:  # нет генераторов
            if self.up_pgen:  # увеличиваем генерацию узла, koef_p_up
                if self.pg_max and self.pg_max > pg_node:
                    if self.pg_min > pg_node + deviation_pg:  # (от 0 до pg_min)
                        if self.pg_min and not self.no_pmin:  # если есть Рмин и учитываем Рмин то
                            if change_p > self.pg_min:
                                self.node_t.cols.Item("pg").SetZ(self.i, self.pg_min)
                                # unbalance_p = unbalance_p + (self.pg_min - deviation_pg)
                                change_p = change_p - self.pg_min
                        else:  # нет Рмин или не учитываем Рмин
                            self.node_t.cols.Item("pg").SetZ(self.i, pg_node + deviation_pg)
                            change_p = change_p - deviation_pg
                    elif self.pg_max > pg_node + deviation_pg and (
                            self.pg_min < pg_node + deviation_pg or self.pg_min == pg_node + deviation_pg):
                        # (от pg_min (включительно) до pg_max)v
                        self.node_t.cols.Item("pg").SetZ(self.i, pg_node + deviation_pg)
                        change_p = change_p - deviation_pg
                    elif self.pg_max < pg_node + deviation_pg or self.pg_max == pg_node + deviation_pg:
                        # (больше или равно pg_max)
                        self.node_t.cols.Item("pg").SetZ(self.i, self.pg_max)
                        change_p = change_p - (self.pg_max - pg_node)

            else:  # снижаем генерацию узла,KefPG_Down
                if self.pg_min < pg_node - deviation_pg or self.pg_min == pg_node - deviation_pg:
                    # (от pg_min (включительно) до pg_node)
                    self.node_t.cols.Item("pg").SetZ(self.i, pg_node - deviation_pg)
                    change_p = change_p - deviation_pg

                elif self.pg_min > pg_node - deviation_pg and (pg_node - deviation_pg) > 0:  # (от 0 до pg_min)
                    if self.pg_min > 0 and not self.no_pmin:  # если есть Рмин и учитываем Рмин то
                        self.node_t.cols.Item("pg").SetZ(self.i, self.pg_min)
                        change_p = change_p - (pg_node - self.pg_min)
                        deviation_pg = deviation_pg - (pg_node - self.pg_min)
                        if change_p > self.pg_min:
                            self.node_t.cols.Item("sta").SetZ(self.i, True)
                            # unbalance_p = unbalance_p + (self.pg_min - deviation_pg)
                            change_p = change_p - self.pg_min

                    else:  # если Рмин не учитываем
                        self.node_t.cols.Item("pg").SetZ(self.i, pg_node - deviation_pg)
                        change_p = change_p - deviation_pg

                elif pg_node - deviation_pg < 0 or pg_node == deviation_pg:  # (меньше или равно 0)
                    self.node_t.cols.Item("pg").SetZ(self.i, 0)
                    change_p = change_p - pg_node


class Gen:
    """Класс для хранения информации о генераторах в узле для изменения мощности в сечении."""

    def __init__(self, i: int, rastr):
        self.reserve_p_up = 0
        self.reserve_p_down = 0
        self.use = True
        self.rastr = rastr
        self.i = i
        self.gen_t = self.rastr.tables("Generator")
        self.Num = self.gen_t.Cols("Num").Z(self.i)
        self.gen_name = self.gen_t.Cols("Name").Z(self.i)
        self.Pmax = self.gen_t.Cols("Pmax").Z(self.i)
        if not self.Pmax:
            log.debug(f"У генератора {self.Num!r}: {self.gen_name!r}  не задано Pmax")
        self.Pmin = self.gen_t.Cols("Pmin").Z(self.i)

    def reserve_p(self):
        if self.gen_t.Cols("sta").Z(self.i):
            self.reserve_p_up = self.Pmax
        else:  # если генератор включен
            self.reserve_p_down = self.gen_t.Cols("P").Z(self.i)
            if self.Pmax:
                self.reserve_p_up = self.Pmax - self.gen_t.Cols("P").Z(self.i)


class CorSheet:
    """
    Клас лист для хранения листов книги excel и работы с ними.
    """
    SHAPE = {"Параметры импорта из файлов RastrWin": 'import_model',  # Импорт из моделей(ИМ)
             'Выполнить изменение модели по строкам': 'list_cor',  # Строковая форма(СФ)
             'Имя таблицы:': 'table_import'}  # Импорт таблиц(ИТ)

    def __init__(self, name: str, obj):
        """
        :param name: Имя листа
        :param obj: Объект лист
        """
        #  Сводная таблица корректировок 'tab_cor', нр корр потребления или нагрузка узлов / имя файла.
        #  Таблица корректировок по списку 'list_cor', нр изм, удалить, снять отметку.
        # 'import_model' - Импорт моделей.
        # 'table_import' - Импорт таблиц(ИТ).
        self.name = name
        self.xls = obj  # xls.cell(1,1).value [строки][столбцы] xls.max_row xls.max_column
        self.import_model_all = []  # Для хранения объектов ImportFromModel
        self.calc_val = self.xls.cell(1, 1).value
        if isinstance(self.calc_val, int):
            self.type = 'tab_cor'
        else:
            try:
                self.type = self.SHAPE[self.calc_val]
            except KeyError:
                raise ValueError(f'Тип задания листа {name!r} не распознан.')

    def run_sheets(self, rm: RastrModel):
        log.info(f'\tВыполнение задания листа {self.name!r}')
        if self.type == 'import_model':
            self.import_model(rm)
        elif self.type == 'list_cor':
            self.list_cor(rm)
        elif self.type == 'tab_cor':
            self.tab_cor(rm)
        elif self.type == 'table_import':
            self.tab_import(rm)
        log.info(f'\tКонец выполнения задания листа {self.name!r}\n')

    def export_model(self):
        """"Экспорт из моделей"""
        for row in range(3, self.xls.max_row + 1):
            if self.xls.cell(row, 1).value and '#' not in self.xls.cell(row, 1).value:
                """ ИД для импорта из модели(выполняется после блока начала)"""
                ifm = ImportFromModel(import_file_name=self.xls.cell(row, 1).value,
                                      criterion_start={"years": self.xls.cell(row, 6).value,
                                                       "season": self.xls.cell(row, 7).value,
                                                       "max_min": self.xls.cell(row, 8).value,
                                                       "add_name": self.xls.cell(row, 9).value},
                                      tables=self.xls.cell(row, 2).value,
                                      param=self.xls.cell(row, 4).value,
                                      sel=self.xls.cell(row, 3).value,
                                      calc=self.xls.cell(row, 5).value)
                self.import_model_all.append(ifm)

    def tab_import(self, rm: RastrModel) -> None:
        """Импорт таблицы из XL в Rastr"""
        tables_name = self.xls.cell(2, 1).value
        type_import = self.xls.cell(4, 1).value
        field_import = []
        field_column = []
        for column in range(2, self.xls.max_column + 1):
            if self.xls.cell(1, column).value:
                field_column.append(column)
                field_import.append(self.xls.cell(1, column).value)
        field_import = ','.join(field_import)
        data = [[self.xls.cell(row, col).value for col in field_column] for row in range(2, self.xls.max_row + 1)]
        table = rm.rastr.Tables(tables_name)
        table.ReadSafeArray(type_import, field_import, data)

    def import_model(self, rm: RastrModel) -> None:
        """Импорт в модели"""
        if self.import_model_all:
            for im in self.import_model_all:
                im.import_data_in_rm(rm)

    def list_cor(self, rm: RastrModel) -> None:
        """
        Таблица корректировок по списку, нр изм, удалить, снять отметку.
        """
        # номера столбцов
        C_SELECTION = 2
        C_VALUE = 3

        for row in range(3, self.xls.max_row + 1):
            name_fun = self.xls.cell(row, 1).value
            if name_fun:
                if '#' not in name_fun:
                    sel = str(self.xls.cell(row, C_SELECTION).value)
                    value = self.xls.cell(row, C_VALUE).value
                    year = self.xls.cell(row, 4).value
                    season = self.xls.cell(row, 5).value
                    max_min = self.xls.cell(row, 6).value
                    add_name = self.xls.cell(row, 7).value
                    statement = self.xls.cell(row, 8).value

                    if any([year, season, max_min, add_name]) and rm.code_name_rg2:  # any если хотя бы один истина
                        if not rm.test_name(condition={"years": year, "season": season,
                                                       "max_min": max_min, "add_name": add_name},
                                            info=f'\t\tcor_x:{sel=}, {value=}'):
                            continue
                    if statement:
                        if not rm.conditions_test(statement):
                            continue
                    log.info(rm.txt_task_cor(name=name_fun, sel=sel, value=value))

    def tab_cor(self, rm: RastrModel) -> None:
        """
        Корректировка моделей по заданию в табличном виде
        """
        name_files = ""
        dict_param_column = {}  # {10: "pn"}-столбец: параметр
        # Шаг по колонкам и запись в словарь всех столбцов для коррекции
        for column_name_file in range(2, self.xls.max_column + 1):
            if self.xls.cell(1, column_name_file).value not in ["", None]:
                name_files = self.xls.cell(1, column_name_file).value.split("|")
            if self.xls.cell(2, column_name_file).value:
                duct_add = False
                for name_file in name_files:
                    if name_file in [rm.name_base, "*"]:
                        duct_add = True
                    if "*" in name_file and len(name_file) > 7:
                        pattern_name = re.compile(r"\[(.*)\].*\[(.*)\].*\[(.*)\].*\[(.*)\]")
                        match = re.search(pattern_name, name_file)
                        if match.re.groups == 4 and rm.code_name_rg2:
                            if rm.test_name(condition={"years": match[1], "season": match[2],
                                                       "max_min": match[3], "add_name": match[4]},
                                            info=f"\tcor_xl, условие: {name_file}, ") or not rm.code_name_rg2:
                                duct_add = True
                if duct_add:
                    _ = self.xls.cell(2, column_name_file).value
                    dict_param_column[column_name_file] = _.replace(' ', '')

        if not dict_param_column:
            log.info(f"\t {rm.name_base} НЕ НАЙДЕН на листе {self.name} книги excel")
        else:
            log.info(f'\t\tРасчетной модели соответствуют столбцы: параметры {dict_param_column}')
            calc_vals = {1: "ЗАМЕНИТЬ", 2: "+", 3: "-", 0: "*"}
            # 1: "ЗАМЕНИТЬ", 2: "ПРИБАВИТЬ", 3: "ВЫЧЕСТЬ", 0: "УМНОЖИТЬ"
            for row in range(3, self.xls.max_row + 1):
                for column, param in dict_param_column.items():
                    short_key = self.xls.cell(row, 1).value
                    if short_key not in [None, ""]:
                        new_val = self.xls.cell(row, column).value
                        if new_val is not None:
                            if param not in ["pop", "pp"]:
                                if self.calc_val == 1:
                                    rm.cor(keys=str(short_key),
                                           values=f"{param}={new_val}",
                                           print_log=True)
                                else:
                                    rm.cor(keys=str(short_key),
                                           values=f"{param}={param}{calc_vals[self.calc_val]}{new_val}",
                                           print_log=True)
                            else:
                                rm.cor_pop(zone=short_key, new_pop=new_val)  # изменить потребление


class CorXL:
    """
    Изменить параметры модели по заданию в таблице excel.
    """

    def __init__(self, excel_file_name: str, sheets: str) -> None:
        """
        Проверить наличие книги и листов, создать классы CorSheet для листов.
        :param excel_file_name: Полное имя файла excel;
        :param sheets: имя листов, нр [импорт из моделей][XL->RastrWin], если '*', то все листы по порядку
        """
        self.sheets_list = []  # для хранения объектов CorSheet
        log.info(f"Изменить модели по заданию из книги: {excel_file_name}, листы: {sheets}")
        if not os.path.exists(excel_file_name):
            raise ValueError("Ошибка в задании, не найден файл: " + excel_file_name)
        else:
            self.excel_file_name = excel_file_name
            # data_only - Загружать расчетные значения ячеек, иначе будут формулы.
            self.wb = load_workbook(excel_file_name, data_only=True)

            if sheets == '*':  # все листы
                self.sheets = self.wb.sheetnames
                for sheet in self.sheets:
                    if '#' not in sheet:  # все листы
                        self.sheets_list.append(CorSheet(name=sheet, obj=self.wb[sheet]))
            else:
                self.sheets = re.findall(r"\[(.+?)\]", sheets)
                for sheet in self.sheets:
                    if sheet not in self.wb.sheetnames:
                        raise ValueError(f"Ошибка в задании, не найден лист: {sheet} в файле {excel_file_name}")
                    else:
                        self.sheets_list.append(CorSheet(name=sheet, obj=self.wb[sheet]))

    def init_export_model(self) -> None:
        """Экспорт данных из растр в csv"""
        for sheet in self.sheets_list:
            if sheet.type == 'import_model':
                sheet.export_model()

    def run_xl(self, rm: RastrModel) -> None:
        """Запуск всех корректировок"""
        for sheet in self.sheets_list:
            sheet.run_sheets(rm)


class ImportFromModel:
    # __slots__ = 'set_import_model', 'calc_str'
    set_import_model = []  # хранение объектов класса ImportFromModel созданных в GUI и коде
    calc_str = {"обновить": 2, "загрузить": 1, "присоединить": 0, "присоединить-обновить": 3, "объединить": 3}
    number = 0  # для создания уникального имени csv файла

    def __init__(self,
                 import_file_name: str,
                 criterion_start: Union[dict, None] = None,
                 tables: str = '',
                 param='',
                 sel: Union[str, None] = '',
                 calc: Union[int, str] = '2',
                 way='array'):
        """
        Импорт данных из файлов РМ ('.rg2', '.rst' и др.) и сохранение их в экземпляре класса или в csv.
        :param import_file_name: Полное имя файла
        :param criterion_start: {"years": "","season": "","max_min": "", "add_name": ""} условие выполнения
        :param tables: таблица для импорта, нр "node, vetv"
        :param param: параметры для импорта: "" все параметры или перечисление, нр 'sel, sta'(ключи необязательно)
        :param sel: выборка нр "sel" или "" - все
        :param calc: число типа int, строка или ключевое слово:
        {"обновить": 2 , "загрузить": 1, "присоединить": 0, "присоединить-обновить": 3}
        :param way: 'csv' или 'array'
        'array' - Создает папку temp в папке с файлом и сохраняет в ней .csv файлы
        """
        self.way = way
        folder_temp = ''
        if not os.path.exists(import_file_name):
            raise ValueError("Ошибка в задании, не найден файл: " + import_file_name)
        else:
            log.info(f'Экспорт данных из файла "{import_file_name}".')
            self.import_file_name = import_file_name

            if way == 'csv':
                ImportFromModel.number += 1
                folder_temp = os.path.dirname(import_file_name) + '\\temp'
                if not os.path.exists(folder_temp):
                    log.debug(f'Создана папка {folder_temp}.')
                    os.mkdir(folder_temp)

            self.criterion_start = criterion_start
            self.sel = sel if sel else ''

            if type(calc) == int:
                self.calc = calc
            elif calc.isdigit():
                self.calc = int(calc)
            else:
                if calc in self.calc_str:
                    self.calc = self.calc_str[calc]
                else:
                    raise ValueError(f"ImportFromModel. Ошибка в задании, не распознано задание '{calc=}'.")

            self.param = []
            self.import_data = []  # if way == 'array': tuple(данных)
            self.import_csv_file = []  # if way == 'csv': полный путь к CSV
            self.tables = tables.replace(' ', '').split(",")  # разделить на ["таблицы"]

            import_rm = RastrModel(full_name=self.import_file_name)
            import_rm.load()

            for i, tabl in enumerate(self.tables):
                # Параметры
                if param:  # Добавить к строке параметров ключи текущей таблицы
                    self.param.append(param + ',' + import_rm.rastr.Tables(tabl).Key)
                else:  # если все параметры
                    self.param.append(import_rm.all_cols(tabl))

                log.info(f"\tТаблица: {tabl}, выборка: {self.sel}, параметры: {self.param[i]!r}.")
                tab = import_rm.rastr.Tables(tabl)
                tab.setsel(self.sel)
                if tab.count:
                    # Данные
                    if way == 'csv':
                        self.import_csv_file.append(f"{folder_temp}\\{os.path.basename(import_file_name)}_{tabl}_"
                                                    f"{ImportFromModel.number}.csv")
                        # Экспорт данных из файла в .csv файлы в папку temp
                        log.info(f"\tФайл CSV: {self.import_csv_file[i]!r}.")
                        tab.WriteCSV(1, self.import_csv_file[i], self.param[i], ";")  # 0 дописать, 1 заменить
                    elif way == 'array':
                        self.import_data.append(tab.writesafearray(self.param[i], "000"))

    def import_data_in_rm(self, rm: RastrModel) -> None:
        """
        Импорт данных в файлы
        """
        log.info(f"\tИмпорт из файла {self.import_file_name} в РМ.")
        if not rm.code_name_rg2 or rm.test_name(condition=self.criterion_start,
                                                info='\tImportFromModel '):
            for i, tab in enumerate(self.tables):
                log.info(f"\tТаблица: {self.tables[i]}, выборка: {self.sel}, тип: {self.calc}, "
                         f"параметры: {self.param[i]}.")
                rm_tab = rm.rastr.Tables(self.tables[i])

                if self.way == 'csv':
                    log.info(f"\tФайл CSV: {self.import_csv_file[i]}")
                    rm_tab.ReadCSV(self.calc, self.import_csv_file[i], self.param[i], ";", '')
                elif self.way == 'array':

                    set_param_in = set(rm.all_cols(self.tables[i], val_return='list'))
                    set_param_out = set(self.param[i].split(','))
                    delta = set_param_out - set_param_in
                    if delta:
                        data = pd.DataFrame(data=self.import_data[i], columns=self.param[i].split(','))
                        for field in delta:
                            data.drop(columns=field, inplace=True)
                        self.param[i] = ','.join(data.columns)
                        self.import_data[i] = tuple(data.itertuples(index=False, name=None))
                    if len(self.import_data) > i:
                        rm_tab.ReadSafeArray(self.calc, self.param[i], self.import_data[i])
        ImportFromModel.number = 0


class Automation:
    """
    Моделирование действия ПА
    """
    def __init__(self, rm: RastrModel):
        log.debug('Инициализация автоматики')
        self.n_action = {}
        self.df_automation = None
        if rm.rastr.tables.Find('automation') > -1:
            if len(rm.rastr.tables('automation')):
                self.df_automation = rm.df_from_table(table_name='automation')
                if rm.rastr.tables.Find('automation_pattern') > -1:
                    df_automation_pattern = rm.df_from_table(table_name='automation_pattern')
                    df_automation_pattern['name'] = df_automation_pattern['name'].str.strip()

                    df_automation_pattern.set_index('name', inplace=True)
                    dict_name_action = df_automation_pattern.to_dict()['pattern']

                    self.df_automation.replace({"action": dict_name_action}, inplace=True)
                    self.df_automation.replace({"condition": dict_name_action}, inplace=True)

    def scheme_description(self, number: str) -> tuple:
        """
        По номеру n в таблице automation возвращает строки задания в текстовом виде
        :param number: "Номер_ПА.номер_ступени"
        :return: (list(название из таблицы automation), list(задание из той же таблицы))
        """
        if number in self.n_action:
            return self.n_action[number][0], self.n_action[number][1]

        names = []
        tasks = []
        number = number.replace(' ', '')
        if '.' in number:
            n, step = number.split('.')
        else:
            n = number
            step = -1
        n = int(n)
        step = int(step)
        cut = self.df_automation[(self.df_automation['n'] == n) & (self.df_automation['sta'] == 0)]
        if step > -1:
            cut = cut[cut['step'] == step]

        if not len(cut):
            raise ValueError(f'В таблице automation отсутствует запись с номером {number!r}')

        if cut['action'].all():

            for i in cut.index:
                task = f"[{cut.loc[i, 'action']}]"
                if cut.loc[i, 'condition']:
                    task += f"{{{cut.loc[i, 'condition']}}}"

                tasks.append(task)
                names.append(cut.loc[i, 'name'])

            self.n_action[number] = (names, tasks)
            return names, tasks
        else:
            raise ValueError(f'В таблице automation в записи с номером {number!r} отсутствует описание действия.')

    def execute_action_pa(self, rm: RastrModel, df_init: pd.DataFrame) -> str:
        # не забыть про restore_only_state
        pass


class PrintXL:
    """Класс печать данных в excel"""
    short_name_tables = {'n': 'node',
                         'v': 'vetv',
                         'g': 'Generator',
                         'na': 'area',
                         'npa': 'area2',
                         'no': 'darea',
                         'nga': 'ngroup',
                         'ns': 'sechen'}

    #  ...._log  лист протокол для сводной

    def __init__(self, task):
        """
        Добавить листы и первая строка с названиями
        """
        self.name_xl_file = ''  # Имя файла EXCEL для сохранения
        self.data_table = {}  # Для хранения ссылок на листы excel {'имя листа=имя таблицы': fd c данными}
        self.data_parameters = pd.DataFrame()
        self.task = task
        self.book = Workbook()
        #  Создать лист xl и присвоить ссылку на него
        for name_table in self.task['set_printXL']:
            if self.task['set_printXL'][name_table]['add']:
                self.data_table[name_table] = pd.DataFrame()

        if self.task['print_parameters']['add']:
            self.set_output_parameters = set()
            for task_i in self.task['print_parameters']['sel'].replace(' ', '').split('/'):
                key_row, key_column = task_i.split(":")  # нр"8;9", "pn;qn"
                for col in key_column.split(';'):  # ['pn','qn']
                    for row in key_row.split(';'):  # ['15105,15113','15038,15037,4']
                        self.set_output_parameters.add(f'{row}:{col}')

        if self.task['print_balance_q']['add']:
            self.sheet_q = self.book.create_sheet("balance_Q")
            self.row_q = {}
            # (имя ключа, название в ячейке XL, комментарий ячейки)
            name_row = (
                ('row_name',
                 'Наименование', ''),
                ('row_qn',
                 'Реактивная мощность нагрузки', 'Calc(sum,area,qn,vibor)'),
                ('row_dq_sum',
                 'Нагрузочные потери', ''),
                ('row_dq_line',
                 'в т.ч. потери в ЛЭП', 'потери в ЛЭП: \nCalc(sum,area,dq_line,vibor)'),
                ('row_dq_tran',
                 'потери в трансформаторах', 'Calc(sum,area,dq_tran,vibor)'),
                ('row_shq_tran',
                 'потери Х.Х. в трансформаторах', 'Calc(sum,area,shq_tran,vibor)'),
                ('row_skrm_potr',
                 'Потребление реактивной мощности СКРМ (ШР, УШР, СК, СТК)',
                 'Calc(sum,node,qsh,qsh>0 & vibor) - Calc(sum,node,qg,qg<0&pg<0.1&pg>-0.1 & vibor)'),
                ('row_sum_port_Q',
                 'Суммарное потребление реактивной мощности', ''),
                ('row_qg',
                 'Генерация реактивной мощности электростанциями', 'Calc(sum,node,qg,(pg>0.1|pg<-0.1) & vibor)'),
                ('row_skrm_gen',
                 'Генерация реактивной мощности СКРМ (БСК, СК, СТК)', ''),
                ('row_qg_min',
                 'Минимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmin,pg>0.1& vibor)'),
                ('row_qg_max',
                 'Максимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmax,pg>0.1& vibor)'),
                ('row_shq_line',
                 'Зарядная мощность ЛЭП', 'Calc(sum,area,shq_line, vibor)'),
                ('row_sum_QG',
                 'Суммарная генерация реактивной мощности', ''),
                ('row_Q_itog',
                 'Внешний переток реактивной мощности (избыток/дефицит +/-)', ''),
                ('row_Q_itog_gmin',
                 'Внешний переток реактивной мощности при минимальной генерации '
                 'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)', ''),
                ('row_Q_itog_gmax',
                 'Внешний переток реактивной мощности при максимальной генерации '
                 'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)',
                 ''),
            )
            self.sheet_q.cell(1, 1, 'Таблица 1 - Баланс реактивной мощности, Мвар')
            for n, row_info in enumerate(name_row, 2):
                self.row_q[row_info[0]] = n
                self.sheet_q.cell(n, 1, row_info[1])
                if row_info[2]:
                    self.sheet_q.cell(n, 1).comment = Comment(row_info[2], '')

    def add_val(self, rm: RastrModel):
        log.info("\tВывод данных из моделей в XL")

        # Добавить значения в вывод таблиц.
        for name_table in self.data_table:
            # проверка наличия таблицы
            if rm.rastr.Tables.Find(name_table) < 0:
                if name_table == 'sechen':
                    rm.downloading_additional_files(['sch'])
            # Считать данные из таблиц растр.

            fields = self.task['set_printXL'][name_table]['par'].replace(' ', '')
            setsel = self.task['set_printXL'][name_table]['sel']
            if not fields:
                fields = rm.all_cols(name_table)

            data = rm.df_from_table(table_name=name_table, fields=fields, setsel=setsel)
            if not data.empty:
                self.data_table[name_table] = pd.concat([self.data_table[name_table],
                                                         data.apply(lambda x: rm.info_file, axis=1).join(other=data)])

        if self.task['print_parameters']['add']:
            self.add_val_parameters(rm)

        if self.task['print_balance_q']['add']:
            self.add_val_balance_q(rm)

    def add_val_parameters(self, rm):
        """
        Вывод заданных параметров в формате: "15105,15113;15038,15037,4:r;x;b / 15198:pg;qg / ns=1(sechen):psech".
        """
        if 'sechen' in self.task['print_parameters']['sel']:
            if rm.rastr.tables('sechen').Find < 0:
                rm.downloading_additional_files(['sch'])
        date = pd.Series(dtype='object')
        for i in self.set_output_parameters:
            k, p = i.split(':')
            table, sel = rm.recognize_key(k)
            if rm.rastr.tables(table).cols(p).Prop(1) == 2:  # если поле типа строка
                date.loc[i] = rm.txt_field_return(table, sel, p)
            else:
                date.loc[i] = rm.rastr.tables(table).cols.Item(p).ZS(rm.index(table_name=table, key_str=sel))
        date = pd.concat([date, rm.info_file])
        self.data_parameters = pd.concat([self.data_parameters, date], axis=1)

    def add_val_balance_q(self, rm):
        column = self.sheet_q.max_column + 1
        choice = self.task["print_balance_q"]["sel"]
        self.sheet_q.cell(2, column, f'{rm.season_name} {rm.god} г ({rm.additional_name})')
        area = rm.rastr.Tables("area")
        area.SetSel(self.task["print_balance_q"]["sel"])
        # ndx = area.FindNextSel(-1)

        # Реактивная мощность нагрузки
        address_qn = self.sheet_q.cell(self.row_q['row_qn'], column,
                                       rm.rastr.Calc("sum", "area", "qn", choice)).coordinate
        # Потери Q в ЛЭП
        address_dq_line = self.sheet_q.cell(self.row_q['row_dq_line'], column,
                                            rm.rastr.Calc("sum", "area", "dq_line", choice)).coordinate
        # Потери Q в трансформаторах
        address_dq_tran = self.sheet_q.cell(self.row_q['row_dq_tran'], column,
                                            rm.rastr.Calc("sum", "area", "dq_tran", choice)).coordinate
        # Потери Q_ХХ в трансформаторах
        address_shq_tran = self.sheet_q.cell(self.row_q['row_shq_tran'], column,
                                             rm.rastr.Calc("sum", "area", "shq_tran", choice)).coordinate
        # ШР УШР без бСК
        skrm = rm.rastr.Calc("sum", "node", "qsh", f"qsh>0&({choice})") - \
               rm.rastr.Calc("sum", "node", "qg", f"qg<0&pg<0.1&pg>-0.1&({choice})")
        address_SHR = self.sheet_q.cell(self.row_q['row_skrm_potr'], column, skrm).coordinate
        # Генерация Q генераторов
        address_qg = self.sheet_q.cell(self.row_q['row_qg'], column,
                                       rm.rastr.Calc("sum", "node", "qg", f"(pg>0.1|pg<-0.1)&({choice})")).coordinate
        # Генерация БСК шунтом и СТК СК
        address_skrm_gen = self.sheet_q.cell(self.row_q['row_skrm_gen'], column,
                                             -rm.rastr.Calc("sum", "node", "qsh", f"qsh<0&({choice})") + rm.rastr.Calc(
                                                 "sum", "node", "qg", f"qg>0&pg<0.1&pg>-0.1&({choice})")).coordinate
        # Минимальная генерация реактивной мощности в узлах выборки
        address_qg_min = self.sheet_q.cell(self.row_q['row_qg_min'], column,
                                           rm.rastr.Calc("sum", "node", "qmin", f"pg>0.1&({choice})")).coordinate
        # Максимальная генерация реактивной мощности в узлах выборки
        address_qg_max = self.sheet_q.cell(self.row_q['row_qg_max'], column,
                                           rm.rastr.Calc("sum", "node", "qmax", f"pg>0.1&({choice})")).coordinate
        # Генерация Q в ЛЭП
        address_shq_line = self.sheet_q.cell(self.row_q['row_shq_line'], column,
                                             - rm.rastr.Calc("sum", "area", "shq_line", choice)).coordinate
        address_losses = self.sheet_q.cell(self.row_q['row_dq_sum'], column,
                                           f"={address_dq_line}+{address_dq_tran}+{address_shq_tran}").coordinate
        address_load = self.sheet_q.cell(self.row_q['row_sum_port_Q'], column,
                                         f"={address_qn}+{address_losses}+{address_SHR}").coordinate
        address_sum_gen = self.sheet_q.cell(self.row_q['row_sum_QG'], column,
                                            f"={address_qg}+{address_shq_line}+{address_skrm_gen}").coordinate
        self.sheet_q.cell(self.row_q['row_Q_itog'], column,
                          f"=-{address_load}+{address_sum_gen}")
        self.sheet_q.cell(self.row_q['row_Q_itog_gmin'], column,
                          f"=-{address_load}+{address_qg_min}+{address_shq_line}")
        self.sheet_q.cell(self.row_q['row_Q_itog_gmax'], column,
                          f"=-{address_load}+{address_qg_max}+{address_shq_line}")

    def finish(self):
        """
        Преобразовать в объект таблицу и удалить листы с одной строкой.
        """

        self.name_xl_file = self.task['name_time'] + ' вывод данных.xlsx'
        self.book.save(self.name_xl_file)
        self.book = None

        for name_table, data in self.data_table.items():
            limitation = ''
            value_p = ''
            for lim in ['pmax', 'set_pop']:
                if lim in data.columns:
                    if len(data[data[lim] != 0]):
                        limitation = lim
                        break
            for val in ['psech', 'pop', 'pp']:
                if val in data.columns:
                    value_p = val
                    break
            if limitation and val:
                data.loc[data[limitation] != 0, 'difference_p'] = data.loc[data[limitation] != 0, value_p] - \
                                                                  data.loc[data[limitation] != 0, limitation]

            with pd.ExcelWriter(path=self.name_xl_file, mode='a', engine="openpyxl") as writer:
                data.to_excel(excel_writer=writer,
                              sheet_name=name_table,
                              header=True,
                              index=False)

        if self.task['print_parameters']['add']:
            with pd.ExcelWriter(path=self.name_xl_file, mode='a', engine="openpyxl") as writer:
                self.data_parameters.T.to_excel(excel_writer=writer,
                                                sheet_name='Значения',
                                                header=True,
                                                index=False)

        self.book = load_workbook(self.name_xl_file)
        for sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            if sheet.max_row < 2:
                del self.book[sheet_name]  # удалить пустой лист
            else:
                if sheet_name != "balance_Q":
                    PrintXL.create_table(sheet, sheet_name)  # Создать объект таблица.

        if self.task['print_balance_q']['add']:
            self.sheet_q = self.book['balance_Q']
            self.sheet_q.row_dimensions[2].height = 140
            self.sheet_q.column_dimensions['A'].width = 40
            thins = Side(border_style="thin", color="000000")
            for row in range(2, self.sheet_q.max_row + 1):
                for col in range(1, self.sheet_q.max_column + 1):
                    if row > 2 and col > 1:
                        self.sheet_q.cell(row, col).number_format = BUILTIN_FORMATS[1]
                    self.sheet_q.cell(row, col).border = Border(thins, thins, thins, thins)
                    self.sheet_q.cell(row, col).font = Font(name='Times New Roman', size=11)
                    if row == 2:
                        self.sheet_q.cell(row, col).alignment = Alignment(text_rotation=90,
                                                                          wrap_text=True, horizontal="center")
                    if col == 1:
                        self.sheet_q.cell(row, col).alignment = Alignment(wrap_text=True)
                    if row in [12, 13, 17, 18]:
                        self.sheet_q.cell(row, col).fill = PatternFill('solid', fgColor="00FF0000")
                    if row in [9, 15, 16]:
                        self.sheet_q.cell(row, col).font = Font(bold=True)

        self.book.save(self.name_xl_file)
        self.book = None

        # Открыть excel через win32com.client и создать сводные.
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.ScreenUpdating = False  # обновление экрана
        # excel.Calculation = -4135  # xlCalculationManual
        excel.EnableEvents = False  # отслеживание событий
        excel.StatusBar = False  # отображение информации в строке статуса excel
        try:
            self.book = excel.Workbooks.Open(self.name_xl_file)
        except Exception:
            raise Exception(f'Ошибка при открытии файла {self.name_xl_file=}')

        for name_sheet in self.data_table:
            rows = self.task['set_printXL'][name_sheet]['rows'].split(",")
            rows = list(set(rows) & set(self.data_table[name_sheet].columns))
            columns = self.task['set_printXL'][name_sheet]['columns'].split(",")
            columns = list(set(columns) & set(self.data_table[name_sheet].columns))
            values = self.task['set_printXL'][name_sheet]['values'].split(",")
            values = list(set(values) & set(self.data_table[name_sheet].columns))

            tab_log = self.book.sheets[name_sheet].ListObjects[0]
            name_pivot_sheet = name_sheet + '_сводная'
            pivot_sheet = self.book.Sheets.Add(After=name_sheet)
            pivot_sheet.Name = name_pivot_sheet

            pt_cache = self.book.PivotCaches().add(1, tab_log)  # создать КЭШ xlDatabase, ListObjects
            pt = pt_cache.CreatePivotTable(TableDestination=name_pivot_sheet + "!R1C1",
                                           TableName="Сводная_" + name_sheet)  # создать сводную таблицу
            pt.ManualUpdate = True  # не обновить сводную
            pt.AddFields(RowFields=rows,
                         ColumnFields=columns,
                         PageFields=["Имя файла"],
                         AddToTable=False)

            for val in values:
                pt.AddDataField(pt.PivotFields(val),
                                val + " ",
                                -4157)  # xlMax -4136 xlSum -4157
                # Field                      Caption             def формула расчета
                pt.PivotFields(val + " ").NumberFormat = "0"

            # .PivotFields("na").ShowDetail = True #  группировка
            pt.RowAxisLayout(1)  # xlTabularRow показывать в табличной форме!!!!
            if len(values) > 1:
                pt.DataPivotField.Orientation = 1  # xlRowField"Значения в столбцах или строках xlColumnField

            # .DataPivotField.Position = 1 #  позиция в строках
            pt.RowGrand = False  # удалить строку общих итогов
            pt.ColumnGrand = False  # удалить столбец общих итогов
            pt.MergeLabels = True  # объединять одинаковые ячейки
            pt.HasAutoFormat = False  # не обновлять ширину при обновлении
            pt.NullString = "--"  # заменять пустые ячейки
            pt.PreserveFormatting = False  # сохранять формат ячеек при обновлении
            pt.ShowDrillIndicators = False  # показывать кнопки свертывания
            # pt.PivotCache.MissingItemsLimit = 0 # xlMissingItemsNone
            # xlMissingItemsNone для норм отображения уникальных значений автофильтра
            for row in rows:
                pt.PivotFields(row).Subtotals = [False, False, False, False, False, False, False, False,
                                                 False, False,
                                                 False, False]  # промежуточные итоги и фильтры
            for column in columns:
                pt.PivotFields(column).Subtotals = [False, False, False, False, False, False, False, False,
                                                    False, False,
                                                    False, False]  # промежуточные итоги и фильтры
            pt.ManualUpdate = False  # обновить сводную
            pt.TableStyle2 = ""  # стиль
            if name_sheet in ["area", "area2", "darea"]:
                pt.ColumnRange.ColumnWidth = 10  # ширина строк
                pt.RowRange.ColumnWidth = 9
                pt.RowRange.Columns(1).ColumnWidth = 7
                pt.RowRange.Columns(2).ColumnWidth = 20
                pt.RowRange.Columns(3).ColumnWidth = 20
                pt.RowRange.Columns(6).ColumnWidth = 20
            pt.DataBodyRange.HorizontalAlignment = -4108  # xlCenter
            # .DataBodyRange.NumberFormat = "#,##0"
            # формат
            pt.TableRange1.WrapText = True  # перенос текста в ячейке
            pt.TableRange1.Borders(7).LineStyle = 1  # лево
            pt.TableRange1.Borders(8).LineStyle = 1  # верх
            pt.TableRange1.Borders(9).LineStyle = 1  # низ
            pt.TableRange1.Borders(10).LineStyle = 1  # право
            pt.TableRange1.Borders(11).LineStyle = 1  # внутри вертикаль
            pt.TableRange1.Borders(12).LineStyle = 1  #

        self.book.Save()
        excel.Visible = True
        excel.ScreenUpdating = True  # обновление экрана
        excel.Calculation = -4105  # xlCalculationAutomatic
        excel.EnableEvents = True  # отслеживание событий
        excel.StatusBar = True  # отображение информации в строке статуса excel

    @staticmethod
    def create_table(sheet, sheet_name, point_start: str = 'A1'):
        """
        Создать объект таблица из всего диапазона листа.
        :param sheet: Объект лист excel
        :param sheet_name: Имя таблицы.
        :param point_start:
        """
        tab = Table(displayName=sheet_name,
                    ref=f'{point_start}:' + get_column_letter(sheet.max_column) + str(sheet.max_row))

        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                            showFirstColumn=False,
                                            showLastColumn=False,
                                            showRowStripes=True,
                                            showColumnStripes=True)
        sheet.add_table(tab)


def str_yeas_in_list(id_str: str):
    """
    Преобразует перечень годов.
    :param id_str: "2021,2023...2025"
    :return: [2021,2023,2024,2025] или []
    """
    years_list = id_str.replace(" ", "").split(',')
    if years_list:
        years_list_new = np.array([], int)
        for it in years_list:
            if "..." in it:
                i_years = it.split('...')
                years_list_new = np.hstack(
                    [years_list_new, np.array(np.arange(int(i_years[0]), int(i_years[1]) + 1), int)])
            else:
                years_list_new = np.hstack([years_list_new, int(it)])
        return np.sort(years_list_new)
    else:
        return []


def test_run(source):
    year = datetime.now().year
    month = datetime.now().month
    # try:
    #     i_date = datetime.strptime(urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8'),
    #                                "%Y-%m-%d %H:%M:%S").date()
    #
    #     year = i_date.year
    #     month = i_date.month
    # except:
    #     pass

    if year > 2023:
        if month > 1:
            if source:
                raise ValueError('Неизвестная ошибка.')


def my_except_hook(func):
    """
    Переназначить функцию для добавления информации об ошибке в диалоговое окно.
    :param func:
    :return:
    """

    def new_func(*args, **kwargs):
        log.error(f"Критическая ошибка: {args[0]}, {args[1]}", exc_info=True)
        mb.showerror("Ошибка", f"Критическая ошибка: {args[0]}, {args[1]}")
        # https://python-scripts.com/python-traceback
        func(*args, **kwargs)

    return new_func


if __name__ == '__main__':
    em = None  # глобальный объект класса EditModel
    cm = None  # глобальный объект класса CalcModel
    sys.excepthook = my_except_hook(sys.excepthook)

    # DEBUG, INFO, WARNING, ERROR и CRITICAL
    # logging.basicConfig(filename="log_file.log", level=logging.DEBUG, filemode='w',
    #                     format='%(asctime)s %(name)s  %(levelname)s:%(message)s')

    log = logging.getLogger(__name__)
    log.setLevel(logging.DEBUG)  # INFO DEBUG
    formatter = logging.Formatter('%(asctime)s %(name)s %(levelname)s:%(message)s')

    file_handler = logging.FileHandler(filename=GeneralSettings.log_file, mode='w')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    # file_handler.close()
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(formatter)

    log.addHandler(file_handler)
    log.addHandler(console_handler)

    app = QtWidgets.QApplication([])  # Новый экземпляр QApplication
    # app.setApplicationName("Правка моделей RastrWin")

    gui_choice_window = MainChoiceWindow()
    gui_choice_window.show()
    gui_calc_ur = CalcWindow()
    # gui_calc_ur.show()
    gui_calc_ur_set = CalcSetWindow()
    # gui_calc_ur_set.show()
    gui_edit = EditWindow()
    # gui_edit.show()
    gui_set = SetWindow()
    # gui_set.show()
    sys.exit(app.exec_())  # Запуск

# TODO дописать: сравнение файлов
# TODO спросить про перезапись файлов
# self.save(full_name_new=r'I:\rastr_add\test\result\1.rg2')

# TODO в иксель не более 1 048 576 строк и 16 384 столбца

# # import time
# start_time = time.time()
# for i in range(1000):
#     pass
#
# print((time.time() - start_time))

# setz - 0,5 мс лучше calc 0,8 мс - для 1 значения
# для корректировки многих значений эффективен calc, через ip.uhom в 5 раз медленнее
# calc в 4-10(узлы - ветви) раз быстрее чем writesafearray - ReadSafeArray
# writesafearray > ReadSafeArray на 10 %
# циклить через FindNextSel крайне медленно