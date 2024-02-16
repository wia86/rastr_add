"""Модуль для вывода параметров РМ в таблице excel."""
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd
import win32com.client

log_print_xl = logging.getLogger(f'__main__.{__name__}')


class PrintXL:
    """Класс печать данных в excel"""
    short_name_tables = {'n': 'node',
                         'v': 'vetv',
                         'g': 'Generator',
                         'na': 'area',
                         'npa': 'area2',
                         'no': 'darea',
                         'nga': 'ngroup',
                         'ns': 'sechen'}
    set_param = {
        "sechen": {"sel": 'ns>0',
                   'par': '',  # "ns,name,pmin,pmax,psech",
                   "rows": "ns,name",  # поля строк в сводной
                   "columns": "Год,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля столбцов в сводной
                   "values": "psech,pmax,difference_p"},
        "area": {"sel": 'na>0',
                 'par': '',  # 'na,name,no,pg,pn,pn_sum,dp,pop,set_pop,qn_sum,pg_max,pg_min,poq,qn,qg,dev_pop'
                 "rows": "na,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                 "columns": "Год",  # поля столбцов в сводной
                 "values": "pop,difference_p"},
        "area2": {"sel": 'npa>0',
                  'par': '',  # 'npa,name,pg,pn,dp,pop,vnp,qg,qn,dq,poq,vnq,pn_sum,qn_sum,set_pop,dev_pop'
                  "rows": "npa,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                  "columns": "Год",  # поля столбцов в сводной
                  "values": "pop,difference_p"},
        "darea": {"sel": 'no>0',
                  'par': '',  # 'no,name,pg,pp,pvn,qn_sum,pnr_sum,pn_sum,set_pop,qvn,qp,qg,dev_pop',
                  "rows": "no,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3",  # поля строк в сводной
                  "columns": "Год",  # поля столбцов в сводной
                  "values": "pp,difference_p"}
    }

    def __init__(self, task):
        """
        Добавить листы и первая строка с названиями
        """
        self.name_xl_file = ''  # Имя файла EXCEL для сохранения
        self.data_table = {}  # Для хранения ссылок на листы excel {'имя листа=имя таблицы': fd c данными}
        self.data_parameters = pd.DataFrame()
        self.task = task
        for tb in self.task['set_printXL']:
            if tb in self.set_param:
                self.task['set_printXL'][tb] = self.task['set_printXL'][tb] | self.set_param[tb]

        self.book = Workbook()
        #  Создать лист xl и присвоить ссылку на него
        for name_table in self.task['set_printXL']:
            if self.task['set_printXL'][name_table]['add']:
                self.data_table[name_table] = pd.DataFrame()

        if self.task['print_parameters']['add']:
            self.set_output_parameters = set()
            for task_i in self.task['print_parameters']['sel'].replace(' ', '').split('/'):
                key_row, key_column = task_i.split(":")  # нр"8;9", "pn;qn"
                for col in key_column.split(';'):  # ['pn','qn']
                    for row in key_row.split(';'):  # ['15105,15113','15038,15037,4']
                        self.set_output_parameters.add(f'{row}:{col}')

        if self.task['print_balance_q']['add']:
            self.sheet_q = self.book.create_sheet("balance_Q")
            self.row_q = {}
            # (имя ключа, название в ячейке XL, комментарий ячейки)
            name_row = (
                ('row_name',
                 'Наименование', ''),
                ('row_qn',
                 'Реактивная мощность нагрузки', 'Calc(sum,area,qn,vibor)'),
                ('row_dq_sum',
                 'Нагрузочные потери', ''),
                ('row_dq_line',
                 'в т.ч. потери в ЛЭП', 'потери в ЛЭП: \nCalc(sum,area,dq_line,vibor)'),
                ('row_dq_tran',
                 'потери в трансформаторах', 'Calc(sum,area,dq_tran,vibor)'),
                ('row_shq_tran',
                 'потери Х.Х. в трансформаторах', 'Calc(sum,area,shq_tran,vibor)'),
                ('row_skrm_potr',
                 'Потребление реактивной мощности СКРМ (ШР, УШР, СК, СТК)',
                 'Calc(sum,node,qsh,qsh>0 & vibor) - Calc(sum,node,qg,qg<0&pg<0.1&pg>-0.1 & vibor)'),
                ('row_sum_port_Q',
                 'Суммарное потребление реактивной мощности', ''),
                ('row_qg',
                 'Генерация реактивной мощности электростанциями', 'Calc(sum,node,qg,(pg>0.1|pg<-0.1) & vibor)'),
                ('row_skrm_gen',
                 'Генерация реактивной мощности СКРМ (БСК, СК, СТК)', ''),
                ('row_qg_min',
                 'Минимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmin,pg>0.1& vibor)'),
                ('row_qg_max',
                 'Максимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmax,pg>0.1& vibor)'),
                ('row_shq_line',
                 'Зарядная мощность ЛЭП', 'Calc(sum,area,shq_line, vibor)'),
                ('row_sum_QG',
                 'Суммарная генерация реактивной мощности', ''),
                ('row_Q_itog',
                 'Внешний переток реактивной мощности (избыток/дефицит +/-)', ''),
                ('row_Q_itog_gmin',
                 'Внешний переток реактивной мощности при минимальной генерации '
                 'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)', ''),
                ('row_Q_itog_gmax',
                 'Внешний переток реактивной мощности при максимальной генерации '
                 'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)',
                 ''),
            )
            self.sheet_q.cell(1, 1, 'Таблица 1 - Баланс реактивной мощности, Мвар')
            for n, row_info in enumerate(name_row, 2):
                self.row_q[row_info[0]] = n
                self.sheet_q.cell(n, 1, row_info[1])
                if row_info[2]:
                    self.sheet_q.cell(n, 1).comment = Comment(row_info[2], '')

    def add_val(self, rm):
        log_print_xl.info("\tВывод данных из моделей в XL")

        # Добавить значения в вывод таблиц.
        for name_table in self.data_table:
            # проверка наличия таблицы
            if rm.rastr.Tables.Find(name_table) < 0:
                if name_table == 'sechen':
                    rm.downloading_additional_files('sch')
            # Считать данные из таблиц растр.

            fields = self.task['set_printXL'][name_table]['par'].replace(' ', '')
            setsel = self.task['set_printXL'][name_table]['sel']
            if not fields:
                fields = rm.all_cols(name_table)

            data = rm.df_from_table(table_name=name_table, fields=fields, setsel=setsel)
            if not data.empty:
                self.data_table[name_table] = pd.concat([self.data_table[name_table],
                                                         data.apply(lambda x: pd.Series(rm.info_file), axis=1).join(
                                                             other=data)])

        if self.task['print_parameters']['add']:
            self.add_val_parameters(rm)

        if self.task['print_balance_q']['add']:
            self.add_val_balance_q(rm)

    def add_val_parameters(self, rm):
        """
        Вывод заданных параметров в формате: "15105,15113;15038,15037,4:r;x;b / 15198:pg;qg / ns=1(sechen):psech".
        """
        if 'sechen' in self.task['print_parameters']['sel']:
            if rm.rastr.tables.Find('sechen') < 0:
                rm.downloading_additional_files('sch')
        date = pd.Series(dtype='object')
        for i in self.set_output_parameters:
            k, p = i.split(':')
            table, sel = rm.recognize_key(k, 'tab sel')
            if rm.rastr.tables(table).cols(p).Prop(1) == 2:  # если поле типа строка
                date.loc[i] = str(rm.txt_field_return(table, sel, p))
            else:
                date.loc[i] = rm.rastr.tables(table).cols.Item(p).ZS(rm.index(table_name=table, key_str=sel))
        date = pd.concat([date, pd.Series(rm.info_file)])
        self.data_parameters = pd.concat([self.data_parameters, date], axis=1)

    def add_val_balance_q(self, rm):
        column = self.sheet_q.max_column + 1
        choice = self.task["print_balance_q"]["sel"]
        self.sheet_q.cell(2, column,
                          f'{rm.info_file["Сезон макс/мин"]} {rm.info_file["Год"]} г ({rm.info_file["Доп. имена"]})')
        area = rm.rastr.Tables("area")
        area.SetSel(self.task["print_balance_q"]["sel"])
        # ndx = area.FindNextSel(-1)

        # Реактивная мощность нагрузки
        address_qn = self.sheet_q.cell(self.row_q['row_qn'], column,
                                       rm.rastr.Calc("sum", "area", "qn", choice)).coordinate
        # Потери Q в ЛЭП
        address_dq_line = self.sheet_q.cell(self.row_q['row_dq_line'], column,
                                            rm.rastr.Calc("sum", "area", "dq_line", choice)).coordinate
        # Потери Q в трансформаторах
        address_dq_tran = self.sheet_q.cell(self.row_q['row_dq_tran'], column,
                                            rm.rastr.Calc("sum", "area", "dq_tran", choice)).coordinate
        # Потери Q_ХХ в трансформаторах
        address_shq_tran = self.sheet_q.cell(self.row_q['row_shq_tran'], column,
                                             rm.rastr.Calc("sum", "area", "shq_tran", choice)).coordinate
        # ШР УШР без бСК
        skrm = (rm.rastr.Calc("sum", "node", "qsh", f"qsh>0&({choice})") -
                rm.rastr.Calc("sum", "node", "qg", f"qg<0&pg<0.1&pg>-0.1&({choice})"))
        address_SHR = self.sheet_q.cell(self.row_q['row_skrm_potr'], column, skrm).coordinate
        # Генерация Q генераторов
        address_qg = self.sheet_q.cell(self.row_q['row_qg'], column,
                                       rm.rastr.Calc("sum", "node", "qg", f"(pg>0.1|pg<-0.1)&({choice})")).coordinate
        # Генерация БСК шунтом и СТК СК
        address_skrm_gen = self.sheet_q.cell(self.row_q['row_skrm_gen'], column,
                                             -rm.rastr.Calc("sum", "node", "qsh", f"qsh<0&({choice})") + rm.rastr.Calc(
                                                 "sum", "node", "qg", f"qg>0&pg<0.1&pg>-0.1&({choice})")).coordinate
        # Минимальная генерация реактивной мощности в узлах выборки
        address_qg_min = self.sheet_q.cell(self.row_q['row_qg_min'], column,
                                           rm.rastr.Calc("sum", "node", "qmin", f"pg>0.1&({choice})")).coordinate
        # Максимальная генерация реактивной мощности в узлах выборки
        address_qg_max = self.sheet_q.cell(self.row_q['row_qg_max'], column,
                                           rm.rastr.Calc("sum", "node", "qmax", f"pg>0.1&({choice})")).coordinate
        # Генерация Q в ЛЭП
        address_shq_line = self.sheet_q.cell(self.row_q['row_shq_line'], column,
                                             - rm.rastr.Calc("sum", "area", "shq_line", choice)).coordinate
        address_losses = self.sheet_q.cell(self.row_q['row_dq_sum'], column,
                                           f"={address_dq_line}+{address_dq_tran}+{address_shq_tran}").coordinate
        address_load = self.sheet_q.cell(self.row_q['row_sum_port_Q'], column,
                                         f"={address_qn}+{address_losses}+{address_SHR}").coordinate
        address_sum_gen = self.sheet_q.cell(self.row_q['row_sum_QG'], column,
                                            f"={address_qg}+{address_shq_line}+{address_skrm_gen}").coordinate
        self.sheet_q.cell(self.row_q['row_Q_itog'], column,
                          f"=-{address_load}+{address_sum_gen}")
        self.sheet_q.cell(self.row_q['row_Q_itog_gmin'], column,
                          f"=-{address_load}+{address_qg_min}+{address_shq_line}")
        self.sheet_q.cell(self.row_q['row_Q_itog_gmax'], column,
                          f"=-{address_load}+{address_qg_max}+{address_shq_line}")

    def finish(self):
        """
        Преобразовать в объект таблицу и удалить листы с одной строкой.
        """

        self.name_xl_file = self.task['name_time'] + ' вывод данных.xlsx'
        self.book.save(self.name_xl_file)
        self.book = None

        for name_table, data in self.data_table.items():
            limitation = ''
            value_p = ''
            for lim in ['pmax', 'set_pop']:
                if lim in data.columns:
                    if len(data[data[lim] != 0]):
                        limitation = lim
                        break
            for val in ['psech', 'pop', 'pp']:
                if val in data.columns:
                    value_p = val
                    break
            if limitation and val:
                data.loc[data[limitation] != 0, 'difference_p'] = data.loc[data[limitation] != 0, value_p] - \
                                                                  data.loc[data[limitation] != 0, limitation]

            with pd.ExcelWriter(path=self.name_xl_file, mode='a', engine="openpyxl") as writer:
                data.to_excel(excel_writer=writer,
                              sheet_name=name_table,
                              header=True,
                              index=False)

        if self.task['print_parameters']['add']:
            with pd.ExcelWriter(path=self.name_xl_file, mode='a', engine="openpyxl") as writer:
                self.data_parameters.T.to_excel(excel_writer=writer,
                                                sheet_name='Значения',
                                                header=True,
                                                index=False)

        self.book = load_workbook(self.name_xl_file)
        for sheet_name in self.book.sheetnames:
            sheet = self.book[sheet_name]
            if sheet.max_row < 2:
                del self.book[sheet_name]  # удалить пустой лист
            else:
                if sheet_name != "balance_Q":
                    PrintXL.create_table(sheet, sheet_name)  # Создать объект таблица.

        if self.task['print_balance_q']['add']:
            self.sheet_q = self.book['balance_Q']
            self.sheet_q.row_dimensions[2].height = 140
            self.sheet_q.column_dimensions['A'].width = 40
            thins = Side(border_style="thin", color="000000")
            for row in range(2, self.sheet_q.max_row + 1):
                for col in range(1, self.sheet_q.max_column + 1):
                    if row > 2 and col > 1:
                        self.sheet_q.cell(row, col).number_format = BUILTIN_FORMATS[1]
                    self.sheet_q.cell(row, col).border = Border(thins, thins, thins, thins)
                    self.sheet_q.cell(row, col).font = Font(name='Times New Roman', size=11)
                    if row == 2:
                        self.sheet_q.cell(row, col).alignment = Alignment(text_rotation=90,
                                                                          wrap_text=True, horizontal="center")
                    if col == 1:
                        self.sheet_q.cell(row, col).alignment = Alignment(wrap_text=True)
                    if row in [12, 13, 17, 18]:
                        self.sheet_q.cell(row, col).fill = PatternFill('solid', fgColor="00FF0000")
                    if row in [9, 15, 16]:
                        self.sheet_q.cell(row, col).font = Font(bold=True)

        self.book.save(self.name_xl_file)
        self.book = None

        # Открыть excel через win32com.client и создать сводные.
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.ScreenUpdating = False  # обновление экрана
        # excel.Calculation = -4135  # xlCalculationManual
        excel.EnableEvents = False  # отслеживание событий
        excel.StatusBar = False  # отображение информации в строке статуса excel
        try:
            self.book = excel.Workbooks.Open(self.name_xl_file)
        except Exception:
            raise Exception(f'Ошибка при открытии файла {self.name_xl_file=}')

        for name_sheet in self.data_table:
            rows = self.task['set_printXL'][name_sheet]['rows'].split(",")
            rows = list(set(rows) & set(self.data_table[name_sheet].columns))
            columns = self.task['set_printXL'][name_sheet]['columns'].split(",")
            columns = list(set(columns) & set(self.data_table[name_sheet].columns))
            values = self.task['set_printXL'][name_sheet]['values'].split(",")
            values = list(set(values) & set(self.data_table[name_sheet].columns))

            tab_log = self.book.sheets[name_sheet].ListObjects[0]
            name_pivot_sheet = name_sheet + '_сводная'
            pivot_sheet = self.book.Sheets.Add(After=name_sheet)
            pivot_sheet.Name = name_pivot_sheet

            pt_cache = self.book.PivotCaches().add(1, tab_log)  # создать КЭШ xlDatabase, ListObjects
            pt = pt_cache.CreatePivotTable(TableDestination=name_pivot_sheet + "!R1C1",
                                           TableName="Сводная_" + name_sheet)  # создать сводную таблицу
            pt.ManualUpdate = True  # не обновить сводную
            pt.AddFields(RowFields=rows,
                         ColumnFields=columns,
                         PageFields=["Имя файла"],
                         AddToTable=False)

            for val in values:
                pt.AddDataField(pt.PivotFields(val),
                                val + " ",
                                -4157)  # xlMax -4136 xlSum -4157
                # Field                      Caption             def формула расчета
                pt.PivotFields(val + " ").NumberFormat = "0"

            # .PivotFields("na").ShowDetail = True #  группировка
            pt.RowAxisLayout(1)  # xlTabularRow показывать в табличной форме!!!!
            if len(values) > 1:
                pt.DataPivotField.Orientation = 1  # xlRowField"Значения в столбцах или строках xlColumnField

            # pt.DataPivotField.Position = 1 # позиция в строках
            pt.RowGrand = False  # удалить строку общих итогов
            pt.ColumnGrand = False  # удалить столбец общих итогов
            pt.MergeLabels = True  # объединять одинаковые ячейки
            pt.HasAutoFormat = False  # не обновлять ширину при обновлении
            pt.NullString = "--"  # заменять пустые ячейки
            pt.PreserveFormatting = False  # сохранять формат ячеек при обновлении
            pt.ShowDrillIndicators = False  # показывать кнопки свертывания
            # pt.PivotCache.MissingItemsLimit = 0 # xlMissingItemsNone
            # xlMissingItemsNone для норм отображения уникальных значений автофильтра
            for row in rows:
                pt.PivotFields(row).Subtotals = [False, False, False, False, False, False, False, False,
                                                 False, False,
                                                 False, False]  # промежуточные итоги и фильтры
            for column in columns:
                pt.PivotFields(column).Subtotals = [False, False, False, False, False, False, False, False,
                                                    False, False,
                                                    False, False]  # промежуточные итоги и фильтры
            pt.ManualUpdate = False  # обновить сводную
            pt.TableStyle2 = ""  # стиль
            if name_sheet in ["area", "area2", "darea"]:
                pt.ColumnRange.ColumnWidth = 10  # ширина строк
                pt.RowRange.ColumnWidth = 9
                pt.RowRange.Columns(1).ColumnWidth = 7
                pt.RowRange.Columns(2).ColumnWidth = 20
                pt.RowRange.Columns(3).ColumnWidth = 20
                pt.RowRange.Columns(6).ColumnWidth = 20
            pt.DataBodyRange.HorizontalAlignment = -4108  # xlCenter
            # .DataBodyRange.NumberFormat = "#,##0"
            # формат
            pt.TableRange1.WrapText = True  # перенос текста в ячейке
            pt.TableRange1.Borders(7).LineStyle = 1  # лево
            pt.TableRange1.Borders(8).LineStyle = 1  # верх
            pt.TableRange1.Borders(9).LineStyle = 1  # низ
            pt.TableRange1.Borders(10).LineStyle = 1  # право
            pt.TableRange1.Borders(11).LineStyle = 1  # внутри вертикаль
            pt.TableRange1.Borders(12).LineStyle = 1  #

        self.book.Save()
        excel.Visible = True
        excel.ScreenUpdating = True  # обновление экрана
        excel.Calculation = -4105  # xlCalculationAutomatic
        excel.EnableEvents = True  # отслеживание событий
        excel.StatusBar = True  # отображение информации в строке статуса excel

    @staticmethod
    def create_table(sheet, sheet_name, point_start: str = 'A1'):
        """
        Создать объект таблица из всего диапазона листа.
        :param sheet: Объект лист excel
        :param sheet_name: Имя таблицы.
        :param point_start:
        """
        tab = Table(displayName=sheet_name,
                    ref=f'{point_start}:' + get_column_letter(sheet.max_column) + str(sheet.max_row))

        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                            showFirstColumn=False,
                                            showLastColumn=False,
                                            showRowStripes=True,
                                            showColumnStripes=True)
        sheet.add_table(tab)
