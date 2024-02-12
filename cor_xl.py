"""Модуль для изменения параметров РМ по заданию в таблице excel."""
import logging
import re
import os
from openpyxl import load_workbook

from rastr_model import *
from import_rm import *

log_cor_xl = logging.getLogger(f'__main__.{__name__}')


class CorXL:
    """
    Изменить параметры модели по заданию в таблице excel.
    """

    def __init__(self, excel_file_name: str, sheets: str) -> None:
        """
        Проверить наличие книги и листов, создать классы CorSheet для листов.
        :param excel_file_name: Полное имя файла excel;
        :param sheets: имя листов, нр [импорт из моделей][XL->RastrWin], если '*', то все листы по порядку
        """
        self.sheets_list = []  # для хранения объектов CorSheet
        log_cor_xl.info(f"Изменить модели по заданию из книги: {excel_file_name}, листы: {sheets}")
        if not os.path.exists(excel_file_name):
            raise ValueError("Ошибка в задании, не найден файл: " + excel_file_name)
        else:
            self.excel_file_name = excel_file_name
            # data_only - Загружать расчетные значения ячеек, иначе будут формулы.
            self.wb = load_workbook(excel_file_name, data_only=True)

            if sheets == '*':  # все листы
                self.sheets = self.wb.sheetnames
                for sheet in self.sheets:
                    if '#' not in sheet:  # все листы
                        self.sheets_list.append(CorSheet(name=sheet, obj=self.wb[sheet]))
            else:
                self.sheets = re.findall(r"\[(.+?)]", sheets)
                for sheet in self.sheets:
                    if sheet not in self.wb.sheetnames:
                        raise ValueError(f"Ошибка в задании, не найден лист: {sheet} в файле {excel_file_name}")
                    else:
                        self.sheets_list.append(CorSheet(name=sheet, obj=self.wb[sheet]))

    def init_export_model(self) -> None:
        """Экспорт данных из растр в csv"""
        for sheet in self.sheets_list:
            if sheet.type == 'import_model':
                sheet.export_model()

    def run_xl(self, rm: RastrModel) -> None:
        """Запуск всех корректировок"""
        for sheet in self.sheets_list:
            sheet.run_sheets(rm)


class CorSheet:
    """
    Клас лист для хранения листов книги excel и работы с ними.
    """
    SHAPE = {"Параметры импорта из файлов RastrWin": 'import_model',  # Импорт из моделей(ИМ)
             'Выполнить изменение модели по строкам': 'list_cor',  # Строковая форма(СФ)
             'Имя таблицы:': 'table_import'}  # Импорт таблиц(ИТ)

    def __init__(self, name: str, obj):
        """
        :param name: Имя листа
        :param obj: Объект лист
        """
        #  Сводная таблица корректировок 'tab_cor', нр корр потребления или нагрузка узлов / имя файла.
        #  Таблица корректировок по списку 'list_cor', нр изм, удалить, снять отметку.
        #  Импорт моделей import_model.
        #  'Импорт таблиц(ИТ) table_import'. 
        self.name = name
        self.xls = obj  # xls.cell(1,1).value [строки][столбцы] xls.max_row xls.max_column
        self.import_model_all = []  # Для хранения объектов ImportFromModel
        self.calc_val = self.xls.cell(1, 1).value
        if isinstance(self.calc_val, int):
            self.type = 'tab_cor'
        else:
            try:
                self.type = self.SHAPE[self.calc_val]
            except KeyError:
                raise ValueError(f'Тип задания листа {name!r} не распознан.')

    def run_sheets(self, rm: RastrModel):
        log_cor_xl.info(f'\tВыполнение задания листа {self.name!r}')
        if self.type == 'import_model':
            self.import_model(rm)
        elif self.type == 'list_cor':
            self.list_cor(rm)
        elif self.type == 'tab_cor':
            self.tab_cor(rm)
        elif self.type == 'table_import':
            self.tab_import(rm)
        log_cor_xl.info(f'\tКонец выполнения задания листа {self.name!r}\n')

    def export_model(self):
        """"Экспорт из моделей"""
        for row in range(3, self.xls.max_row + 1):
            if self.xls.cell(row, 1).value and '#' not in self.xls.cell(row, 1).value:
                """ ИД для импорта из модели(выполняется после блока начала)"""
                ifm = ImportFromModel(RastrModel(full_name=self.xls.cell(row, 1).value, not_calculated=True),
                                      criterion_start={"years": self.xls.cell(row, 6).value,
                                                       "season": self.xls.cell(row, 7).value,
                                                       "max_min": self.xls.cell(row, 8).value,
                                                       "add_name": self.xls.cell(row, 9).value},
                                      tables=self.xls.cell(row, 2).value,
                                      param=self.xls.cell(row, 4).value,
                                      sel=self.xls.cell(row, 3).value,
                                      calc=self.xls.cell(row, 5).value)
                self.import_model_all.append(ifm)

    def tab_import(self, rm: RastrModel) -> None:
        """Импорт таблицы из XL в Rastr"""
        tables_name = self.xls.cell(2, 1).value
        type_import = self.xls.cell(4, 1).value
        field_import = []
        field_column = []
        for column in range(2, self.xls.max_column + 1):
            if self.xls.cell(1, column).value:
                field_column.append(column)
                field_import.append(self.xls.cell(1, column).value)
        field_import = ','.join(field_import)
        data = [[self.xls.cell(row, col).value for col in field_column] for row in range(2, self.xls.max_row + 1)]
        table = rm.rastr.Tables(tables_name)
        table.ReadSafeArray(type_import, field_import, data)

    def import_model(self, rm: RastrModel) -> None:
        """Импорт в модели"""
        if self.import_model_all:
            for im in self.import_model_all:
                im.import_data_in_rm(rm)

    def list_cor(self, rm: RastrModel) -> None:
        """
        Таблица корректировок по списку, нр изм, удалить, снять отметку.
        """
        # номера столбцов
        C_SELECTION = 2
        C_VALUE = 3

        for row in range(3, self.xls.max_row + 1):
            name_fun = self.xls.cell(row, 1).value
            if name_fun:
                if '#' not in name_fun:
                    sel = str(self.xls.cell(row, C_SELECTION).value)
                    value = self.xls.cell(row, C_VALUE).value
                    year = self.xls.cell(row, 4).value
                    season = self.xls.cell(row, 5).value
                    max_min = self.xls.cell(row, 6).value
                    add_name = self.xls.cell(row, 7).value
                    statement = self.xls.cell(row, 8).value

                    if any([year, season, max_min, add_name]) and rm.code_name_rg2:  # any если хотя бы один истина
                        if not rm.test_name(condition={"years": year, "season": season,
                                                       "max_min": max_min, "add_name": add_name},
                                            info=f'\t\tcor_x:{sel=}, {value=}'):
                            continue
                    if statement:
                        if not rm.conditions_test(statement):
                            continue
                    log_cor_xl.info(rm.txt_task_cor(name=name_fun, sel=sel, value=value))

    def tab_cor(self, rm: RastrModel) -> None:
        """
        Корректировка моделей по заданию в табличном виде
        """
        name_files = ""
        dict_param_column = {}  # {10: "pn"} - столбец: параметр
        # Шаг по колонкам и запись в словарь всех столбцов для коррекции
        for column_name_file in range(2, self.xls.max_column + 1):
            if self.xls.cell(1, column_name_file).value not in ["", None]:
                name_files = self.xls.cell(1, column_name_file).value.split("|")
            if self.xls.cell(2, column_name_file).value:
                duct_add = False
                for name_file in name_files:
                    if name_file in [rm.name_base, "*"]:
                        duct_add = True
                    if "*" in name_file and len(name_file) > 7:
                        match = re.search(re.compile(r"\[(.*)].*\[(.*)].*\[(.*)].*\[(.*)]"),
                                          name_file)
                        if match.re.groups == 4 and rm.code_name_rg2:
                            if rm.test_name(condition={"years": match[1], "season": match[2],
                                                       "max_min": match[3], "add_name": match[4]},
                                            info=f"\tcor_xl, условие: {name_file}, ") or not rm.code_name_rg2:
                                duct_add = True
                if duct_add:
                    _ = self.xls.cell(2, column_name_file).value
                    dict_param_column[column_name_file] = _.replace(' ', '')

        if not dict_param_column:
            log_cor_xl.info(f"\t {rm.name_base} НЕ НАЙДЕН на листе {self.name} книги excel")
        else:
            log_cor_xl.info(f'\t\tРасчетной модели соответствуют столбцы: параметры {dict_param_column}')
            calc_vals = {1: "ЗАМЕНИТЬ", 2: "+", 3: "-", 0: "*"}
            # 1: "ЗАМЕНИТЬ", 2: "ПРИБАВИТЬ", 3: "ВЫЧЕСТЬ", 0: "УМНОЖИТЬ"
            for row in range(3, self.xls.max_row + 1):
                for column, param in dict_param_column.items():
                    short_keys = self.xls.cell(row, 1).value
                    if short_keys not in [None, ""]:
                        short_keys = str(short_keys)
                        new_val = self.xls.cell(row, column).value
                        if new_val is not None:
                            # for short_key in short_keys.split(';'):
                            if param in ["pop", "pp"]:
                                rm.cor_pop(zone=short_keys,
                                           new_pop=new_val)  # изменить потребление
                            else:
                                if self.calc_val == 1:
                                    rm.cor(keys=short_keys,
                                           values=f"{param}={new_val}",
                                           print_log=True)
                                else:
                                    rm.cor(keys=short_keys,
                                           values=f"{param}={param}{calc_vals[self.calc_val]}{new_val}",
                                           print_log=True)
