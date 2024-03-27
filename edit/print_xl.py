__all__ = ['PrintXL', 'BalanceQ']

import os

from openpyxl.reader.excel import load_workbook

from pivot_table import make_pivot_tables

"""Модуль для вывода параметров РМ в таблице excel."""
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS

from collection_func import from_list1_only_exists_in_list2

log_print_xl = logging.getLogger(f'__main__.{__name__}')


class PrintXL:
    """Класс печать данных в excel"""
    short_name_tables = {'n': 'node',
                         'v': 'vetv',
                         'g': 'Generator',
                         'na': 'area',
                         'npa': 'area2',
                         'no': 'darea',
                         'nga': 'ngroup',
                         'ns': 'sechen'}
    set_param = {
        'sechen': {'sel': 'ns>0',
                   'par': '',  # 'ns,name,pmin,pmax,psech',
                   'rows': 'ns,name',  # поля строк в сводной
                   'columns': 'Год,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3',  # поля столбцов в сводной
                   'values': 'psech,pmax,difference_p'},
        'area': {'sel': 'na>0',
                 'par': '',  # 'na,name,no,pg,pn,pn_sum,dp,pop,set_pop,qn_sum,pg_max,pg_min,poq,qn,qg,dev_pop'
                 'rows': 'na,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3',  # поля строк в сводной
                 'columns': 'Год',  # поля столбцов в сводной
                 'values': 'pop,difference_p'},
        'area2': {'sel': 'npa>0',
                  'par': '',  # 'npa,name,pg,pn,dp,pop,vnp,qg,qn,dq,poq,vnq,pn_sum,qn_sum,set_pop,dev_pop'
                  'rows': 'npa,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3',  # поля строк в сводной
                  'columns': 'Год',  # поля столбцов в сводной
                  'values': 'pop,difference_p'},
        'darea': {'sel': 'no>0',
                  'par': '',  # 'no,name,pg,pp,pvn,qn_sum,pnr_sum,pn_sum,set_pop,qvn,qp,qg,dev_pop',
                  'rows': 'no,name,Сезон макс/мин,Доп. имя1,Доп. имя2,Доп. имя3',  # поля строк в сводной
                  'columns': 'Год',  # поля столбцов в сводной
                  'values': 'pp,difference_p'}
    }

    def __init__(self, task):
        """
        Добавить листы и первая строка с названиями
        """
        self.data_table = {}  # {'имя листа=имя таблицы': fd c данными}
        self.task = task
        for tb in self.task['set_printXL']:
            if tb in self.set_param:
                self.task['set_printXL'][tb] = self.task['set_printXL'][tb] | self.set_param[tb]

        #  Создать лист xl и присвоить ссылку на него
        for name_table in self.task['set_printXL']:
            if self.task['set_printXL'][name_table]['add']:
                self.data_table[name_table] = pd.DataFrame()

    def add_val(self, rm):
        """
        Добавить значения.
        :param rm:
        """
        log_print_xl.info('\tВывод данных из моделей в XL')

        for name_table in self.data_table:
            # Проверка наличия таблицы
            if rm.rastr.Tables.Find(name_table) < 0:
                if name_table == 'sechen':
                    rm.downloading_additional_files('sch')
                else:
                    raise ValueError(f'Отсутствует таблица {name_table}')

            # Считать данные из таблиц растр.

            fields = self.task['set_printXL'][name_table]['par'].replace(' ', '')
            setsel = self.task['set_printXL'][name_table]['sel']

            if not fields:
                fields = rm.all_cols(name_table)

            data = rm.df_from_table(table_name=name_table,
                                    fields=fields,
                                    setsel=setsel)
            if data is not None:
                data_extended = data.apply(lambda x: pd.Series(rm.info_file), axis=1).join(other=data)
                self.data_table[name_table] = pd.concat([self.data_table[name_table], data_extended])

    def save_to_xl(self,
                   path_xl_book: str):
        """
        Сохранить данные в excel и сформировать сводные.
        :param path_xl_book: Путь для сохранения excel.
        """
        if not self.data_table:
            log_print_xl.debug('Данные для сохранения в excel отсутствуют.')
            return
        dict_columns = {}
        for name_table, data in self.data_table.items():

            # Добавить столбец difference_p.
            limitation = ''
            value_p = ''
            for lim in ['pmax', 'set_pop']:
                if lim in data.columns:
                    if len(data[data[lim] != 0]):
                        limitation = lim
                        break
            for val in ['psech', 'pop', 'pp']:
                if val in data.columns:
                    value_p = val
                    break
            if limitation and val:
                data.loc[data[limitation] != 0, 'difference_p'] = data.loc[data[limitation] != 0, value_p] - \
                                                                  data.loc[data[limitation] != 0, limitation]
            # Записать в excel
            mode = 'a' if os.path.exists(path_xl_book) else 'w'
            with pd.ExcelWriter(path=path_xl_book, mode=mode) as writer:
                data.to_excel(excel_writer=writer,
                              sheet_name=name_table,
                              float_format='%.2f',
                              freeze_panes=(1, 1),
                              index=False)
            dict_columns[name_table] = tuple(data.columns)

        task_pivot = self.create_task_for_pivot_tables(dict_columns)
        make_pivot_tables(book_path=path_xl_book,
                          sheets_info=task_pivot)

    def create_task_for_pivot_tables(self,
                                     dict_columns: dict) -> dict:
        """
        Формирование задания для сводной.
        :param dict_columns: Словарь с перечнем имен столбцов.
        :return: См.sheets_info в make_pivot_tables
        """
        sheets_info = {}

        for key in dict_columns:
            sheets_info[key] = {}
            columns = dict_columns[key]

            sheet_info = sheets_info[key]
            sheet_info['sheet_name'] = f'Сводная_{key}'
            sheet_info['pt_name'] = f'pt_{key}'

            sheet_info['row_fields'] = self.task['set_printXL'][key]['rows'].split(',')
            sheet_info['row_fields'] = list(set(sheet_info['row_fields']) & set(self.data_table[key].columns))

            sheet_info['row_fields'] = from_list1_only_exists_in_list2(sheet_info['row_fields'],
                                                                       columns)

            sheet_info['column_fields'] = self.task['set_printXL'][key]['columns'].split(',')
            sheet_info['column_fields'] = list(set(sheet_info['column_fields']) & set(self.data_table[key].columns))

            # sheet_info['column_fields'] = ['Год', 'Сезон макс/мин'] + [col for col in columns if 'Доп. имя' in col]
            sheet_info['column_fields'] = from_list1_only_exists_in_list2(sheet_info['column_fields'],
                                                                          columns)

            sheet_info['page_fields'] = ['Имя файла']

            sheet_info['page_fields'] = from_list1_only_exists_in_list2(sheet_info['page_fields'],
                                                                        columns)

            data_list = self.task['set_printXL'][key]['values'].split(',')
            data_list = list(set(data_list) & set(self.data_table[key].columns))
            sheet_info['data_field'] = {}
            for i in data_list:
                sheet_info['data_field'][i] = i + ' '
            sheet_info['conditional_formatting'] = False
        return sheets_info


class BalanceQ:

    def __init__(self, sel: str):
        """
        Вывод в excel таблиц c балансом реактивной мощности.
        :param sel: Выборка в таблице node.
        """
        self.sel = sel
        self.book = Workbook()
        self.sheet_q = self.book.create_sheet('balance_Q')
        self.row_q = {}
        # (имя ключа, название в ячейке XL, комментарий ячейки)
        name_row = (
            ('row_name',
             'Наименование', ''),
            ('row_qn',
             'Реактивная мощность нагрузки', 'Calc(sum,area,qn,vibor)'),
            ('row_dq_sum',
             'Нагрузочные потери', ''),
            ('row_dq_line',
             'в т.ч. потери в ЛЭП', 'потери в ЛЭП: \nCalc(sum,area,dq_line,vibor)'),
            ('row_dq_tran',
             'потери в трансформаторах', 'Calc(sum,area,dq_tran,vibor)'),
            ('row_shq_tran',
             'потери Х.Х. в трансформаторах', 'Calc(sum,area,shq_tran,vibor)'),
            ('row_skrm_potr',
             'Потребление реактивной мощности СКРМ (ШР, УШР, СК, СТК)',
             'Calc(sum,node,qsh,qsh>0 & vibor) - Calc(sum,node,qg,qg<0&pg<0.1&pg>-0.1 & vibor)'),
            ('row_sum_port_Q',
             'Суммарное потребление реактивной мощности', ''),
            ('row_qg',
             'Генерация реактивной мощности электростанциями', 'Calc(sum,node,qg,(pg>0.1|pg<-0.1) & vibor)'),
            ('row_skrm_gen',
             'Генерация реактивной мощности СКРМ (БСК, СК, СТК)', ''),
            ('row_qg_min',
             'Минимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmin,pg>0.1& vibor)'),
            ('row_qg_max',
             'Максимальная генерация реактивной мощности электростанциями', 'Calc(sum,node,qmax,pg>0.1& vibor)'),
            ('row_shq_line',
             'Зарядная мощность ЛЭП', 'Calc(sum,area,shq_line, vibor)'),
            ('row_sum_QG',
             'Суммарная генерация реактивной мощности', ''),
            ('row_Q_itog',
             'Внешний переток реактивной мощности (избыток/дефицит +/-)', ''),
            ('row_Q_itog_gmin',
             'Внешний переток реактивной мощности при минимальной генерации '
             'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)', ''),
            ('row_Q_itog_gmax',
             'Внешний переток реактивной мощности при максимальной генерации '
             'реактивной мощности электростанциями и КУ(избыток/дефицит +/-)',
             ''),
        )
        self.sheet_q.cell(1, 1, 'Таблица 1 - Баланс реактивной мощности, Мвар')
        for n, row_info in enumerate(name_row, 2):
            self.row_q[row_info[0]] = n
            self.sheet_q.cell(n, 1, row_info[1])
            if row_info[2]:
                self.sheet_q.cell(n, 1).comment = Comment(row_info[2], '')

    def add_val(self, rm):

        column = self.sheet_q.max_column + 1
        self.sheet_q.cell(2, column,
                          f'{rm.info_file["Сезон макс/мин"]} {rm.info_file["Год"]} г ({rm.info_file["Доп. имена"]})')
        area = rm.rastr.Tables('area')
        area.SetSel(self.sel)
        # ndx = area.FindNextSel(-1)

        # Реактивная мощность нагрузки
        address_qn = self.sheet_q.cell(self.row_q['row_qn'], column,
                                       rm.rastr.Calc('sum', 'area', 'qn', self.sel)).coordinate
        # Потери Q в ЛЭП
        address_dq_line = self.sheet_q.cell(self.row_q['row_dq_line'], column,
                                            rm.rastr.Calc('sum', 'area', 'dq_line', self.sel)).coordinate
        # Потери Q в трансформаторах
        address_dq_tran = self.sheet_q.cell(self.row_q['row_dq_tran'], column,
                                            rm.rastr.Calc('sum', 'area', 'dq_tran', self.sel)).coordinate
        # Потери Q_ХХ в трансформаторах
        address_shq_tran = self.sheet_q.cell(self.row_q['row_shq_tran'], column,
                                             rm.rastr.Calc('sum', 'area', 'shq_tran', self.sel)).coordinate
        # ШР УШР без бСК
        skrm = (rm.rastr.Calc('sum', 'node', 'qsh', f'qsh>0&({self.sel})') -
                rm.rastr.Calc('sum', 'node', 'qg', f'qg<0&pg<0.1&pg>-0.1&({self.sel})'))
        address_SHR = self.sheet_q.cell(self.row_q['row_skrm_potr'], column, skrm).coordinate
        # Генерация Q генераторов
        address_qg = self.sheet_q.cell(self.row_q['row_qg'], column,
                                       rm.rastr.Calc('sum', 'node', 'qg', f'(pg>0.1|pg<-0.1)&({self.sel})')).coordinate
        # Генерация БСК шунтом и СТК СК
        address_skrm_gen = self.sheet_q.cell(self.row_q['row_skrm_gen'], column,
                                             -rm.rastr.Calc('sum', 'node', 'qsh',
                                                            f'qsh<0&({self.sel})') + rm.rastr.Calc(
                                                 'sum', 'node', 'qg', f'qg>0&pg<0.1&pg>-0.1&({self.sel})')).coordinate
        # Минимальная генерация реактивной мощности в узлах выборки
        address_qg_min = self.sheet_q.cell(self.row_q['row_qg_min'], column,
                                           rm.rastr.Calc('sum', 'node', 'qmin', f'pg>0.1&({self.sel})')).coordinate
        # Максимальная генерация реактивной мощности в узлах выборки
        address_qg_max = self.sheet_q.cell(self.row_q['row_qg_max'], column,
                                           rm.rastr.Calc('sum', 'node', 'qmax', f'pg>0.1&({self.sel})')).coordinate
        # Генерация Q в ЛЭП
        address_shq_line = self.sheet_q.cell(self.row_q['row_shq_line'], column,
                                             - rm.rastr.Calc('sum', 'area', 'shq_line', self.sel)).coordinate
        address_losses = self.sheet_q.cell(self.row_q['row_dq_sum'], column,
                                           f'={address_dq_line}+{address_dq_tran}+{address_shq_tran}').coordinate
        address_load = self.sheet_q.cell(self.row_q['row_sum_port_Q'], column,
                                         f'={address_qn}+{address_losses}+{address_SHR}').coordinate
        address_sum_gen = self.sheet_q.cell(self.row_q['row_sum_QG'], column,
                                            f'={address_qg}+{address_shq_line}+{address_skrm_gen}').coordinate
        self.sheet_q.cell(self.row_q['row_Q_itog'], column,
                          f'=-{address_load}+{address_sum_gen}')
        self.sheet_q.cell(self.row_q['row_Q_itog_gmin'], column,
                          f'=-{address_load}+{address_qg_min}+{address_shq_line}')
        self.sheet_q.cell(self.row_q['row_Q_itog_gmax'], column,
                          f'=-{address_load}+{address_qg_max}+{address_shq_line}')

    def save_to_xl(self,
                   path_xl_book: str):
        self.book.save(path_xl_book)
        # self.book = load_workbook(path_xl_book)
        # sheet_q = self.book['balance_Q']
        #
        # sheet_q.row_dimensions['2'].height = 140
        # sheet_q.column_dimensions['A'].width = 40
        # thins = Side(border_style='thin', color='000000')
        # for row in range(2, sheet_q.max_row + 1):
        #     for col in range(1, sheet_q.max_column + 1):
        #         if row > 2 and col > 1:
        #             sheet_q.cell(row, col).number_format = BUILTIN_FORMATS[1]
        #         sheet_q.cell(row, col).border = Border(thins, thins, thins, thins)
        #         sheet_q.cell(row, col).font = Font(name='Times New Roman', size=11)
        #         if row == 2:
        #             sheet_q.cell(row, col).alignment = Alignment(text_rotation=90,
        #                                                          wrap_text=True, horizontal='center')
        #         if col == 1:
        #             sheet_q.cell(row, col).alignment = Alignment(wrap_text=True)
        #         if row in [12, 13, 17, 18]:
        #             sheet_q.cell(row, col).fill = PatternFill('solid', fgColor='00FF0000')
        #         if row in [9, 15, 16]:
        #             sheet_q.cell(row, col).font = Font(bold=True)
        # self.book.save(path_xl_book)

