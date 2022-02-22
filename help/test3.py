import os
from openpyxl import Workbook, load_workbook

book = Workbook()
s= book.create_sheet('лист')


d={0:0}
d[1]=s
print(d, type(d))
print(d.keys())
print(max(d.keys()))
os.path.dirname(r"I:\rastr_add\rastr example\pattern\режим.rg2")
r=1
