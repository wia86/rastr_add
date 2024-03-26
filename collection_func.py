import re
import sqlite3

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

KEY_TABLES = {'ny': 'node',
              'ip': 'vetv',
              'Num': 'Generator',
              # 'g': 'Generator',
              'na': 'area',
              'npa': 'area2',
              'no': 'darea',
              'nga': 'ngroup',
              'ns': 'sechen'}


def convert_s_key(s_key):
    """ Преобразовать '1,2' в (1, 2, 0).
     Число оставить числом"""
    match s_key:
        case str():
            return s_key_vetv_in_tuple(s_key)
        case int():
            return s_key
        case _:
            try:
                return int(s_key)
            except TypeError:
                raise TypeError(f'Неверный тип данных {type(s_key)!r}')


def s_key_vetv_in_tuple(s_key: str) -> tuple:
    """ Преобразовать '1,2' в (1, 2, 0) """
    if not isinstance(s_key, str):
        raise TypeError('Неверный тип данных')
    key = s_key.split(',')
    if len(key) == 2:
        key.append('0')
    key = tuple(map(int, key))
    if len(key) != 3:
        raise ValueError(f'//Ошибка входных параметров {s_key}')
    return key


def str_yeas_in_list(str_init: str, sep: tuple = ('...', '…')) -> list:
    """
    Преобразует перечень годов с диапазонами в отсортированный массив.
    :param str_init: '2021,2023...2025'
    :param sep: Картеж разделителей.
    :return: [2021,2023,2024,2025] или []
    """
    list_init = str_init.replace(' ', '').split(',')
    if not list_init:
        return []
    years_list_new = []
    for list_init_i in list_init:
        for sep_i in sep:
            if sep_i in list_init_i:
                min_val, max_val = list_init_i.split(sep_i)
                min_val = int(min_val)
                max_val = int(max_val)
                if min_val > max_val:
                    raise ValueError(f'Неверный формат: {str_init}')
                min_max = list(range(min_val, max_val + 1))
                years_list_new.extend(min_max)
                break
        else:
            years_list_new.append(int(list_init_i))
    return sorted(years_list_new)


def split_task_action(txt: str) -> list | bool:
    """
    Разделить строку по запятым, если запятая не внутри [] {}
    :param txt: '[1,2,0:sta=1],[2,3:sta=0]{5,7:sta==1},[9,8:sta=1],6'
    :return: ['[1,2,0:sta=1]', '[2,3:sta=0]{5,7:sta==1}', '[9,8:sta=1]', '6'] или  False
    """
    if not txt:
        return False
    # Вычленить значения в [] и {}.
    actions = re.findall(re.compile(r'\[(.+?)]'), txt)
    conditions = re.findall(re.compile(r'\{(.+?)}'), txt)

    # Заменить значения в [ ] и { } на act_cond_{n}
    dict_key = {}  # замена, действие
    for n, action in enumerate(actions + conditions):
        dict_key[f'act_cond_{n}'] = action
        txt = txt.replace(action, f'act_cond_{n}')

    #  Заменить act_cond_{n} на значения в [ ] и { }.
    result = []
    for part in txt.split(','):
        for key in dict_key:
            if key in part:
                part = part.replace(key, dict_key[key])
        result.append(part)
    return result


def name_table_from_key(task_key: str):
    """
    По ключу строки (нр ny=1) определяет имя таблицы.
    :return: table:str 'node' or False если не найдено.
    """
    for key_tables in KEY_TABLES:
        if key_tables in task_key:
            return KEY_TABLES[key_tables]
    return False


def recognize_key(common_key: str, back: str = 'all'):
    """
    Распознать:
     -имя таблицы;
     -короткий ключ числовой (генератор, узел) или строковый (ветвь '12,13') (s_key);
     -короткий ключ числовой (s_key_digit);
     -выборку в таблице (sel).
    :param common_key: например:['na=11(node)','125', 'g=125', '12,13,0', '12,13', 'ip=5&iq=3&np=0']
    :param back: тип возвращаемого значения
    :return:'all' (имя таблицы: str, выборка: str, ключ: int|tuple(int,int,int))
            'tab' имя таблицы: str
            's_key_digit'  ключ: int|tuple(int,int,int)
            'tab s_key_digit'
            'sel' выборка: str
            'tab sel'
    """
    s_key = None
    if isinstance(common_key, str):
        common_key = common_key.replace(' ', '')
    else:
        common_key = str(int(common_key))
    selection_in_table = common_key  # выборка в таблице
    if common_key == '-1':
        rastr_table = 'vetv'
        s_key_digit = -1
    else:
        rastr_table = ''  # имя таблицы
        # проверка наличия явного указания таблицы
        match = re.search(re.compile(r'\((.+?)\)'), common_key)
        s_key_digit = 0
        if match:  # таблица указана
            rastr_table = match[1]
            selection_in_table = selection_in_table.split('(', 1)[0]

        if selection_in_table:
            # разделение ключей для распознания
            key_comma = selection_in_table.split(',')  # нр для ветви [ , , ], прочее []
            key_equally = selection_in_table.split('=')  # есть = [, ], нет равно []
            if ',' in selection_in_table:  # vetv
                if len(key_comma) > 3:
                    raise ValueError(f'Ошибка в задании {common_key=}')
                rastr_table = 'vetv'
                if len(key_comma) == 3:
                    selection_in_table = f'ip={key_comma[0]}&iq={key_comma[1]}&np={key_comma[2]}'
                    s_key_digit = int(key_comma[0]), int(key_comma[1]), int(key_comma[2])
                    s_key = f'{key_comma[0]},{key_comma[1]},{key_comma[2]}'
                if len(key_comma) == 2:
                    selection_in_table = f'ip={key_comma[0]}&iq={key_comma[1]}&np=0'
                    s_key_digit = int(key_comma[0]), int(key_comma[1]), 0
                    s_key = f'{key_comma[0]},{key_comma[1]}'
            else:
                if selection_in_table.isdigit():
                    rastr_table = 'node'
                    s_key_digit = int(selection_in_table)
                    selection_in_table = 'ny=' + selection_in_table
                elif 'g' == key_equally[0]:
                    rastr_table = 'Generator'
                    s_key_digit = int(key_equally[1])
                    selection_in_table = 'Num=' + key_equally[1]
                elif len(key_equally) == 2:
                    s_key_digit = int(key_equally[1])
                    if not rastr_table:
                        if key_equally[0] in KEY_TABLES:
                            rastr_table = KEY_TABLES[key_equally[0]]  # вернет имя таблицы
                elif len(key_equally) > 2:  # 'ip = 1&iq = 2&np = 0'
                    rastr_table = 'vetv'
                    ip = int(key_equally[1].split('&')[0])
                    iq = int(key_equally[2].split('&')[0])
                    np_ = 0 if len(key_equally) == 3 else int(key_equally[3])
                    s_key_digit = ip, iq, np_  # if np_ else (ip, iq)
        if not rastr_table:
            raise ValueError(f'Таблица не определена: {common_key=}')
    if not s_key:
        s_key = s_key_digit

    if back == 's_key_digit':
        return s_key_digit
    elif back == 'sel':
        return selection_in_table
    elif back == 'tab':
        return rastr_table
    elif back == 'tab s_key_digit':
        return rastr_table, s_key_digit
    elif back == 'tab sel':
        return rastr_table, selection_in_table
    return rastr_table, selection_in_table, s_key_digit, s_key


def create_table(sheet, sheet_name, point_start: str = 'A1'):
    """
    Создать объект таблица из всего диапазона листа (openpyxl).
    :param sheet: Объект лист excel
    :param sheet_name: Имя таблицы.
    :param point_start:
    """
    tab = Table(displayName=sheet_name,
                ref=f'{point_start}:' + get_column_letter(sheet.max_column) + str(sheet.max_row))

    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9',
                                        showFirstColumn=False,
                                        showLastColumn=False,
                                        showRowStripes=True,
                                        showColumnStripes=True)
    sheet.add_table(tab)


def from_list1_only_exists_in_list2(list1: list, list2) -> list:
    """
    Вернуть лист с элементами list1 которые есть в list2 (в той же очередности)
    :param list1: Итерируемый объект не строка
    :param list2: Итерируемый объект не строка
    :return: Отфильтрованный лист list1
    """
    if isinstance(list1, str) or isinstance(list2, str):
        raise TypeError
    list1_filter = []
    for x in list1:
        if x in list2:
            list1_filter.append(x)
    return list1_filter


def save_to_sqlite(path_db: str,
                   dict_df: dict):

    con = sqlite3.connect(path_db)
    for key in dict_df:
        dict_df[key].to_sql(key,
                            con,
                            if_exists='replace')
    con.commit()
    con.close()


if __name__ == '__main__':
    # print(str_yeas_in_list('2021...2025'))
    print(split_task_action('[1,2,0:sta=1],[2,3:sta=0]{5,7:sta==1},[9,8:sta=1],6'))
